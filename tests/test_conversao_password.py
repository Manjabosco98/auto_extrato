import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch, MagicMock

import pandas as pd
from googleapiclient.errors import HttpError
from httplib2 import Response
from openpyxl import Workbook, load_workbook

from src.app.supabase import supabase_api
from src.services import conversao


class FakeDrive:
    def __init__(self, temp_dir, root_pdfs=None, password_pdfs=None, existing_folders=None):
        self.temp_dir = Path(temp_dir)
        self.root_pdfs = root_pdfs if root_pdfs is not None else [
            {
                "id": "pdf-1",
                "name": "0126_EXTBAN_TESTE_CLIENTE_AG 1_CC 1.pdf",
                "mimeType": "application/pdf",
            }
        ]
        self.password_pdfs = password_pdfs if password_pdfs is not None else []
        self.moved = []
        self.renamed = []
        self.uploaded = []
        self.updated = []
        self.trashed = set()
        self.history_file = None
        self.history_rows = []
        self.folders = []
        self.pdfs_calls = []
        self._existing_folders = existing_folders or set()

    def get_or_create_folder(self, folder_id_pai, name_folder):
        self.folders.append((folder_id_pai, name_folder))
        return {"id": f"id-{name_folder}", "name": name_folder}

    def list_folder_by_name(self, folder_id, name_folder):
        for pai, nome in self.folders:
            if pai == folder_id and nome == name_folder:
                return {"id": f"id-{name_folder}", "name": name_folder}
        if (folder_id, name_folder) in self._existing_folders:
            return {"id": f"id-{name_folder}", "name": name_folder}
        return None

    def get_file_info(self, file_id):
        raise AssertionError("get_file_info nao deve ser chamado para resolver SRVARQ > EMP")

    def find_folder_by_name(self, name_folder, shared_with_me=False):
        if name_folder == "SRVARQ":
            return {"id": "id-SRVARQ", "name": "SRVARQ"}

        return None

    def pdfs(self, folder_id, pdf_type=None):
        self.pdfs_calls.append((folder_id, pdf_type))

        if folder_id == "id-PDFS_COM_SENHAS":
            return [
                arquivo
                for arquivo in self.password_pdfs
                if arquivo["id"] not in self.trashed
            ]

        if folder_id != "id-EXT":
            return []

        return [
            arquivo
            for arquivo in self.root_pdfs
            if arquivo["id"] not in self.trashed
        ]

    def download(self, file_id, destino_local):
        destino = Path(destino_local)

        if self.history_file and file_id == self.history_file["id"]:
            destino.write_bytes(self.history_file["content"])
            return str(destino)

        destino.write_bytes(b"%PDF-1.4")
        return str(destino)

    def upload(self, caminho_local, folder_id_destino, type_file, name_drive=None):
        upload = {
            "id": f"upload-{len(self.uploaded) + 1}",
            "name": name_drive or Path(caminho_local).name,
            "folder_id_destino": folder_id_destino,
            "type_file": type_file,
        }
        self.uploaded.append(upload)
        self._capture_history_rows(caminho_local, upload["name"])
        return upload

    def find_file_by_name(self, folder_id, name, mime_type=None):
        if self.history_file and name == self.history_file["name"]:
            return {
                "id": self.history_file["id"],
                "name": self.history_file["name"],
                "mimeType": mime_type,
            }

        return None

    def update_file(self, file_id, caminho_local, type_file, name_drive=None):
        update = {
            "id": file_id,
            "name": name_drive or Path(caminho_local).name,
            "type_file": type_file,
        }
        self.updated.append(update)
        self._capture_history_rows(caminho_local, update["name"])
        return update

    def move_file(self, file_id, folder_id_destino):
        self.moved.append((file_id, folder_id_destino))

    def rename_file(self, file_id, new_name):
        self.renamed.append((file_id, new_name))
        return {"id": file_id, "name": new_name}

    def list_children(self, folder_id):
        if folder_id != "id-PDFS_COM_SENHAS":
            return []

        return [
            arquivo
            for arquivo in self.password_pdfs
            if arquivo["id"] not in self.trashed
        ]

    def trash_file(self, file_id):
        self.trashed.add(file_id)
        return {"id": file_id, "trashed": True}

    def _capture_history_rows(self, caminho_local, name):
        if name != conversao.HISTORICO_CONVERSOES:
            return

        workbook = load_workbook(caminho_local)
        worksheet = workbook.active
        self.history_rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]


class ConversaoPasswordTest(unittest.TestCase):
    def test_extrai_senha_e_nome_limpo(self):
        resultado = conversao.extrair_senha_nome_pdf(
            "0526_EXTBAN C6BANK SENHA 639256_GRB.pdf"
        )

        self.assertEqual(resultado, ("639256", "0526_EXTBAN C6BANK_GRB.pdf"))

    def test_extrai_senha_preservando_sm_no_nome_limpo(self):
        resultado = conversao.extrair_senha_nome_pdf(
            "0526_EXTBAN C6BANK SENHA 663861 SM_METRIKAL.pdf"
        )

        self.assertEqual(resultado, ("663861", "0526_EXTBAN C6BANK SM_METRIKAL.pdf"))

    def test_retorna_none_quando_nome_nao_tem_padrao_senha(self):
        resultado = conversao.extrair_senha_nome_pdf("0526_EXTBAN C6BANK_GRB.pdf")

        self.assertIsNone(resultado)

    def test_extrai_cliente_do_nome_do_arquivo(self):
        resultado = conversao.extrair_cliente_nome_arquivo(
            "0526_EXTBAN ITAU_CAMARGOS.pdf"
        )

        self.assertEqual(resultado, "CAMARGOS")

    def test_extrai_cliente_com_agencia(self):
        resultado = conversao.extrair_cliente_nome_arquivo(
            "0126_EXTBAN_ITAU_ACAO_AG 2903_CC 99019-6.pdf"
        )

        self.assertEqual(resultado, "ACAO")

    def test_extrai_cliente_com_agencia_e_sm(self):
        resultado = conversao.extrair_cliente_nome_arquivo(
            "0526_EXTBAN_NUBANK_SM_AMP ENG_AG XXX_CC 99287663-4.pdf"
        )

        self.assertEqual(resultado, "AMP ENG")

    def test_extrai_cliente_com_agencia_e_senha(self):
        resultado = conversao.extrair_cliente_nome_arquivo(
            "0526_EXTBAN_NUBANK_SENHA XXXX_AMP ENG_AG XXX_CC 99287663-4.pdf"
        )

        self.assertEqual(resultado, "AMP ENG")

    def test_extrai_cliente_com_agencia_e_senha_sm(self):
        resultado = conversao.extrair_cliente_nome_arquivo(
            "0526_EXTBAN_NUBANK_SENHA XXXX_SM_AMP ENG_AG XXX_CC 99287663-4.pdf"
        )

        self.assertEqual(resultado, "AMP ENG")

    def test_extrai_banco_do_nome_do_pdf(self):
        resultado = conversao.extrair_banco_nome_arquivo(
            "0526_EXTBAN SICOOB_SILVA.pdf"
        )

        self.assertEqual(resultado, "SICOOB")

    def test_extrai_banco_removendo_conta_do_nome_lancamento(self):
        resultado = conversao.extrair_banco_nome_arquivo(
            "0526_LANCBAN ITAU 98313-2_SILVA.xlsm"
        )

        self.assertEqual(resultado, "ITAU")

    def test_extrai_periodo_do_nome_do_arquivo(self):
        resultado = conversao.extrair_periodo_nome_arquivo(
            "0526_EXTBAN ITAU_CAMARGOS.pdf"
        )

        self.assertEqual(resultado, ("05", "26"))

    def test_formata_mensagem_google_chat(self):
        mensagem = conversao.formatar_mensagem_google_chat(
            ["0626_EXTBAN SICOO_ALE", "0626_EXTBAN SICOO_LF"],
            momento=conversao.datetime(2026, 6, 15, 12, 31, 0),
        )

        self.assertEqual(
            mensagem,
            (
                "Extratos salvos 15/06/26 as 12:31 e atualizado na Base de Dados:\n\n"
                "0626_EXTBAN SICOO_ALE\n"
                "0626_EXTBAN SICOO_LF"
            ),
        )

    def test_formata_mensagem_google_chat_com_convertidos_e_sem_movimentacao(self):
        mensagem = conversao.formatar_mensagem_google_chat(
            ["0626_EXTBAN SICOO_ALE"],
            pdfs_sem_movimentacao=[
                "0526_EXTBAN ITAU 45440-7 SM_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT",
                "0526_EXTBAN NORMAL_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT",
            ],
            momento=conversao.datetime(2026, 6, 17, 10, 15, 0),
        )

        self.assertEqual(
            mensagem,
            (
                "Extratos salvos 17/06/26 as 10:15 e atualizado na Base de Dados:\n\n"
                "0626_EXTBAN SICOO_ALE\n\n"
                "PDFs sem movimentacao salvos na pasta da empresa:\n\n"
                "0526_EXTBAN ITAU 45440-7 SM_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT\n"
                "0526_EXTBAN NORMAL_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT"
            ),
        )

    def test_formata_mensagem_google_chat_somente_sem_movimentacao(self):
        mensagem = conversao.formatar_mensagem_google_chat(
            [],
            pdfs_sem_movimentacao=[
                "0526_EXTBAN ITAU 45440-7 SM_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT",
            ],
            momento=conversao.datetime(2026, 6, 17, 10, 15, 0),
        )

        self.assertEqual(
            mensagem,
            (
                "PDFs sem movimentacao salvos na pasta da empresa:\n\n"
                "0526_EXTBAN ITAU 45440-7 SM_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT"
            ),
        )

    def test_formata_mensagem_google_chat_somente_nao_legiveis(self):
        mensagem = conversao.formatar_mensagem_google_chat(
            [],
            pdfs_nao_legiveis=[
                "[NAO LEGIVEL] - 0526_EXTBAN IMAGEM_CIAL.pdf",
            ],
            momento=conversao.datetime(2026, 6, 17, 10, 15, 0),
        )

        self.assertEqual(
            mensagem,
            (
                "PDFs nao legiveis movidos para 00_INVALIDOS:\n\n"
                "[NAO LEGIVEL] - 0526_EXTBAN IMAGEM_CIAL.pdf"
            ),
        )

    def test_envia_notificacao_google_chat_com_payload_correto(self):
        class FakeResponse:
            status_code = 200
            text = "ok"

        with patch.object(conversao.requests, "post", return_value=FakeResponse()) as post:
            resultado = conversao.enviar_notificacao_google_chat(
                ["0626_EXTBAN SICOO_ALE"],
                momento=conversao.datetime(2026, 6, 15, 12, 31, 0),
            )

        self.assertTrue(resultado)
        post.assert_called_once_with(
            conversao.GOOGLE_CHAT_SEND_URL,
            json={
                "space_name": "spaces/AAQAQEHQc-k",
                "message": (
                    "Extratos salvos 15/06/26 as 12:31 e atualizado na Base de Dados:\n\n"
                    "0626_EXTBAN SICOO_ALE"
                ),
            },
            timeout=30,
        )

    def test_envia_notificacao_google_chat_somente_sem_movimentacao(self):
        class FakeResponse:
            status_code = 200
            text = "ok"

        with patch.object(conversao.requests, "post", return_value=FakeResponse()) as post:
            resultado = conversao.enviar_notificacao_google_chat(
                [],
                pdfs_sem_movimentacao=[
                    "0526_EXTBAN ITAU 45440-7 SM_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT",
                ],
                momento=conversao.datetime(2026, 6, 17, 10, 15, 0),
            )

        self.assertTrue(resultado)
        post.assert_called_once_with(
            conversao.GOOGLE_CHAT_SEND_URL,
            json={
                "space_name": "spaces/AAQAQEHQc-k",
                "message": (
                    "PDFs sem movimentacao salvos na pasta da empresa:\n\n"
                    "0526_EXTBAN ITAU 45440-7 SM_CIAL.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT"
                ),
            },
            timeout=30,
        )

    def test_resposta_nao_2xx_google_chat_retorna_false(self):
        class FakeResponse:
            status_code = 500
            text = "erro interno"

        with patch.object(conversao.requests, "post", return_value=FakeResponse()):
            resultado = conversao.enviar_notificacao_google_chat(
                ["0626_EXTBAN SICOO_ALE"],
                momento=conversao.datetime(2026, 6, 15, 12, 31, 0),
            )

        self.assertFalse(resultado)

    def test_falha_notificacao_google_chat_nao_quebra_fluxo(self):
        with patch.object(conversao.requests, "post", side_effect=RuntimeError("offline")):
            resultado = conversao.enviar_notificacao_google_chat(
                ["0626_EXTBAN SICOO_ALE"],
                momento=conversao.datetime(2026, 6, 15, 12, 31, 0),
            )

        self.assertFalse(resultado)

    def test_notificacao_google_chat_ignora_lista_vazia(self):
        with (
            patch.object(conversao.requests, "post") as post,
            self.assertLogs(conversao.logger, level="INFO") as logs,
        ):
            resultado = conversao.enviar_notificacao_google_chat([])

        self.assertFalse(resultado)
        post.assert_not_called()
        self.assertIn(
            "Notificacao Google Chat nao enviada: nenhum resultado para informar",
            "\n".join(logs.output),
        )

    def test_nome_pasta_empresa_usa_id_e_razao_social(self):
        resultado = conversao.nome_pasta_empresa(23, " Camargos ")

        self.assertEqual(resultado, "23_CAMARGOS")

    def test_resolve_srvarq_por_nome_quando_id_configurado_nao_abre(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(temp_dir)
            chamadas = []

            def get_or_create_folder(folder_id_pai, name_folder):
                chamadas.append((folder_id_pai, name_folder))

                if folder_id_pai == "srvarq-inacessivel":
                    response = Response({"status": "404", "reason": "Not Found"})
                    raise HttpError(response, b"File not found")

                return {"id": f"id-{name_folder}", "name": name_folder}

            fake_drive.get_or_create_folder = get_or_create_folder

            resultado = conversao.resolver_pasta_emp_base(
                google_drive=fake_drive,
                pasta_base_id="srvarq-inacessivel",
            )

        self.assertEqual(resultado, "id-EMP")
        self.assertEqual(
            chamadas,
            [
                ("srvarq-inacessivel", "EMP"),
                ("id-SRVARQ", "EMP"),
            ],
        )

    def test_carrega_empresa_ativa_por_razao_social(self):
        empresas = {"CAMARGOS": ("437", "CAMARGOS")}
        with patch.object(conversao, "carregar_empresas_ativas", return_value=empresas):
            resultado = conversao.buscar_empresa_por_cliente(
                " camargos ",
                empresas=empresas,
            )

        self.assertEqual(resultado, ("437", "CAMARGOS"))

    def test_cliente_inexistente_retorna_campos_vazios(self):
        resultado = conversao.buscar_empresa_por_cliente(
            "NAO EXISTE",
            empresas={},
        )

        self.assertEqual(resultado, ("", ""))

    def test_registrar_historico_cria_planilha_na_raiz(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(temp_dir)

            with patch.object(
                conversao,
                "buscar_empresa_por_cliente",
                return_value=(437, "CAMARGOS"),
            ):
                conversao.registrar_historico_conversao(
                    google_drive=fake_drive,
                    pasta_raiz_id="root-folder",
                    temp_dir=Path(temp_dir),
                    nomes_arquivos=[
                        "0526_EXTBAN C6BANK_CAMARGOS.pdf",
                        "0526_EXTBAN C6BANK_CAMARGOS.xlsx",
                        "0526_LANCBAN C6BANK_CAMARGOS.xlsm",
                    ],
                    pasta_destino="EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    data_hora=conversao.datetime(2026, 6, 12, 9, 30, 0),
                    data_hora_movimento=conversao.datetime(2026, 6, 12, 9, 31, 5),
                )

        self.assertEqual(fake_drive.updated, [])
        self.assertEqual(fake_drive.uploaded[-1]["name"], conversao.HISTORICO_CONVERSOES)
        self.assertEqual(fake_drive.uploaded[-1]["folder_id_destino"], "root-folder")
        self.assertEqual(
            fake_drive.history_rows,
            [
                conversao.HISTORICO_HEADERS,
                (
                    437,
                    "CAMARGOS",
                    "2026-06-12",
                    "09:30:00",
                    "0526_EXTBAN C6BANK_CAMARGOS.pdf",
                    "EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    "2026-06-12 09:31:05",
                    "C6BANK",
                ),
                (
                    437,
                    "CAMARGOS",
                    "2026-06-12",
                    "09:30:00",
                    "0526_EXTBAN C6BANK_CAMARGOS.xlsx",
                    "EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    "2026-06-12 09:31:05",
                    "C6BANK",
                ),
                (
                    437,
                    "CAMARGOS",
                    "2026-06-12",
                    "09:30:00",
                    "0526_LANCBAN C6BANK_CAMARGOS.xlsm",
                    "EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    "2026-06-12 09:31:05",
                    "C6BANK",
                ),
            ],
        )

    def test_registrar_historico_atualiza_planilha_existente(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_dir_path = Path(temp_dir)
            historico_existente = temp_dir_path / conversao.HISTORICO_CONVERSOES
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(conversao.HISTORICO_HEADERS_ANTIGO)
            worksheet.append(
                [
                    "0426_EXTBAN C6BANK_LF.pdf",
                    "2026-06-11 08:00:00",
                    "00_CONVERTIDOS/0426_EXTBAN C6BANK_LF",
                ]
            )
            workbook.save(historico_existente)

            fake_drive = FakeDrive(temp_dir)
            fake_drive.history_file = {
                "id": "history-id",
                "name": conversao.HISTORICO_CONVERSOES,
                "content": historico_existente.read_bytes(),
            }

            with patch.object(
                conversao,
                "buscar_empresa_por_cliente",
                return_value=(437, "CAMARGOS"),
            ):
                conversao.registrar_historico_conversao(
                    google_drive=fake_drive,
                    pasta_raiz_id="root-folder",
                    temp_dir=temp_dir_path,
                    nomes_arquivos=[
                        "0526_EXTBAN C6BANK_CAMARGOS.pdf",
                        "0526_EXTBAN C6BANK_CAMARGOS.xlsx",
                        "0526_LANCBAN C6BANK_CAMARGOS.xlsm",
                    ],
                    pasta_destino="EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    data_hora=conversao.datetime(2026, 6, 12, 9, 30, 0),
                    data_hora_movimento=conversao.datetime(2026, 6, 12, 9, 31, 5),
                )

        self.assertEqual(fake_drive.updated[-1]["id"], "history-id")
        self.assertEqual(
            fake_drive.history_rows,
            [
                conversao.HISTORICO_HEADERS,
                (
                    None,
                    None,
                    "2026-06-11",
                    "08:00:00",
                    "0426_EXTBAN C6BANK_LF.pdf",
                    "00_CONVERTIDOS/0426_EXTBAN C6BANK_LF",
                    None,
                    "C6BANK",
                ),
                (
                    437,
                    "CAMARGOS",
                    "2026-06-12",
                    "09:30:00",
                    "0526_EXTBAN C6BANK_CAMARGOS.pdf",
                    "EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    "2026-06-12 09:31:05",
                    "C6BANK",
                ),
                (
                    437,
                    "CAMARGOS",
                    "2026-06-12",
                    "09:30:00",
                    "0526_EXTBAN C6BANK_CAMARGOS.xlsx",
                    "EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    "2026-06-12 09:31:05",
                    "C6BANK",
                ),
                (
                    437,
                    "CAMARGOS",
                    "2026-06-12",
                    "09:30:00",
                    "0526_LANCBAN C6BANK_CAMARGOS.xlsm",
                    "EMP/437_CAMARGOS/MOV/CONT/26/05/EXT",
                    "2026-06-12 09:31:05",
                    "C6BANK",
                ),
            ],
        )

    def test_historico_existente_com_banco_vazio_preenche_pelo_nome(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(conversao.HISTORICO_HEADERS)
        worksheet.append(
            [
                22,
                "SILVA",
                "2026-06-13",
                "11:45:06",
                "0526_LANCBAN ITAU 98313-2_SILVA.xlsm",
                "EMP/22_SILVA/MOV/CONT/26/05/EXT",
                "2026-06-13 11:45:12",
                "",
            ]
        )

        conversao.migrar_historico_antigo(worksheet)

        self.assertEqual(worksheet.cell(row=2, column=8).value, "ITAU")

    def test_move_pdf_com_senha_sem_interromper_conversao(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[{
                    "id": "pdf-1",
                    "name": "0526_EXTBAN_C6BANK_SENHA 639256_SM_GRB_AG 1_CC 123.pdf",
                    "mimeType": "application/pdf",
                }],
                existing_folders={("id-EMP", "23_GRB")},
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"GRB": (23, "GRB")}),
                patch.object(conversao, "pdf_possui_senha", return_value=True),
                patch.object(
                    conversao,
                    "remover_senha_pdf",
                    side_effect=lambda caminho_pdf, senha, caminho_saida: Path(caminho_saida).write_bytes(
                        b"%PDF-1.4 sem senha"
                    ),
                ) as remover_senha,
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-23"),
                patch.object(conversao, "baixar_controle_supabase", return_value=(True, None)),
                patch.object(conversao, "enviar_notificacao_google_chat"),
            ):
                conversao.executar_conversao()

        remover_senha.assert_called_once()
        self.assertEqual(fake_drive.updated[0]["id"], "pdf-1")
        self.assertEqual(
            fake_drive.updated[0]["name"],
            "0526_EXTBAN_C6BANK_SM_GRB_AG 1_CC 123.pdf",
        )
        self.assertIn(("pdf-1", "id-EXT"), fake_drive.moved)
        self.assertIn((conversao.GOOGLE_DRIVE_FOLDER_ID, "EXT"), fake_drive.folders)
        self.assertNotIn((conversao.GOOGLE_DRIVE_FOLDER_ID, "PDFS_COM_SENHAS"), fake_drive.folders)
        self.assertIn(("id-EXT", conversao.PDF_MIME_TYPE), fake_drive.pdfs_calls)
        pdf_extractor.assert_not_called()

    def test_pdf_convertido_registra_historico_apos_mover_para_convertidos(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-1",
                        "name": "0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf",
                        "mimeType": "application/pdf",
                    }
                ],
                existing_folders={("id-EMP", "23_CAMARGOS")},
            )
            eventos = []

            def copy_modelo(origem, destino):
                Path(destino).write_bytes(b"modelo")

            def move_file(file_id, folder_id_destino):
                fake_drive.moved.append((file_id, folder_id_destino))
                eventos.append(f"move:{folder_id_destino}")

            def registrar_historico(**kwargs):
                eventos.append("historico")
                self.assertEqual(
                    kwargs["nomes_arquivos"],
                    [
                        "0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf",
                        "0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.xlsx",
                        "0526_LANCBAN_C6BANK_CAMARGOS_AG 1_CC 123.xlsm",
                    ],
                )
                self.assertEqual(
                    kwargs["pasta_destino"],
                    "EMP/23_CAMARGOS/MOV/CONT/26/05/EXT",
                )
                self.assertEqual(kwargs["empresa_id"], 23)
                self.assertEqual(kwargs["empresa_nome"], "CAMARGOS")

            def enviar_notificacao(nomes_extratos, pdfs_sem_movimentacao=None, pdfs_nao_legiveis=None, atualizados_sge=None, **kwargs):
                eventos.append("chat")
                self.assertEqual(nomes_extratos, ["0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123"])
                self.assertEqual(pdfs_sem_movimentacao, [])
                self.assertEqual(pdfs_nao_legiveis, [])
                return True

            fake_drive.move_file = move_file

            df = pd.DataFrame(
                [{"DATA": "01/05/2026", "VALOR": 10.0, "TIPO": "C", "DESCRIÇÃO": "PIX"}]
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"CAMARGOS": (23, "CAMARGOS")}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "dispatch", return_value=df),
                patch.object(conversao, "planilha_lancamento"),
                patch.object(conversao.shutil, "copy2", side_effect=copy_modelo),
                patch.object(conversao, "registrar_historico_conversao", side_effect=registrar_historico),
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-23"),
                patch.object(conversao, "enviar_notificacao_google_chat", side_effect=enviar_notificacao),
            ):
                pdf_extractor.return_value.extract.return_value = (["texto extraido"], 1)
                conversao.executar_conversao()

        self.assertEqual(
            eventos,
            [
                "move:id-EXT",
                "historico",
                "chat",
            ],
        )
        self.assertIn((conversao.GOOGLE_DRIVE_FOLDER_ID, "EXT"), fake_drive.folders)
        self.assertIn((conversao.GOOGLE_DRIVE_FOLDER_ID, "00_INVALIDOS"), fake_drive.folders)
        self.assertNotIn((conversao.GOOGLE_DRIVE_FOLDER_ID, "PDFS_COM_SENHAS"), fake_drive.folders)
        self.assertIn(("id-EXT", conversao.PDF_MIME_TYPE), fake_drive.pdfs_calls)
        self.assertIn((conversao.GOOGLE_DRIVE_EMP_FOLDER_ID, "EMP"), fake_drive.folders)
        self.assertIn(("id-MOV", "CONT"), fake_drive.folders)

    def test_interpreta_quatro_formatos_canonicos(self):
        casos = [
            ("0526_EXTBAN_NUBANK_AMP ENG_AG XXX_CC 99287663-4.pdf", None, False),
            ("0526_EXTBAN_NUBANK_SM_AMP ENG_AG XXX_CC 99287663-4.pdf", None, True),
            ("0526_EXTBAN_NUBANK_SENHA XXXX_AMP ENG_AG XXX_CC 99287663-4.pdf", "XXXX", False),
            ("0526_EXTBAN_NUBANK_SENHA XXXX_SM_AMP ENG_AG XXX_CC 99287663-4.pdf", "XXXX", True),
        ]

        for nome, senha, sem_movimentacao in casos:
            with self.subTest(nome=nome):
                dados = conversao.interpretar_nome_extrato(nome)
                self.assertEqual(dados.competencia, "05-2026")
                self.assertEqual(dados.banco, "NUBANK")
                self.assertEqual(dados.empresa, "AMP ENG")
                self.assertEqual(dados.senha, senha)
                self.assertEqual(dados.sem_movimentacao, sem_movimentacao)

    def test_banco_canonico_e_extraido_de_pdf_xlsx_e_xlsm(self):
        self.assertEqual(
            conversao.extrair_banco_nome_arquivo(
                "0526_EXTBAN_NUBANK_AMP ENG_AG XXX_CC 99287663-4.pdf"
            ),
            "NUBANK",
        )
        self.assertEqual(
            conversao.extrair_banco_nome_arquivo(
                "0526_EXTBAN_NUBANK_AMP ENG_AG XXX_CC 99287663-4.xlsx"
            ),
            "NUBANK",
        )
        self.assertEqual(
            conversao.extrair_banco_nome_arquivo(
                "0526_LANCBAN_NUBANK_AMP ENG_AG XXX_CC 99287663-4.xlsm"
            ),
            "NUBANK",
        )

    def test_sm_so_e_reconhecido_como_segmento_exato(self):
        self.assertFalse(
            conversao.nome_indica_sem_movimentacao(
                "0526_EXTBAN_NUBANK_SMART EMP_AG XXX_CC 99287663-4.pdf"
            )
        )

    def test_pdf_sem_criptografia_com_senha_no_nome_apenas_renomeia(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[{
                    "id": "pdf-1",
                    "name": "0526_EXTBAN_NUBANK_SENHA XXXX_SM_AMP ENG_AG XXX_CC 99287663-4.pdf",
                    "mimeType": "application/pdf",
                }],
                existing_folders={("id-EMP", "23_AMP ENG")},
            )
            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"AMP ENG": (23, "AMP ENG")}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-23"),
                patch.object(conversao, "baixar_controle_supabase", return_value=(True, None)),
                patch.object(conversao, "enviar_notificacao_google_chat"),
            ):
                conversao.executar_conversao()

        self.assertEqual(
            fake_drive.renamed[0],
            ("pdf-1", "0526_EXTBAN_NUBANK_SM_AMP ENG_AG XXX_CC 99287663-4.pdf"),
        )
        self.assertIn(("pdf-1", "id-EXT"), fake_drive.moved)

    def test_senha_invalida_move_pdf_sanitizado_para_invalidos(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[{
                    "id": "pdf-1",
                    "name": "0526_EXTBAN_NUBANK_SENHA 000000_AMP ENG_AG XXX_CC 99287663-4.pdf",
                    "mimeType": "application/pdf",
                }],
            )
            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "pdf_possui_senha", return_value=True),
                patch.object(conversao, "remover_senha_pdf", side_effect=ValueError("Senha invalida")),
                patch.object(conversao, "enviar_notificacao_google_chat"),
                patch.object(conversao, "carregar_empresas_ativas", return_value={}),
            ):
                conversao.executar_conversao()

        self.assertEqual(
            fake_drive.renamed[0],
            ("pdf-1", "[ERRO SENHA] - 0526_EXTBAN_NUBANK_AMP ENG_AG XXX_CC 99287663-4.pdf"),
        )
        self.assertIn(("pdf-1", "id-00_INVALIDOS"), fake_drive.moved)

    def test_nome_invalido_move_para_invalidos(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[{
                    "id": "pdf-1",
                    "name": "0526_EXTBAN_NUBANK_EMPRESA.pdf",
                    "mimeType": "application/pdf",
                }],
            )
            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "enviar_notificacao_google_chat") as mock_chat,
                patch.object(conversao, "carregar_empresas_ativas", return_value={}),
            ):
                conversao.executar_conversao()

        self.assertEqual(fake_drive.renamed, [])
        self.assertEqual(fake_drive.moved, [])
        chat_kwargs = mock_chat.call_args.kwargs
        nomes_invalidos = chat_kwargs.get("nomes_invalidos", [])
        self.assertTrue(any("0526_EXTBAN_NUBANK_EMPRESA.pdf" in n for n in nomes_invalidos))

    def test_erro_em_um_pdf_nao_interrompe_o_proximo(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-erro",
                        "name": "0526_EXTBAN_NUBANK_EMPRESA_AG 1_CC 1.pdf",
                        "mimeType": "application/pdf",
                    },
                    {
                        "id": "pdf-ok",
                        "name": "0526_EXTBAN_NUBANK_EMPRESA_AG 1_CC 2.pdf",
                        "mimeType": "application/pdf",
                    },
                ],
                existing_folders={("id-EMP", "23_EMPRESA")},
            )

            def copy_modelo(origem, destino):
                Path(destino).write_bytes(b"modelo")

            df = pd.DataFrame(
                [{"DATA": "01/05/2026", "VALOR": 10.0, "TIPO": "C", "DESCRIÇÃO": "PIX"}]
            )
            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"EMPRESA": (23, "EMPRESA")}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "dispatch", return_value=df),
                patch.object(conversao, "planilha_lancamento"),
                patch.object(conversao.shutil, "copy2", side_effect=copy_modelo),
                patch.object(conversao, "registrar_historico_conversao"),
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-23"),
                patch.object(conversao, "baixar_controle_supabase", return_value=(True, None)),
                patch.object(conversao, "enviar_notificacao_google_chat"),
            ):
                pdf_extractor.return_value.extract.side_effect = [
                    RuntimeError("PDF quebrado"),
                    (["texto extraido"], 1),
                ]
                conversao.executar_conversao()

        self.assertIn(("pdf-ok", "id-EXT"), fake_drive.moved)

    def test_nome_com_sm_move_para_pasta_empresa_sem_renomear(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-sm",
                        "name": "0526_EXTBAN_ITAU_SM_CIAL_AG 1_CC 45440-7.pdf",
                        "mimeType": "application/pdf",
                    }
                ],
                existing_folders={("id-EMP", "123_CIAL")},
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"CIAL": (123, "CIAL")}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-123"),
                patch.object(conversao, "baixar_controle_supabase", return_value=(True, None)) as mock_baixa,
                patch.object(conversao, "enviar_notificacao_google_chat") as notificacao,
            ):
                conversao.executar_conversao()

        self.assertEqual(fake_drive.renamed, [])
        self.assertIn(("pdf-sm", "id-EXT"), fake_drive.moved)
        self.assertEqual(fake_drive.history_rows[0], conversao.HISTORICO_HEADERS)
        self.assertEqual(
            fake_drive.history_rows[1][4],
            "0526_EXTBAN_ITAU_SM_CIAL_AG 1_CC 45440-7.pdf",
        )
        self.assertEqual(fake_drive.history_rows[1][5], "EMP/123_CIAL/MOV/CONT/26/05/EXT")
        notificacao.assert_called_once()
        self.assertEqual(
            notificacao.call_args.kwargs["pdfs_sem_movimentacao"],
            ["0526_EXTBAN_ITAU_SM_CIAL_AG 1_CC 45440-7.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT"],
        )
        self.assertEqual(notificacao.call_args.kwargs["status_execucao"], "SUCESSO")
        pdf_extractor.assert_not_called()
        mock_baixa.assert_called_once()
        self.assertEqual(mock_baixa.call_args[1]["status"], "Não Aplicável")

    def test_sem_movimentacao_sem_empresa_permanece_na_ext(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-sm",
                        "name": "0526_EXTBAN_ITAU_SM_CIAL_AG 1_CC 45440-7.pdf",
                        "mimeType": "application/pdf",
                    }
                ],
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "enviar_notificacao_google_chat") as notificacao,
            ):
                conversao.executar_conversao()

        self.assertEqual(fake_drive.renamed, [])
        self.assertEqual(fake_drive.moved, [])
        self.assertEqual(fake_drive.history_rows, [])
        notificacao.assert_called_once()
        self.assertEqual(
            notificacao.call_args.kwargs["erros_processamento"],
            ["0526_EXTBAN_ITAU_SM_CIAL_AG 1_CC 45440-7.pdf - empresa ou pasta nao encontrada; mantido na EXT"],
        )
        self.assertEqual(
            notificacao.call_args.kwargs["status_execucao"],
            "SUCESSO COM ALERTAS",
        )
        pdf_extractor.assert_not_called()

    def test_pdf_sem_texto_move_para_invalidos_com_prefixo_sem_imagem(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-imagem",
                        "name": "0526_EXTBAN_IMAGEM_CIAL_AG 1_CC 1.pdf",
                        "mimeType": "application/pdf",
                    }
                ],
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "dispatch") as dispatch,
                patch.object(conversao, "enviar_notificacao_google_chat") as notificacao,
                patch.object(conversao, "carregar_empresas_ativas", return_value={}),
            ):
                pdf_extractor.return_value.extract.return_value = ([], 0)
                conversao.executar_conversao()

        self.assertEqual(
            fake_drive.renamed,
            [
                (
                    "pdf-imagem",
                    "[NAO LEGIVEL] - 0526_EXTBAN_IMAGEM_CIAL_AG 1_CC 1.pdf",
                )
            ],
        )
        self.assertIn(("pdf-imagem", "id-00_INVALIDOS"), fake_drive.moved)
        self.assertEqual(fake_drive.history_rows, [])
        notificacao.assert_called_once()
        self.assertEqual(
            notificacao.call_args.kwargs["pdfs_nao_legiveis"],
            ["[NAO LEGIVEL] - 0526_EXTBAN_IMAGEM_CIAL_AG 1_CC 1.pdf"],
        )
        dispatch.assert_not_called()

    def test_dataframe_vazio_move_para_pasta_empresa_sem_renomear(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-vazio",
                        "name": "0526_EXTBAN_NORMAL_CIAL_AG 1_CC 1.pdf",
                        "mimeType": "application/pdf",
                    }
                ],
                existing_folders={("id-EMP", "123_CIAL")},
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"CIAL": (123, "CIAL")}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "dispatch", return_value=pd.DataFrame()),
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-123"),
                patch.object(conversao, "baixar_controle_supabase", return_value=(True, None)) as mock_baixa,
                patch.object(conversao, "enviar_notificacao_google_chat") as notificacao,
            ):
                pdf_extractor.return_value.extract.return_value = (["texto extraido"], 1)
                conversao.executar_conversao()

        self.assertEqual(fake_drive.renamed, [])
        self.assertIn(("pdf-vazio", "id-EXT"), fake_drive.moved)
        self.assertEqual(fake_drive.history_rows[0], conversao.HISTORICO_HEADERS)
        self.assertEqual(
            fake_drive.history_rows[1][4],
            "0526_EXTBAN_NORMAL_CIAL_AG 1_CC 1.pdf",
        )
        self.assertEqual(fake_drive.history_rows[1][5], "EMP/123_CIAL/MOV/CONT/26/05/EXT")
        notificacao.assert_called_once()
        self.assertEqual(
            notificacao.call_args.kwargs["pdfs_sem_movimentacao"],
            ["0526_EXTBAN_NORMAL_CIAL_AG 1_CC 1.pdf - EMP/123_CIAL/MOV/CONT/26/05/EXT"],
        )
        mock_baixa.assert_called_once()
        self.assertEqual(mock_baixa.call_args[1]["status"], "Não Aplicável")

    def test_extrai_competencia_do_nome_do_arquivo(self):
        resultado = conversao.extrair_competencia_nome_arquivo(
            "0526_EXTBAN_NUBANK_AMP ENG_99287663-4.pdf"
        )
        self.assertEqual(resultado, "05-2026")

    def test_extrai_competencia_dezembro(self):
        resultado = conversao.extrair_competencia_nome_arquivo(
            "1225_EXTBAN_ITAU_GRB_12345-6.pdf"
        )
        self.assertEqual(resultado, "12-2025")

    def test_extrai_competencia_erro_nome_invalido(self):
        with self.assertRaises(ValueError):
            conversao.extrair_competencia_nome_arquivo("EXTBAN_NUBANK_AMP.pdf")

    def test_extrai_codigo_documento_extban(self):
        resultado = conversao.extrair_codigo_documento(
            "0526_EXTBAN_NUBANK_AMP ENG_99287663-4.pdf"
        )
        self.assertEqual(resultado, "EXTBAN")

    def test_extrai_codigo_documento_curto(self):
        resultado = conversao.extrair_codigo_documento("0526_EXTBAN_.pdf")
        self.assertEqual(resultado, "EXTBAN")

    def test_extrai_codigo_documento_maiusculo(self):
        resultado = conversao.extrair_codigo_documento(
            "0526_extban_nubank_amp.pdf"
        )
        self.assertEqual(resultado, "EXTBAN")

    def test_formatar_mensagem_com_sge(self):
        mensagem = conversao.formatar_mensagem_google_chat(
            ["0526_EXTBAN_NUBANK_AMP ENG_99287663-4"],
            atualizados_sge=2,
            momento=conversao.datetime(2026, 6, 19, 14, 30, 0),
        )

        self.assertIn("Baixa dada no portal SGE: 2 documento(s) atualizado(s)", mensagem)
        self.assertIn("Extratos salvos 19/06/26 as 14:30", mensagem)

    def test_formatar_mensagem_sge_zero_nao_aparece(self):
        mensagem = conversao.formatar_mensagem_google_chat(
            ["0526_EXTBAN_NUBANK_AMP ENG_99287663-4"],
            atualizados_sge=0,
            momento=conversao.datetime(2026, 6, 19, 14, 30, 0),
        )

        self.assertNotIn("SGE", mensagem)

    def test_formatar_mensagem_sge_none_nao_aparece(self):
        mensagem = conversao.formatar_mensagem_google_chat(
            ["0526_EXTBAN_NUBANK_AMP ENG_99287663-4"],
            momento=conversao.datetime(2026, 6, 19, 14, 30, 0),
        )

        self.assertNotIn("SGE", mensagem)

    def test_enviar_notificacao_passa_sge(self):
        class FakeResponse:
            status_code = 200
            text = "ok"

        with patch.object(conversao.requests, "post", return_value=FakeResponse()) as post:
            resultado = conversao.enviar_notificacao_google_chat(
                ["0526_EXTBAN_NUBANK_AMP ENG_99287663-4"],
                atualizados_sge=3,
                momento=conversao.datetime(2026, 6, 19, 14, 30, 0),
            )

        self.assertTrue(resultado)
        call_kwargs = post.call_args
        mensagem_enviada = call_kwargs[1]["json"]["message"]
        self.assertIn("Baixa dada no portal SGE: 3 documento(s) atualizado(s)", mensagem_enviada)

    def test_fluxo_completo_chama_supabase(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-1",
                        "name": "0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf",
                        "mimeType": "application/pdf",
                    }
                ],
                existing_folders={("id-EMP", "23_CAMARGOS")},
            )

            def copy_modelo(origem, destino):
                Path(destino).write_bytes(b"modelo")

            df = pd.DataFrame(
                [{"DATA": "01/05/2026", "VALOR": 10.0, "TIPO": "C", "DESCRIÇÃO": "PIX"}]
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"CAMARGOS": (23, "CAMARGOS")}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "dispatch", return_value=df),
                patch.object(conversao, "planilha_lancamento"),
                patch.object(conversao.shutil, "copy2", side_effect=copy_modelo),
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-23"),
                patch.object(conversao, "baixar_controle_supabase", return_value=(True, None)) as mock_sge,
                patch.object(conversao, "enviar_notificacao_google_chat") as notificacao,
            ):
                pdf_extractor.return_value.extract.return_value = (["texto extraido"], 1)
                conversao.executar_conversao()

        mock_sge.assert_called_once()
        call_kwargs = mock_sge.call_args[1]
        self.assertEqual(call_kwargs["empresa_codigo"], "23")
        self.assertEqual(call_kwargs["competencia"], "05-2026")
        self.assertEqual(call_kwargs["codigo_documento"], "EXTBAN")
        self.assertEqual(call_kwargs["banco"], "C6BANK")
        self.assertEqual(call_kwargs["agencia"], "1")
        self.assertEqual(call_kwargs["conta"], "123")
        self.assertEqual(call_kwargs["quantidade_arquivos"], 3)
        self.assertEqual(call_kwargs["nome_arquivo"], "0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf")
        self.assertIn("Google Drive /", call_kwargs["local_arquivo"])
        self.assertEqual(call_kwargs["status"], "Enviado")

        notificacao.assert_called_once()
        notificacao_kwargs = notificacao.call_args[1]
        self.assertEqual(notificacao_kwargs["atualizados_sge"], 1)

    def test_falha_sge_nao_quebra_fluxo(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDrive(
                temp_dir,
                root_pdfs=[
                    {
                        "id": "pdf-1",
                        "name": "0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf",
                        "mimeType": "application/pdf",
                    }
                ],
                existing_folders={("id-EMP", "23_CAMARGOS")},
            )

            def copy_modelo(origem, destino):
                Path(destino).write_bytes(b"modelo")

            df = pd.DataFrame(
                [{"DATA": "01/05/2026", "VALOR": 10.0, "TIPO": "C", "DESCRIÇÃO": "PIX"}]
            )

            with (
                patch.object(conversao, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(conversao, "carregar_empresas_ativas", return_value={"CAMARGOS": (23, "CAMARGOS")}),
                patch.object(conversao, "pdf_possui_senha", return_value=False),
                patch.object(conversao, "PDFExtractor") as pdf_extractor,
                patch.object(conversao, "dispatch", return_value=df),
                patch.object(conversao, "planilha_lancamento"),
                patch.object(conversao.shutil, "copy2", side_effect=copy_modelo),
                patch.object(conversao, "buscar_id_empresa_supabase", return_value="uuid-empresa-23"),
                patch.object(conversao, "baixar_controle_supabase", return_value=(False, None)),
                patch.object(conversao, "enviar_notificacao_google_chat") as notificacao,
            ):
                pdf_extractor.return_value.extract.return_value = (["texto extraido"], 1)
                conversao.executar_conversao()

        notificacao.assert_called_once()
        notificacao_kwargs = notificacao.call_args[1]
        self.assertEqual(notificacao_kwargs["atualizados_sge"], 0)

    def test_buscar_id_empresa_retorna_uuid(self):
        fake_response = MagicMock()
        fake_response.status_code = 200
        fake_response.json.return_value = {
            "data": [{"id_empresa": "3f6b2660-b83b-4da4-ae4f-20cd8df4e6de"}],
        }

        with patch.object(supabase_api.requests, "get", return_value=fake_response):
            resultado = supabase_api.buscar_id_empresa_supabase(
                empresa_codigo="428",
                competencia="02-2026",
                codigo_documento="EXTBAN",
            )

        self.assertEqual(resultado, "3f6b2660-b83b-4da4-ae4f-20cd8df4e6de")

    def test_buscar_id_empresa_retorna_none_sem_registro(self):
        fake_response = MagicMock()
        fake_response.status_code = 200
        fake_response.json.return_value = {"data": []}

        with patch.object(supabase_api.requests, "get", return_value=fake_response):
            resultado = supabase_api.buscar_id_empresa_supabase(
                empresa_codigo="999",
                competencia="01-2026",
                codigo_documento="EXTBAN",
            )

        self.assertIsNone(resultado)

    def test_buscar_id_empresa_retorna_none_em_erro_http(self):
        fake_response = MagicMock()
        fake_response.status_code = 500
        fake_response.text = "erro interno"

        with patch.object(supabase_api.requests, "get", return_value=fake_response):
            resultado = supabase_api.buscar_id_empresa_supabase(
                empresa_codigo="428",
                competencia="02-2026",
                codigo_documento="EXTBAN",
            )

        self.assertIsNone(resultado)

    def test_buscar_id_empresa_retorna_none_em_excecao(self):
        with patch.object(
            supabase_api.requests, "get", side_effect=RuntimeError("offline")
        ):
            resultado = supabase_api.buscar_id_empresa_supabase(
                empresa_codigo="428",
                competencia="02-2026",
                codigo_documento="EXTBAN",
            )

        self.assertIsNone(resultado)

    def test_atualizar_controle_busca_uuid_antes_do_put(self):
        fake_get = MagicMock()
        fake_get.status_code = 200
        fake_get.json.return_value = {
            "data": [{"id_empresa": "uuid-empresa-123"}],
        }

        fake_put = MagicMock()
        fake_put.status_code = 200
        fake_put.text = "ok"

        with (
            patch.object(supabase_api.requests, "get", return_value=fake_get),
            patch.object(supabase_api.requests, "put", return_value=fake_put) as mock_put,
        ):
            resultado = supabase_api.atualizar_controle_supabase(
                empresa_codigo="428",
                competencia="02-2026",
                codigo_documento="EXTBAN",
                data_recebimento="2026-06-19",
                quantidade_arquivos=3,
                nome_arquivo="0226_EXTBAN_INTER_ACAO.pdf",
                local_arquivo="Google Drive / EMP/428_ACAO/MOV/CONT/26/02/EXT",
            )

        self.assertTrue(resultado)
        put_kwargs = mock_put.call_args[1]["json"]
        self.assertEqual(put_kwargs["id_empresa"], "uuid-empresa-123")
        self.assertEqual(put_kwargs["empresa_codigo"], "428")

    def test_atualizar_controle_falha_se_uuid_nao_encontrado(self):
        fake_get = MagicMock()
        fake_get.status_code = 200
        fake_get.json.return_value = {"data": []}

        with (
            patch.object(supabase_api.requests, "get", return_value=fake_get),
            patch.object(supabase_api.requests, "put") as mock_put,
        ):
            resultado = supabase_api.atualizar_controle_supabase(
                empresa_codigo="999",
                competencia="01-2026",
                codigo_documento="EXTBAN",
                data_recebimento="2026-06-19",
                quantidade_arquivos=3,
                nome_arquivo="0126_EXTBAN_XXX.pdf",
                local_arquivo="Google Drive / XXX",
            )

        self.assertFalse(resultado)
        mock_put.assert_not_called()

    def test_extrai_senha_com_underscore_antes_de_senha(self):
        resultado = conversao.extrair_senha_nome_pdf(
            "0526_EXTBAN_C6BANK_SENHA 663861_ SM_METRIKAL_AG 1_CC 421987570.pdf"
        )

        self.assertIsNotNone(resultado)
        senha, nome_limpo = resultado
        self.assertEqual(senha, "663861")
        self.assertEqual(nome_limpo, "0526_EXTBAN_C6BANK_SM_METRIKAL_AG 1_CC 421987570.pdf")

    def test_extrai_senha_com_underscore_e_barra_baixo(self):
        resultado = conversao.extrair_senha_nome_pdf(
            "0526_EXTBAN_C6BANK_SENHA 123456_GRB.pdf"
        )

        self.assertIsNotNone(resultado)
        senha, nome_limpo = resultado
        self.assertEqual(senha, "123456")

    def test_resolver_pasta_destino_emp_erro_pasta_cliente_nao_existe(self):
        fake_drive = FakeDrive(
            "dummy",
            existing_folders=set(),
        )

        with self.assertRaises(ValueError) as ctx:
            conversao.resolver_pasta_destino_emp(
                google_drive=fake_drive,
                pasta_emp_id="id-EMP",
                arquivo_nome="0526_EXTBAN C6BANK_CAMARGOS.pdf",
                empresa_id=23,
                empresa_nome="CAMARGOS",
            )

        self.assertIn("23_CAMARGOS", str(ctx.exception))

    def test_resolver_pasta_destino_emp_funciona_pasta_existente(self):
        fake_drive = FakeDrive(
            "dummy",
            existing_folders={("id-EMP", "23_CAMARGOS")},
        )

        pasta_id, caminho = conversao.resolver_pasta_destino_emp(
            google_drive=fake_drive,
            pasta_emp_id="id-EMP",
            arquivo_nome="0526_EXTBAN C6BANK_CAMARGOS.pdf",
            empresa_id=23,
            empresa_nome="CAMARGOS",
        )

        self.assertEqual(pasta_id, "id-EXT")
        self.assertEqual(caminho, "EMP/23_CAMARGOS/MOV/CONT/26/05/EXT")

    def test_preparacao_rejeita_modelo_sem_plan1(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            modelo_path = Path(temp_dir) / "Lancamentos_Contabeis.xlsm"

            modelo = Workbook()
            modelo.active.title = "OutraAba"
            modelo.save(modelo_path)

            with self.assertRaisesRegex(ValueError, "Plan1"):
                conversao.validar_preparacao_fluxo(modelo_path)

    def test_preparacao_falha_antes_de_autenticar_no_drive(self):
        with (
            patch.object(conversao, "validar_preparacao_fluxo", side_effect=ValueError("modelo invalido")),
            patch.object(conversao, "GoogleDriveAuth") as google_drive,
            patch.object(conversao, "enviar_notificacao_google_chat") as notificacao,
        ):
            with self.assertRaisesRegex(ValueError, "modelo invalido"):
                conversao.executar_conversao()

        google_drive.assert_called_once()
        notificacao.assert_called_once()
        self.assertEqual(notificacao.call_args.kwargs["tipo"], "FALHA")
        self.assertEqual(notificacao.call_args.kwargs["status_execucao"], "FALHA")

    def test_baixar_controle_envia_post_para_url_correta(self):
        fake_post = MagicMock()
        fake_post.status_code = 200
        fake_post.text = "ok"

        with patch.object(supabase_api.requests, "post", return_value=fake_post) as mock_post:
            sucesso, detalhe = supabase_api.baixar_controle_supabase(
                empresa_codigo="23",
                codigo_documento="EXTBAN",
                competencia="05-2026",
                banco="C6BANK",
                agencia="1",
                conta="123",
                nome_arquivo="0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf",
                local_arquivo="Google Drive / EMP/23_CAMARGOS/MOV/CONT/26/05/EXT",
                quantidade_arquivos=3,
            )

        self.assertTrue(sucesso)
        self.assertIsNone(detalhe)
        mock_post.assert_called_once()
        call_args = mock_post.call_args
        self.assertEqual(call_args[0][0], supabase_api.SUPABASE_BAIXA_URL)
        self.assertEqual(call_args[1]["json"]["empresa_codigo"], "23")
        self.assertEqual(call_args[1]["json"]["banco"], "C6BANK")
        self.assertEqual(call_args[1]["json"]["agencia"], "1")
        self.assertEqual(call_args[1]["json"]["conta"], "123")
        self.assertEqual(call_args[1]["json"]["competencia"], "2026-05")
        self.assertEqual(call_args[1]["json"]["codigo_documento"], "EXTBAN")
        self.assertEqual(call_args[1]["json"]["quantidade_arquivos"], 3)
        self.assertEqual(call_args[1]["json"]["status_envio"], "Enviado")
        self.assertEqual(call_args[1]["json"]["data_recebimento"], date.today().isoformat())
        self.assertNotIn("id_empresa", call_args[1]["json"])

    def test_baixar_controle_retorna_false_em_erro_http(self):
        fake_response = MagicMock()
        fake_response.status_code = 500
        fake_response.text = "erro interno"

        with patch.object(supabase_api.requests, "post", return_value=fake_response):
            sucesso, detalhe = supabase_api.baixar_controle_supabase(
                empresa_codigo="23",
                codigo_documento="EXTBAN",
                competencia="05-2026",
                banco="C6BANK",
                agencia="1",
                conta="123",
                nome_arquivo="0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf",
                local_arquivo="Google Drive / EMP/23_CAMARGOS/MOV/CONT/26/05/EXT",
                quantidade_arquivos=3,
            )

        self.assertFalse(sucesso)
        self.assertIsNone(detalhe)

    def test_baixar_controle_retorna_false_em_excecao(self):
        with patch.object(
            supabase_api.requests, "post", side_effect=RuntimeError("offline")
        ):
            sucesso, detalhe = supabase_api.baixar_controle_supabase(
                empresa_codigo="23",
                codigo_documento="EXTBAN",
                competencia="05-2026",
                banco="C6BANK",
                agencia="1",
                conta="123",
                nome_arquivo="0526_EXTBAN_C6BANK_CAMARGOS_AG 1_CC 123.pdf",
                local_arquivo="Google Drive / EMP/23_CAMARGOS/MOV/CONT/26/05/EXT",
                quantidade_arquivos=3,
            )

        self.assertFalse(sucesso)
        self.assertIsNone(detalhe)

    def test_baixar_controle_envia_status_nao_aplicavel(self):
        fake_post = MagicMock()
        fake_post.status_code = 200
        fake_post.text = "ok"

        with patch.object(supabase_api.requests, "post", return_value=fake_post) as mock_post:
            sucesso, detalhe = supabase_api.baixar_controle_supabase(
                empresa_codigo="47",
                codigo_documento="EXTBAN",
                competencia="05-2026",
                banco="ITAU",
                agencia="2903",
                conta="97630-2",
                nome_arquivo="0526_EXTBAN_ITAU_SM_SEVEN_AG 2903_CC 97630-2.pdf",
                local_arquivo="Google Drive / EMP/47_SEVEN/MOV/CONT/26/05/EXT",
                quantidade_arquivos=1,
                status="Não Aplicável",
            )

        self.assertTrue(sucesso)
        payload = mock_post.call_args[1]["json"]
        self.assertEqual(payload["status_envio"], "Não Aplicável")
        self.assertEqual(payload["empresa_codigo"], "47")
        self.assertEqual(payload["data_recebimento"], date.today().isoformat())

    def test_baixar_controle_retorna_false_empresa_nao_cadastrada(self):
        fake_response = MagicMock()
        fake_response.status_code = 404
        fake_response.text = '{"error":{"message":"Empresa nao encontrada."}}'

        with patch.object(supabase_api.requests, "post", return_value=fake_response):
            sucesso, detalhe = supabase_api.baixar_controle_supabase(
                empresa_codigo="449",
                codigo_documento="EXTBAN",
                competencia="05-2026",
                banco="C6BANK",
                agencia="1",
                conta="421987570",
                nome_arquivo="0526_EXTBAN_C6BANK_SM_METRIKAL_AG 1_CC 421987570.pdf",
                local_arquivo="Google Drive / EMP/449_METRIKAL/MOV/CONT/26/05/EXT",
                quantidade_arquivos=1,
                status="Não Aplicável",
            )

        self.assertFalse(sucesso)
        self.assertIn("449", detalhe)
        self.assertIn("nao esta cadastrada", detalhe)


if __name__ == "__main__":
    unittest.main()
