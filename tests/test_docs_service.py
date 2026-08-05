import tempfile
import unittest
import unittest.mock
from pathlib import Path
from unittest.mock import patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.api.endpoints import docs as docs_endpoint
from src.app.supabase import supabase_api
from src.services import docs


# Mapa padrao codigo -> destino_final usado nos testes (antes vinha da
# planilha DOCS E CAMINHO.xlsx; hoje vem do cadastro de Documentos do SGE).
CAMINHOS_PADRAO = {"JURATPAS": "SRVARQ/EMP/{EMPRESA}/MOV/CONT/{ANO}/{MES}/REL"}


class FakeDriveDocs:
    def __init__(self, temp_dir, existing_folders=None, children=None, folder_children=None):
        self.temp_dir = Path(temp_dir)
        self.children = children if children is not None else [
            {
                "id": "doc-1",
                "name": "0626_JURATPAS_BRITO.docx",
                "mimeType": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            }
        ]
        self.folder_children = folder_children or {}
        self.folders = []
        self.moved = []
        self.uploaded = []
        self.updated = []
        self.history_rows = []
        self.history_file = None
        self._existing_folders = existing_folders or set()

    def get_or_create_folder(self, folder_id_pai, name_folder):
        self.folders.append((folder_id_pai, name_folder))
        return {"id": f"id-{name_folder}", "name": name_folder}

    def list_folder_by_name(self, folder_id, name_folder):
        for pai, nome in self.folders:
            if pai == folder_id and nome == name_folder:
                return {"id": f"id-{name_folder}", "name": name_folder}
        if (folder_id, name_folder) in self._existing_folders:
            return {"id": f"id-{name_folder}", "name": name_folder}
        return None

    def find_file_by_name(self, folder_id, name, mime_type=None):
        if self.history_file and name == self.history_file["name"]:
            return {
                "id": self.history_file["id"],
                "name": self.history_file["name"],
                "mimeType": mime_type,
            }

        return None

    def download(self, file_id, destino_local):
        destino = Path(destino_local)

        if self.history_file and file_id == self.history_file["id"]:
            destino.write_bytes(self.history_file["content"])
            return str(destino)

        raise AssertionError(f"download inesperado: {file_id}")

    def list_children(self, folder_id):
        if folder_id == "id-DOCS":
            return self.children
        if folder_id in self.folder_children:
            return self.folder_children[folder_id]

        return []

    def move_file(self, file_id, folder_id_destino):
        self.moved.append((file_id, folder_id_destino))

    def upload(self, caminho_local, folder_id_destino, type_file, name_drive=None):
        upload = {
            "id": f"upload-{len(self.uploaded) + 1}",
            "name": name_drive or Path(caminho_local).name,
            "folder_id_destino": folder_id_destino,
            "type_file": type_file,
        }
        self.uploaded.append(upload)
        self._capture_history(caminho_local, upload["name"])
        return upload

    def update_file(self, file_id, caminho_local, type_file, name_drive=None):
        update = {
            "id": file_id,
            "name": name_drive or Path(caminho_local).name,
            "type_file": type_file,
        }
        self.updated.append(update)
        self._capture_history(caminho_local, update["name"])
        return update

    def _capture_history(self, caminho_local, name):
        if name != docs.HISTORICO_DOCS:
            return

        workbook = load_workbook(caminho_local)
        worksheet = workbook.active
        self.history_rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]


class DocsServiceTest(unittest.TestCase):
    def test_parse_nome_documento(self):
        resultado = docs.parse_nome_documento("0626_JURATPAS_BRITO.docx")

        self.assertEqual(
            resultado,
            {
                "mes": "06",
                "ano": "26",
                "codigo": "JURATPAS",
                "cliente": "BRITO",
                "banco": "",
                "agencia": "",
                "conta": "",
            },
        )

    def test_monta_destino_docs(self):
        resultado = docs.montar_destino_docs(
            destino_template="SRVARQ/EMP/{EMPRESA}/MOV/CONT/{ANO}/{MÊS}/REL",
            empresa_chave="196_BRITO",
            ano="26",
            mes="06",
        )

        self.assertEqual(
            resultado,
            ("EMP", ["196_BRITO", "MOV", "CONT", "26", "06", "REL"]),
        )

    def test_monta_destino_docs_remove_prefixo_emp(self):
        resultado = docs.montar_destino_docs(
            destino_template="EMP/{EMPRESA}/MOV/CONT/{ANO}/{MES}/REL",
            empresa_chave="196_BRITO",
            ano="26",
            mes="06",
        )

        self.assertEqual(
            resultado,
            ("EMP", ["196_BRITO", "MOV", "CONT", "26", "06", "REL"]),
        )

    def test_monta_destino_docs_aceita_caminho_iniciando_na_empresa(self):
        resultado = docs.montar_destino_docs(
            destino_template="{EMPRESA}/MOV/CONT/{ANO}/{MES}/REL",
            empresa_chave="196_BRITO",
            ano="26",
            mes="06",
        )

        self.assertEqual(
            resultado,
            ("EMP", ["196_BRITO", "MOV", "CONT", "26", "06", "REL"]),
        )

    def test_monta_destino_docs_raiz_publico(self):
        # destino UNC vindo do cadastro de Documentos do SGE (raiz PUBLICO)
        resultado = docs.montar_destino_docs(
            destino_template="\\\\SRVARQ\\PUBLICO\\{EMPRESA}\\MOV\\CONT\\{ANO}\\{MES}\\REL",
            empresa_chave="196_BRITO",
            ano="26",
            mes="06",
        )

        self.assertEqual(
            resultado,
            ("PUBLICO", ["196_BRITO", "MOV", "CONT", "26", "06", "REL"]),
        )

    def test_monta_destino_docs_alias_pub_vira_publico(self):
        # grafia legada da planilha: SRVARQ\PUB -> pasta real PUBLICO
        resultado = docs.montar_destino_docs(
            destino_template="SRVARQ/PUB/{EMPRESA}/MOV/CONT/{ANO}/{MES}/EST",
            empresa_chave="196_BRITO",
            ano="26",
            mes="06",
        )

        self.assertEqual(
            resultado,
            ("PUBLICO", ["196_BRITO", "MOV", "CONT", "26", "06", "EST"]),
        )

    def test_carrega_caminhos_documentos_api(self):
        resposta = unittest.mock.Mock()
        resposta.status_code = 200
        resposta.json.return_value = {
            "count": 3,
            "data": [
                {
                    "codigo_documento": "JURATPAS",
                    "destino_final": "SRVARQ\\EMP\\{EMPRESA}\\MOV\\CONT\\{ANO}\\{MES}\\REL",
                    "ativo": True,
                },
                {
                    "codigo_documento": "ANTIGO",
                    "destino_final": "SRVARQ\\EMP\\{EMPRESA}\\X",
                    "ativo": False,
                },
                {"codigo_documento": "SEMDESTINO", "destino_final": "", "ativo": True},
            ],
        }

        with patch.object(supabase_api.requests, "get", return_value=resposta):
            resultado = supabase_api.carregar_caminhos_documentos_api()

        # apenas ativos com destino_final entram no mapa
        self.assertEqual(
            resultado,
            {"JURATPAS": "SRVARQ\\EMP\\{EMPRESA}\\MOV\\CONT\\{ANO}\\{MES}\\REL"},
        )

    def test_carrega_caminhos_documentos_api_falha_levanta_erro(self):
        resposta = unittest.mock.Mock()
        resposta.status_code = 500
        resposta.text = "erro interno"

        with patch.object(supabase_api.requests, "get", return_value=resposta):
            with self.assertRaises(RuntimeError):
                supabase_api.carregar_caminhos_documentos_api()

    def test_executar_docs_move_arquivo_registra_historico_e_notifica(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(
                temp_dir,
                existing_folders={("id-EMP", "196_BRITO")},
            )

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(docs, "carregar_caminhos_documentos_api", return_value=CAMINHOS_PADRAO),
                patch.object(docs, "carregar_empresas_ativas", return_value={"BRITO": (196, "BRITO")}),
                patch.object(docs, "enviar_notificacao_docs_google_chat") as notificacao,
                patch.object(docs, "consultar_situacao_controle", return_value=supabase_api.SituacaoControle(cadastrado=True, id_empresa="uuid-196")),
                patch.object(docs, "baixar_controle_supabase", return_value=(True, None)),
                patch.object(docs, "agora_historico", return_value=docs.datetime(2026, 6, 16, 12, 30, 5)),
            ):
                resultado = docs.executar_docs()

        self.assertEqual(resultado, {
            "processados": 1,
            "movidos": 1,
            "ignorados": 0,
            "erros": 0,
            "pendencias": 0,
            "erros_sge": 0,
            "baixas_preservadas": 0,
        })
        self.assertIn(("root-folder", "DOCS"), fake_drive.folders)
        self.assertIn(("srvarq-folder", "EMP"), fake_drive.folders)
        self.assertEqual(
            fake_drive.folders[-5:],
            [
                ("id-196_BRITO", "MOV"),
                ("id-MOV", "CONT"),
                ("id-CONT", "26"),
                ("id-26", "06"),
                ("id-06", "REL"),
            ],
        )
        self.assertEqual(fake_drive.moved, [("doc-1", "id-REL")])
        self.assertEqual(fake_drive.uploaded[-1]["name"], docs.HISTORICO_DOCS)
        self.assertEqual(
            fake_drive.history_rows,
            [
                docs.HISTORICO_DOCS_HEADERS,
                (
                    "196",
                    "BRITO",
                    "2026-06-16",
                    "12:30:05",
                    "JURATPAS",
                    "0626_JURATPAS_BRITO.docx",
                    "EMP/196_BRITO/MOV/CONT/26/06/REL",
                    "2026-06-16 12:30:05",
                    "Baixado",
                ),
            ],
        )
        notificacao.assert_called_once()
        call_args, call_kwargs = notificacao.call_args
        self.assertEqual(
            call_args[0],
            ["0626_JURATPAS_BRITO.docx - EMP/196_BRITO/MOV/CONT/26/06/REL"],
        )
        self.assertEqual(call_kwargs["pendencias"], [])
        self.assertEqual(call_kwargs["erros_sge"], 0)
        self.assertEqual(call_kwargs["erros_sge_detalhe"], [])
        self.assertIs(call_kwargs["google_drive"], fake_drive)
        self.assertEqual(call_kwargs["pasta_raiz_id"], "root-folder")
        self.assertTrue(call_kwargs["execucao_id"].startswith("DOCS-"))

    def test_executar_docs_ignora_empresa_inexistente(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(temp_dir)

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "carregar_caminhos_documentos_api", return_value=CAMINHOS_PADRAO),
                patch.object(docs, "carregar_empresas_ativas", return_value={}),
                patch.object(docs, "enviar_notificacao_docs_google_chat") as notificacao,
            ):
                resultado = docs.executar_docs()

        self.assertEqual(resultado, {
            "processados": 1,
            "movidos": 0,
            "ignorados": 1,
            "erros": 0,
            "pendencias": 1,
            "erros_sge": 0,
            "baixas_preservadas": 0,
        })
        self.assertEqual(fake_drive.moved, [])
        self.assertEqual(fake_drive.uploaded, [])
        notificacao.assert_called_once()
        call_args, call_kwargs = notificacao.call_args
        self.assertEqual(call_args[0], [])
        self.assertEqual(len(call_kwargs["pendencias"]), 1)
        self.assertEqual(call_kwargs["erros_sge"], 0)
        self.assertEqual(call_kwargs["erros_sge_detalhe"], [])
        self.assertIs(call_kwargs["google_drive"], fake_drive)
        self.assertEqual(call_kwargs["pasta_raiz_id"], "root-folder")
        self.assertTrue(call_kwargs["execucao_id"].startswith("DOCS-"))

    def test_executar_docs_pasta_do_cliente_ausente_reporta_motivo(self):
        # Empresa cadastrada no SGE mas sem pasta ID_NOME em EMP: a pendencia
        # precisa dizer QUAL pasta falta, nao "erro ao mover/processar".
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(temp_dir)

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(docs, "carregar_caminhos_documentos_api", return_value=CAMINHOS_PADRAO),
                patch.object(docs, "carregar_empresas_ativas", return_value={"BRITO": (196, "BRITO")}),
                patch.object(docs, "enviar_notificacao_docs_google_chat") as notificacao,
                patch.object(docs, "consultar_situacao_controle") as mock_busca_sge,
                patch.object(docs, "baixar_controle_supabase") as mock_baixa,
            ):
                resultado = docs.executar_docs()

        self.assertEqual(resultado, {
            "processados": 1,
            "movidos": 0,
            "ignorados": 0,
            "erros": 1,
            "pendencias": 1,
            "erros_sge": 0,
            "baixas_preservadas": 0,
        })
        self.assertEqual(fake_drive.moved, [])
        # o destino e resolvido antes da baixa: nada e enviado ao SGE
        mock_busca_sge.assert_not_called()
        mock_baixa.assert_not_called()

        pendencia = notificacao.call_args[1]["pendencias"][0]
        self.assertIn("0626_JURATPAS_BRITO.docx", pendencia)
        self.assertIn("196_BRITO", pendencia)
        self.assertIn("nao encontrada em EMP", pendencia)
        self.assertNotIn("erro ao mover/processar", pendencia)

    def test_notificacao_docs_google_chat_payload(self):
        with patch.object(docs, "registrar_e_enviar_notificacao", return_value=True) as mock_registrar:
            resultado = docs.enviar_notificacao_docs_google_chat(
                ["0626_JURATPAS_BRITO.docx - EMP/196_BRITO/MOV/CONT/26/06/REL"],
                momento=docs.datetime(2026, 6, 16, 12, 30, 0),
                google_drive="fake-drive",
                pasta_raiz_id="root-folder",
                execucao_id="DOCS-20260616123000",
            )

        self.assertTrue(resultado)
        mock_registrar.assert_called_once()
        call_kwargs = mock_registrar.call_args[1]
        self.assertEqual(call_kwargs["google_drive"], "fake-drive")
        self.assertEqual(call_kwargs["pasta_raiz_id"], "root-folder")
        self.assertEqual(call_kwargs["execucao_id"], "DOCS-20260616123000")
        self.assertEqual(call_kwargs["tipo"], "DOCS")
        self.assertIn("Documentos salvos", call_kwargs["mensagem"])

    def test_notificacao_docs_inclui_bloco_de_pendencias(self):
        with patch.object(docs, "registrar_e_enviar_notificacao", return_value=True) as mock_registrar:
            resultado = docs.enviar_notificacao_docs_google_chat(
                [],
                pendencias=[
                    "0626_COMPSN_DISLEY.pdf - codigo 'COMPSN' nao cadastrado em Documentos no portal SGE",
                ],
                momento=docs.datetime(2026, 6, 16, 12, 30, 0),
                google_drive="fake-drive",
                pasta_raiz_id="root-folder",
                execucao_id="DOCS-20260616123000",
            )

        # envia mesmo sem documentos movidos, so com pendencias
        self.assertTrue(resultado)
        mensagem = mock_registrar.call_args[1]["mensagem"]
        self.assertIn("Documentos pendentes em DOCS (nao movidos)", mensagem)
        self.assertIn("COMPSN", mensagem)

    def test_conta_arquivos_pasta_ignora_subpastas_e_desktop_ini(self):
        fake = FakeDriveDocs.__new__(FakeDriveDocs)
        fake.folder_children = {
            "pasta-x": [
                {"id": "a", "name": "nota1.pdf", "mimeType": "application/pdf"},
                {"id": "b", "name": "nota2.pdf", "mimeType": "application/pdf"},
                {"id": "c", "name": "desktop.ini", "mimeType": "application/octet-stream"},
                {"id": "d", "name": "sub", "mimeType": docs.GoogleDriveAuth.FOLDER_MIME_TYPE},
            ]
        }
        self.assertEqual(docs._contar_arquivos_pasta(fake, "pasta-x"), 2)

    def test_executar_docs_processa_pasta_conta_arquivos_e_baixa(self):
        folder_mime = docs.GoogleDriveAuth.FOLDER_MIME_TYPE
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(
                temp_dir,
                existing_folders={("id-EMP", "156_CEMAF OPE")},
                children=[
                    {"id": "pasta-1", "name": "0626_NFSENT_CEMAF OPE", "mimeType": folder_mime}
                ],
                folder_children={
                    "pasta-1": [
                        {"id": "n1", "name": "nota1.pdf", "mimeType": "application/pdf"},
                        {"id": "n2", "name": "nota2.pdf", "mimeType": "application/pdf"},
                        {"id": "n3", "name": "nota3.pdf", "mimeType": "application/pdf"},
                    ]
                },
            )

            with (
                patch.object(docs, "GoogleDriveAuth") as mock_auth,
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(
                    docs,
                    "carregar_caminhos_documentos_api",
                    return_value={"NFSENT": "SRVARQ/EMP/{EMPRESA}/MOV/CONT/{ANO}/{MES}/REL/NFSENT"},
                ),
                patch.object(docs, "carregar_empresas_ativas", return_value={"CEMAF OPE": (156, "CEMAF OPE")}),
                patch.object(docs, "enviar_notificacao_docs_google_chat"),
                patch.object(docs, "consultar_situacao_controle", return_value=supabase_api.SituacaoControle(cadastrado=True, id_empresa="uuid-156")),
                patch.object(docs, "baixar_controle_supabase", return_value=(True, None)) as mock_baixa,
                patch.object(docs, "agora_historico", return_value=docs.datetime(2026, 6, 26, 8, 40, 0)),
            ):
                # preserva o FOLDER_MIME_TYPE real para o filtro de subpastas funcionar
                mock_auth.return_value = fake_drive
                mock_auth.FOLDER_MIME_TYPE = folder_mime
                resultado = docs.executar_docs()

        self.assertEqual(resultado["movidos"], 1)
        self.assertEqual(resultado["erros_sge"], 0)
        # a pasta inteira foi movida para o destino do codigo NFSENT
        self.assertEqual(fake_drive.moved, [("pasta-1", "id-NFSENT")])
        # baixa: codigo NFSENT, sem instancia, e Arquivos = 3 (itens dentro da pasta)
        kwargs = mock_baixa.call_args.kwargs
        self.assertEqual(kwargs["codigo_documento"], "NFSENT")
        self.assertEqual(kwargs["competencia"], "06-2026")
        self.assertEqual(kwargs["empresa_codigo"], "156")
        self.assertEqual(kwargs["quantidade_arquivos"], 3)
        self.assertEqual(kwargs["status"], "Enviado")

    def test_executar_docs_destino_raiz_publico(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(
                temp_dir,
                existing_folders={("id-PUBLICO", "196_BRITO")},
                children=[
                    {"id": "doc-1", "name": "0626_RELVEND_BRITO.pdf", "mimeType": "application/pdf"}
                ],
            )

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(
                    docs,
                    "carregar_caminhos_documentos_api",
                    return_value={"RELVEND": "\\\\SRVARQ\\PUBLICO\\{EMPRESA}\\MOV\\CONT\\{ANO}\\{MES}\\REL"},
                ),
                patch.object(docs, "carregar_empresas_ativas", return_value={"BRITO": (196, "BRITO")}),
                patch.object(docs, "enviar_notificacao_docs_google_chat"),
                patch.object(docs, "consultar_situacao_controle", return_value=supabase_api.SituacaoControle(cadastrado=True, id_empresa="uuid-196")),
                patch.object(docs, "baixar_controle_supabase", return_value=(True, None)),
                patch.object(docs, "agora_historico", return_value=docs.datetime(2026, 6, 16, 12, 30, 5)),
            ):
                resultado = docs.executar_docs()

        self.assertEqual(resultado["movidos"], 1)
        self.assertEqual(resultado["pendencias"], 0)
        # raiz PUBLICO resolvida dentro de SRVARQ e usada como base do destino
        self.assertIn(("srvarq-folder", "PUBLICO"), fake_drive.folders)
        self.assertEqual(fake_drive.moved, [("doc-1", "id-REL")])
        # historico registra o caminho a partir da raiz PUBLICO
        self.assertEqual(
            fake_drive.history_rows[-1][6],
            "PUBLICO/196_BRITO/MOV/CONT/26/06/REL",
        )

    def test_rota_docs_executar_responde_202(self):
        app = FastAPI()
        app.include_router(docs_endpoint.router, prefix="/docs")

        with patch.object(docs_endpoint, "executar_docs"):
            response = TestClient(app).post("/docs/executar")

        self.assertEqual(response.status_code, 202)
        self.assertEqual(response.json(), {"message": "Fluxo DOCS iniciado"})

    def test_rota_docs_executar_responde_409_quando_em_execucao(self):
        app = FastAPI()
        app.include_router(docs_endpoint.router, prefix="/docs")
        docs_endpoint._docs_lock.acquire()

        try:
            response = TestClient(app).post("/docs/executar")
        finally:
            docs_endpoint._docs_lock.release()

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json(), {"message": "Fluxo DOCS ja esta em execucao"})

    def test_parse_nome_documento_extban_extrai_banco_agencia_conta(self):
        resultado = docs.parse_nome_documento(
            "0526_EXTBAN_KAMINO_UP380_AG 1_CC 30513944.xls"
        )

        self.assertEqual(
            resultado,
            {
                "mes": "05",
                "ano": "26",
                "codigo": "EXTBAN",
                "cliente": "UP380",
                "banco": "KAMINO",
                "agencia": "1",
                "conta": "30513944",
            },
        )

    def test_parse_nome_documento_extban_sem_agencia_conta(self):
        resultado = docs.parse_nome_documento("0526_EXTBAN_ITAU_SMART.xlsx")

        self.assertEqual(
            resultado,
            {
                "mes": "05",
                "ano": "26",
                "codigo": "EXTBAN",
                "cliente": "SMART",
                "banco": "ITAU",
                "agencia": "",
                "conta": "",
            },
        )

    def test_parse_nome_documento_nao_extban_campos_vazios(self):
        resultado = docs.parse_nome_documento("0526_RELCONFIM_KAMINO_UP380.xlsx")

        self.assertEqual(
            resultado,
            {
                "mes": "05",
                "ano": "26",
                "codigo": "RELCONFIM",
                "cliente": "UP380",
                "banco": "",
                "agencia": "",
                "conta": "",
            },
        )

    def test_parse_aplica_alias_de_codigo(self):
        # FATCAR/EXTMAQCART/EXTBANAPL sao grafias que mapeiam para os codigos canonicos.
        self.assertEqual(docs.parse_nome_documento("0626_FATCAR_BORBA.pdf")["codigo"], "FATCART")
        self.assertEqual(
            docs.parse_nome_documento("0626_EXTMAQCART_BORBA.xlsx")["codigo"], "EXTMAQCAR"
        )
        self.assertEqual(
            docs.parse_nome_documento("0626_EXTBANAPL_ENFOCA.pdf")["codigo"], "EXTAPL"
        )
        self.assertEqual(
            docs.parse_nome_documento("0626_EXTAPL_SICOOB_SAG.pdf")["codigo"], "EXTAPL"
        )

    def test_parse_subtipo_colado_extrai_codigo_base_e_banco(self):
        resultado = docs.parse_nome_documento("0626_EXTAPL FIC GIRO_C6 BANK_GEO ENG.pdf")
        self.assertEqual(resultado["codigo"], "EXTAPL")
        self.assertEqual(resultado["banco"], "C6 BANK")
        self.assertEqual(resultado["cliente"], "GEO ENG")

    def test_executar_docs_extapl_usa_codigo_atual_da_planilha_e_sge(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(
                temp_dir,
                existing_folders={("id-EMP", "510_CASA CARLOS")},
                children=[
                    {"id": "doc-1", "name": "0626_EXTAPL_CANTINA.pdf", "mimeType": "application/pdf"},
                    {"id": "doc-2", "name": "0626_EXTBANAPL_ENFOCA.pdf", "mimeType": "application/pdf"},
                ],
            )

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(
                    docs,
                    "carregar_caminhos_documentos_api",
                    return_value={"EXTAPL": "SRVARQ/EMP/{EMPRESA}/MOV/CONT/{ANO}/{MES}/EXT"},
                ),
                patch.object(
                    docs,
                    "carregar_empresas_ativas",
                    return_value={
                        "CANTINA": (510, "CASA CARLOS"),
                        "ENFOCA": (510, "CASA CARLOS"),
                    },
                ),
                patch.object(docs, "enviar_notificacao_docs_google_chat"),
                patch.object(docs, "consultar_situacao_controle", return_value=supabase_api.SituacaoControle(cadastrado=True, id_empresa="uuid-510")),
                patch.object(docs, "baixar_controle_supabase", return_value=(True, None)) as mock_baixa,
                patch.object(docs, "agora_historico", return_value=docs.datetime(2026, 6, 26, 8, 40, 0)),
            ):
                resultado = docs.executar_docs()

        self.assertEqual(resultado["movidos"], 2)
        self.assertEqual(resultado["pendencias"], 0)
        self.assertEqual(
            [call.kwargs["codigo_documento"] for call in mock_baixa.call_args_list],
            ["EXTAPL", "EXTAPL"],
        )

    def test_parse_codigo_novo_usa_palavra_base(self):
        # "CONTEMP PRONAMP" -> codigo base "CONTEMP" (nao cadastrado -> pendencia)
        self.assertEqual(
            docs.parse_nome_documento("0626_ CONTEMP PRONAMP_DISLEY.pdf")["codigo"], "CONTEMP"
        )
        self.assertEqual(
            docs.parse_nome_documento("0626_RELCAIXA_COLEGIO WR.pdf")["codigo"], "RELCAIXA"
        )

    def test_parse_extban_sem_cliente_nao_usa_cc_como_cliente(self):
        resultado = docs.parse_nome_documento("0626_EXTBAN_CAIXA_AG 0636-8_CC 578111288-6.ofx")
        self.assertEqual(resultado["cliente"], "")
        self.assertEqual(resultado["banco"], "CAIXA")
        self.assertEqual(resultado["conta"], "578111288-6")

    def test_executar_docs_empresa_nao_cadastrada_sge(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(
                temp_dir,
                existing_folders={("id-EMP", "196_BRITO")},
            )

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(docs, "carregar_caminhos_documentos_api", return_value=CAMINHOS_PADRAO),
                patch.object(docs, "carregar_empresas_ativas", return_value={"BRITO": (196, "BRITO")}),
                patch.object(docs, "enviar_notificacao_docs_google_chat") as notificacao,
                patch.object(docs, "consultar_situacao_controle", return_value=supabase_api.SituacaoControle(cadastrado=False)),
                patch.object(docs, "baixar_controle_supabase") as mock_baixa,
                patch.object(docs, "agora_historico", return_value=docs.datetime(2026, 6, 16, 12, 30, 5)),
            ):
                resultado = docs.executar_docs()

        # controle inexistente no SGE: nem tenta a baixa e nao move o arquivo
        mock_baixa.assert_not_called()
        self.assertEqual(resultado["movidos"], 0)
        self.assertEqual(resultado["ignorados"], 1)
        self.assertEqual(resultado["pendencias"], 1)
        self.assertEqual(fake_drive.moved, [])
        self.assertEqual(fake_drive.history_rows, [])
        notificacao.assert_called_once()
        call_kwargs = notificacao.call_args[1]
        self.assertEqual(call_kwargs["erros_sge"], 0)
        self.assertIn("nao cadastrada no SGE", call_kwargs["pendencias"][0])

    def test_executar_docs_erro_api_sge(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(
                temp_dir,
                existing_folders={("id-EMP", "196_BRITO")},
            )

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(docs, "carregar_caminhos_documentos_api", return_value=CAMINHOS_PADRAO),
                patch.object(docs, "carregar_empresas_ativas", return_value={"BRITO": (196, "BRITO")}),
                patch.object(docs, "enviar_notificacao_docs_google_chat") as notificacao,
                patch.object(docs, "consultar_situacao_controle", return_value=supabase_api.SituacaoControle(cadastrado=True, id_empresa="uuid-196")),
                patch.object(docs, "baixar_controle_supabase", return_value=(False, "Empresa 196 nao cadastrada")),
                patch.object(docs, "agora_historico", return_value=docs.datetime(2026, 6, 16, 12, 30, 5)),
            ):
                resultado = docs.executar_docs()

        # baixa recusada pelo SGE: arquivo fica em DOCS e o motivo vai pro Chat
        self.assertEqual(resultado["movidos"], 0)
        self.assertEqual(resultado["ignorados"], 1)
        self.assertEqual(resultado["erros_sge"], 1)
        self.assertEqual(fake_drive.moved, [])
        self.assertEqual(fake_drive.history_rows, [])
        notificacao.assert_called_once()
        call_kwargs = notificacao.call_args[1]
        self.assertEqual(call_kwargs["erros_sge"], 1)
        self.assertEqual(call_kwargs["erros_sge_detalhe"], ["Empresa 196 nao cadastrada"])
        self.assertIn("Empresa 196 nao cadastrada", call_kwargs["pendencias"][0])

    def test_executar_docs_ja_baixado_preserva_baixa_e_move(self):
        # Documento ja baixado no SGE: a baixa nao pode ser sobrescrita, mas o
        # arquivo segue para a pasta da empresa em vez de ficar preso em DOCS.
        with tempfile.TemporaryDirectory() as temp_dir:
            fake_drive = FakeDriveDocs(
                temp_dir,
                existing_folders={("id-EMP", "196_BRITO")},
            )
            situacao = supabase_api.SituacaoControle(
                cadastrado=True,
                id_empresa="uuid-196",
                status_envio="Enviado",
                nome_arquivo="0626_JURATPAS_BRITO.docx",
                data_recebimento="2026-06-10",
            )

            with (
                patch.object(docs, "GoogleDriveAuth", return_value=fake_drive),
                patch.object(docs, "GOOGLE_DRIVE_FOLDER_ID", "root-folder"),
                patch.object(docs, "GOOGLE_DRIVE_EMP_FOLDER_ID", "srvarq-folder"),
                patch.object(docs, "carregar_caminhos_documentos_api", return_value=CAMINHOS_PADRAO),
                patch.object(docs, "carregar_empresas_ativas", return_value={"BRITO": (196, "BRITO")}),
                patch.object(docs, "enviar_notificacao_docs_google_chat") as notificacao,
                patch.object(docs, "consultar_situacao_controle", return_value=situacao),
                patch.object(docs, "baixar_controle_supabase") as mock_baixa,
                patch.object(docs, "agora_historico", return_value=docs.datetime(2026, 6, 16, 12, 30, 5)),
            ):
                resultado = docs.executar_docs()

        mock_baixa.assert_not_called()
        self.assertEqual(resultado["movidos"], 1)
        self.assertEqual(resultado["baixas_preservadas"], 1)
        self.assertEqual(resultado["erros_sge"], 0)
        self.assertEqual(fake_drive.moved, [("doc-1", "id-REL")])
        self.assertEqual(fake_drive.history_rows[1][-1], "Baixa ja existente")

        call_kwargs = notificacao.call_args[1]
        self.assertEqual(call_kwargs["pendencias"], [])
        preservadas = call_kwargs["baixas_preservadas"]
        self.assertEqual(len(preservadas), 1)
        self.assertIn("0626_JURATPAS_BRITO.docx", preservadas[0])
        self.assertIn("2026-06-10", preservadas[0])

    def test_notificacao_docs_google_chat_com_sge(self):
        with patch.object(docs, "registrar_e_enviar_notificacao", return_value=True) as mock_registrar:
            resultado = docs.enviar_notificacao_docs_google_chat(
                ["0626_JURATPAS_BRITO.docx - EMP/196_BRITO/MOV/CONT/26/06/REL"],
                momento=docs.datetime(2026, 6, 16, 12, 30, 0),
                atualizados_sge=1,
                empresas_nao_cadastradas_sge=[],
                erros_sge=0,
                erros_sge_detalhe=[],
                google_drive="fake-drive",
                pasta_raiz_id="root-folder",
                execucao_id="DOCS-20260616123000",
            )

        self.assertTrue(resultado)
        mock_registrar.assert_called_once()
        call_kwargs = mock_registrar.call_args[1]
        self.assertIn("Baixa dada no portal SGE: 1 documento(s) atualizado(s)", call_kwargs["mensagem"])


if __name__ == "__main__":
    unittest.main()
