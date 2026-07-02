"""
Validação cruzada: HISTORICO_CONVERSOES.xlsx vs Plataforma SGE (Supabase).

Baixa a planilha de histórico do Google Drive, consulta a API do SGE
para cada empresa/competência e gera um relatório Excel com as divergências.
"""

import logging
import sys
import time
from collections import defaultdict
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── setup ──────────────────────────────────────────────────────────────
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)

# ── imports do projeto ─────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.app.gdrive.google_drive_auth import GoogleDriveAuth
from src.app.gdrive.settings import (
    GOOGLE_DRIVE_FOLDER_ID,
    GOOGLE_OAUTH_CREDENTIALS,
    GOOGLE_OAUTH_TOKEN,
    GOOGLE_OAUTH_TOKEN_SECRET,
    XLSX_MIME_TYPE,
)
from src.app.supabase.supabase_api import (
    SUPABASE_CONTROLE_GET_URL,
    SUPABASE_CONTROLE_KEY,
    SUPABASE_EMPRESAS_URL,
    _converter_competencia,
    _normalizar_nome_empresa,
)
from src.services.conversao import normalizar_id_empresa


# ── constantes ─────────────────────────────────────────────────────────
HISTORICO_nome = "HISTORICO_CONVERSOES.xlsx"
SGE_PAGE_SIZE = 500


# ── helpers ────────────────────────────────────────────────────────────
def carregar_mapa_empresas() -> dict[str, str]:
    """Carrega mapeamento nome_normalizado → id_empresa do SGE."""
    headers = {"Authorization": f"Bearer {SUPABASE_CONTROLE_KEY}"}
    r = requests.get(SUPABASE_EMPRESAS_URL, headers=headers, params={"limit": 500}, timeout=30)
    r.raise_for_status()
    data = r.json().get("data", [])
    mapa = {}
    for item in data:
        nome = (item.get("nome") or "").strip().upper()
        id_emp = str(item.get("id_empresa") or "").strip()
        if nome and id_emp:
            mapa[nome] = id_emp
    logger.info("Mapa de empresas carregado: %s entradas", len(mapa))
    return mapa


# ── helpers ────────────────────────────────────────────────────────────
def competencia_de_data_hora(valor) -> str:
    """Extrai competência 'YYYY-MM' a partir de 'YYYY-MM-DD HH:MM:SS' ou date/datetime."""
    if pd.isna(valor) or valor is None:
        return ""
    texto = str(valor).strip()
    if len(texto) >= 10:
        partes = texto[:10].split("-")
        if len(partes) == 2:
            return f"{partes[0]}-{partes[1]}"
        if len(partes) == 3:
            return f"{partes[0]}-{partes[1]}"
    return ""


def baixar_historico_drive(temp_dir: Path) -> pd.DataFrame:
    """Baixa HISTORICO_CONVERSOES.xlsx do Google Drive e retorna DataFrame."""
    logger.info("Autenticando Google Drive")
    gdrive = GoogleDriveAuth(
        credentials_path=GOOGLE_OAUTH_CREDENTIALS,
        token_path=GOOGLE_OAUTH_TOKEN,
        token_secret_path=GOOGLE_OAUTH_TOKEN_SECRET,
    )

    logger.info("Procurando %s na pasta raiz", HISTORICO_nome)
    arquivo = gdrive.find_file_by_name(
        folder_id=GOOGLE_DRIVE_FOLDER_ID,
        name=HISTORICO_nome,
        mime_type=XLSX_MIME_TYPE,
    )
    if not arquivo:
        raise FileNotFoundError(f"Arquivo {HISTORICO_nome} nao encontrado no Google Drive")

    destino = temp_dir / HISTORICO_nome
    logger.info("Baixando %s (id=%s)", HISTORICO_nome, arquivo["id"])
    gdrive.download(file_id=arquivo["id"], destino_local=destino)

    logger.info("Lendo planilha")
    df = pd.read_excel(destino, engine="openpyxl")
    logger.info("Registros no histórico: %s", len(df))
    return df


def consultar_sge_empresa(
    empresa_codigo: str,
    competencia: str,
    *,
    url: str = SUPABASE_CONTROLE_GET_URL,
    key: str = SUPABASE_CONTROLE_KEY,
) -> list[dict]:
    """Consulta SGE para uma empresa+competência. Retorna lista de registros."""
    headers = {"Authorization": f"Bearer {key}"}
    params = {
        "empresa_codigo": empresa_codigo,
        "competencia": _converter_competencia(competencia),
        "cod_doc": "EXTBAN",
        "limit": SGE_PAGE_SIZE,
    }

    try:
        r = requests.get(url, headers=headers, params=params, timeout=30)
        if r.status_code < 200 or r.status_code >= 300:
            logger.warning("SGE erro %s para empresa=%s comp=%s", r.status_code, empresa_codigo, competencia)
            return []
        return r.json().get("data", [])
    except Exception:
        logger.exception("Erro ao consultar SGE: empresa=%s comp=%s", empresa_codigo, competencia)
        return []


def gerar_relatorio(df_hist: pd.DataFrame, resultados: list[dict], saida: Path) -> None:
    """Gera o relatório Excel com abas por status."""
    wb = Workbook()

    # ── Aba Resumo ─────────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    resumo_contadores = defaultdict(int)
    for r in resultados:
        resumo_contadores[r["status"]] += 1

    ws_resumo.append(["Status", "Quantidade"])
    for cell in ws_resumo[1]:
        cell.font = header_font
        cell.fill = header_fill
    for status, qtd in sorted(resumo_contadores.items(), key=lambda x: -x[1]):
        ws_resumo.append([status, qtd])
    ws_resumo.append(["Total", len(resultados)])
    ws_resumo.auto_width = True

    # ── Abas detalhadas ────────────────────────────────────────────────
    colunas = [
        "Empresa ID", "Empresa Nome", "Competência", "Arquivo Histórico",
        "Banco", "Data/Hora Movimento", "Status SGE", "Arquivo SGE",
        "Recebimento SGE", "Detalhe",
    ]
    abas = {
        "OK": [],
        "Divergências": [],
        "Não Baixados": [],
        "Não Encontrados": [],
    }

    for r in resultados:
        linha = [
            r.get("empresa_id", ""),
            r.get("empresa_nome", ""),
            r.get("competencia", ""),
            r.get("arquivo_historico", ""),
            r.get("banco", ""),
            r.get("data_hora_movimento", ""),
            r.get("status_sge", ""),
            r.get("arquivo_sge", ""),
            r.get("recebimento_sge", ""),
            r.get("detalhe", ""),
        ]
        status = r["status"]
        if status in abas:
            abas[status].append(linha)
        else:
            abas.setdefault(status, []).append(linha)

    for nome_aba, linhas in abas.items():
        ws = wb.create_sheet(title=nome_aba[:31])
        ws.append(colunas)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for linha in linhas:
            ws.append(linha)
        ws.auto_width = True

    # ── Salvar ─────────────────────────────────────────────────────────
    wb.save(saida)
    logger.info("Relatório salvo em: %s", saida)


# ── main ───────────────────────────────────────────────────────────────
def main() -> None:
    t0 = time.perf_counter()

    with TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)

        # 1. Baixar histórico
        df_hist = baixar_historico_drive(temp_path)

        # Normalizar colunas
        df_hist.columns = [str(c).strip().upper() for c in df_hist.columns]

        # Detectar coluna de empresa
        col_empresa = None
        for cand in ["EMP", "ID", "EMPRESA_ID", "EMPRESA"]:
            if cand in df_hist.columns:
                col_empresa = cand
                break
        if col_empresa is None:
            raise ValueError(f"Coluna de empresa nao encontrada. Colunas: {list(df_hist.columns)}")

        # Detectar coluna de nome arquivo
        col_arquivo = None
        for cand in ["NOME ARQUIVO", "NOME_ARQUIVO", "ARQUIVO"]:
            if cand in df_hist.columns:
                col_arquivo = cand
                break

        # Detectar coluna de data/hora movimento
        col_dt_mov = None
        for cand in ["DATA HORA MOVIMENTO", "DATA_HORA_MOVIMENTO", "DATA HORA"]:
            if cand in df_hist.columns:
                col_dt_mov = cand
                break

        # Detectar coluna de banco
        col_banco = None
        for cand in ["BANCO", "BANK"]:
            if cand in df_hist.columns:
                col_banco = cand
                break

        logger.info("Colunas detectadas: empresa=%s arquivo=%s dt_mov=%s banco=%s",
                     col_empresa, col_arquivo, col_dt_mov, col_banco)

        # 2. Carregar mapeamento empresas do SGE
        mapa_empresas = carregar_mapa_empresas()

        # 2.1 Mapear nomes do histórico para códigos SGE
        def resolver_empresa(valor) -> tuple[str, str]:
            """Retorna (id_sge, nome_original) a partir do valor do histórico."""
            nome = str(valor or "").strip().upper()
            if not nome:
                return "", ""
            # Tentar direto (se já for numérico)
            if nome.isdigit():
                return nome, nome
            # Buscar no mapa
            id_sge = mapa_empresas.get(nome)
            if id_sge:
                return id_sge, nome
            # Tentar normalizar
            nome_norm = _normalizar_nome_empresa(nome)
            id_sge = mapa_empresas.get(nome_norm)
            if id_sge:
                return id_sge, nome
            return "", nome

        # 2.2 Extrair combinações únicas (empresa_codigo, competência)
        combos = set()
        empresa_cache_nome: dict[str, str] = {}  # id_sge → nome do histórico
        for _, row in df_hist.iterrows():
            id_sge, nome_hist = resolver_empresa(row.get(col_empresa))
            dt_mov = row.get(col_dt_mov) if col_dt_mov else None
            comp = competencia_de_data_hora(dt_mov)
            if id_sge and comp:
                combos.add((id_sge, comp))
                empresa_cache_nome[id_sge] = nome_hist

        logger.info("Combinações únicas empresa+competência: %s", len(combos))

        # 3. Consultar SGE para cada combinação
        cache_sge: dict[tuple[str, str], list[dict]] = {}
        total_combos = len(combos)
        for idx, (emp_id, comp) in enumerate(sorted(combos), 1):
            if idx % 20 == 0 or idx == 1:
                logger.info("Consultando SGE: %s/%s (empresa=%s comp=%s)", idx, total_combos, emp_id, comp)
            registros = consultar_sge_empresa(emp_id, comp)
            cache_sge[(emp_id, comp)] = registros
            time.sleep(0.1)  # rate limit

        # 4. Cruzamento
        resultados = []
        for _, row in df_hist.iterrows():
            id_sge, nome_hist = resolver_empresa(row.get(col_empresa))
            dt_mov = row.get(col_dt_mov) if col_dt_mov else None
            comp = competencia_de_data_hora(dt_mov)
            arquivo_hist = str(row.get(col_arquivo, "")).strip() if col_arquivo else ""
            banco = str(row.get(col_banco, "")).strip() if col_banco else ""
            dt_mov_str = str(dt_mov).strip() if dt_mov else ""

            registros_sge = cache_sge.get((id_sge, comp), [])

            if not registros_sge:
                resultados.append({
                    "empresa_id": id_sge,
                    "empresa_nome": nome_hist,
                    "competencia": comp,
                    "arquivo_historico": arquivo_hist,
                    "banco": banco,
                    "data_hora_movimento": dt_mov_str,
                    "status": "Não Encontrados",
                    "status_sge": "",
                    "arquivo_sge": "",
                    "recebimento_sge": "",
                    "detalhe": f"Empresa '{nome_hist}' (id={id_sge})/competencia nao encontrada no SGE",
                })
                continue

            # Verificar se algum registro SGE tem baixa
            arquivo_hist_stem = Path(arquivo_hist).stem if arquivo_hist else ""
            banco_upper = banco.upper().strip()
            # Normalizar: extrair competência MMAA do nome do arquivo
            # Ex: "0526_EXTBAN_C6BANK_..." → "0526"
            partes_nome = arquivo_hist.split("_") if arquivo_hist else []
            competencia_mmaa = partes_nome[0] if partes_nome and len(partes_nome[0]) == 4 else ""
            # "0526" → banco deve conter "C6" ou "C6BANK" etc.

            # Match 1: stem exato (sem extensão) bate com stem do SGE
            melhor_match = None
            for reg in registros_sge:
                nome_sge = str(reg.get("nome_arquivo", "")).strip()
                status_sge = str(reg.get("status_envio", "")).strip()
                stem_sge = Path(nome_sge).stem if nome_sge else ""
                if status_sge == "Enviado" and stem_sge and arquivo_hist_stem and stem_sge == arquivo_hist_stem:
                    melhor_match = reg
                    break

            # Match 2: banco + competência (arquivo SGE contém nome do banco E MMAA)
            if not melhor_match and banco_upper and competencia_mmaa:
                for reg in registros_sge:
                    nome_sge = str(reg.get("nome_arquivo", "")).upper().strip()
                    status_sge = str(reg.get("status_envio", "")).strip()
                    if (status_sge == "Enviado" and nome_sge
                            and banco_upper in nome_sge
                            and competencia_mmaa in nome_sge):
                        melhor_match = reg
                        break

            # Match 3: apenas banco (sem exigir competência no nome)
            if not melhor_match and banco_upper:
                for reg in registros_sge:
                    nome_sge = str(reg.get("nome_arquivo", "")).upper().strip()
                    status_sge = str(reg.get("status_envio", "")).strip()
                    if status_sge == "Enviado" and nome_sge and banco_upper in nome_sge:
                        melhor_match = reg
                        break

            if melhor_match:
                status_sge = melhor_match.get("status_envio", "")
                arq_sge = melhor_match.get("nome_arquivo", "")
                receb_sge = melhor_match.get("data_recebimento", "")

                if status_sge == "Enviado" and arq_sge:
                    resultados.append({
                        "empresa_id": id_sge,
                        "empresa_nome": nome_hist,
                        "competencia": comp,
                        "arquivo_historico": arquivo_hist,
                        "banco": banco,
                        "data_hora_movimento": dt_mov_str,
                        "status": "OK",
                        "status_sge": status_sge,
                        "arquivo_sge": arq_sge,
                        "recebimento_sge": receb_sge,
                        "detalhe": "Conferido",
                    })
                elif status_sge == "Enviado" and not arq_sge:
                    resultados.append({
                        "empresa_id": id_sge,
                        "empresa_nome": nome_hist,
                        "competencia": comp,
                        "arquivo_historico": arquivo_hist,
                        "banco": banco,
                        "data_hora_movimento": dt_mov_str,
                        "status": "Divergências",
                        "status_sge": status_sge,
                        "arquivo_sge": arq_sge,
                        "recebimento_sge": receb_sge,
                        "detalhe": "SGE marcado como Enviado mas sem nome de arquivo",
                    })
                else:
                    resultados.append({
                        "empresa_id": id_sge,
                        "empresa_nome": nome_hist,
                        "competencia": comp,
                        "arquivo_historico": arquivo_hist,
                        "banco": banco,
                        "data_hora_movimento": dt_mov_str,
                        "status": "Não Baixados",
                        "status_sge": status_sge,
                        "arquivo_sge": arq_sge,
                        "recebimento_sge": receb_sge,
                        "detalhe": f"Status SGE: {status_sge}",
                    })
            else:
                # Tem registros no SGE mas nenhum com status Enviado
                status_sge = registros_sge[0].get("status_envio", "") if registros_sge else ""
                resultados.append({
                    "empresa_id": id_sge,
                    "empresa_nome": nome_hist,
                    "competencia": comp,
                    "arquivo_historico": arquivo_hist,
                    "banco": banco,
                    "data_hora_movimento": dt_mov_str,
                    "status": "Não Baixados",
                    "status_sge": status_sge,
                    "arquivo_sge": "",
                    "recebimento_sge": "",
                    "detalhe": f"Nenhum registro Enviado no SGE ({len(registros_sge)} registros)",
                })

        # 5. Gerar relatório
        saida = Path.cwd() / "validacao_historico_sge.xlsx"
        gerar_relatorio(df_hist, resultados, saida)

    elapsed = time.perf_counter() - t0
    logger.info("Validação concluída em %.1fs", elapsed)

    # Resumo no console
    contadores = defaultdict(int)
    for r in resultados:
        contadores[r["status"]] += 1
    print("\n=== RESUMO DA VALIDAÇÃO ===")
    for status, qtd in sorted(contadores.items(), key=lambda x: -x[1]):
        print(f"  {status}: {qtd}")
    print(f"  Total: {len(resultados)}")
    print(f"\nRelatório: {saida}")


if __name__ == "__main__":
    main()
