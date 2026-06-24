import gc
import logging
import os
import re
import shutil
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

import requests
from googleapiclient.errors import HttpError
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

from src.app.gdrive.google_drive_auth import GoogleDriveAuth
from src.app.gdrive.settings import (
    GOOGLE_DRIVE_EMP_FOLDER_ID,
    GOOGLE_DRIVE_FOLDER_ID,
    GOOGLE_OAUTH_CREDENTIALS,
    GOOGLE_OAUTH_TOKEN,
    PDF_MIME_TYPE,
    XLS_MIME_TYPE,
    XLSX_MIME_TYPE,
    XLSM_MIME_TYPE,
    GOOGLE_OAUTH_TOKEN_SECRET
)
from src.app.supabase.supabase_api import (
    baixar_controle_supabase,
    buscar_id_empresa_supabase,
)
from src.schemas import LayoutNotRecognized, dispatch
from src.schemas.parsers.pdf_extractor import PDFExtractor
from src.services.chat_notifications import (
    GOOGLE_CHAT_SEND_URL,
    GOOGLE_CHAT_SPACE_NAME,
    registrar_e_enviar_notificacao,
)
from src.utils.helpers import log_memoria, planilha_lancamento, remover_senha_pdf, totalizador
from src.utils.logging_config import setup_logging


logger = logging.getLogger(__name__)

SEM_MOV_PREFIX = "[SEM MOV] - "
SEM_IMAGEM_PREFIX = "[SEM IMAGEM] - "
NAO_LEGIVEL_PREFIX = "[NAO LEGIVEL] - "
NOME_INVALIDO_PREFIX = "[NOME INVALIDO] - "
ERRO_SENHA_PREFIX = "[ERRO SENHA] - "
LAYOUT_INVALIDO_PREFIX = "[LAYOUT NAO RECONHECIDO] - "
PREFIXOS_INVALIDOS = (
    SEM_MOV_PREFIX,
    SEM_IMAGEM_PREFIX,
    NAO_LEGIVEL_PREFIX,
    NOME_INVALIDO_PREFIX,
    ERRO_SENHA_PREFIX,
    LAYOUT_INVALIDO_PREFIX,
)
HISTORICO_CONVERSOES = "HISTORICO_CONVERSOES.xlsx"
TIMEZONE_HISTORICO = ZoneInfo("America/Sao_Paulo")
HISTORICO_HEADERS = ("ID", "EMP", "DATA", "HORA", "NOME ARQUIVO", "PASTA DESTINO", "DATA HORA MOVIMENTO", "BANCO")
HISTORICO_HEADERS_SEM_BANCO = ("ID", "EMP", "DATA", "HORA", "NOME ARQUIVO", "PASTA DESTINO", "DATA HORA MOVIMENTO")
HISTORICO_HEADERS_COM_HORA_MOVIMENTO = ("ID", "EMP", "DATA", "HORA", "HORA MOVIMENTO", "NOME ARQUIVO", "PASTA DESTINO")
HISTORICO_HEADERS_SEM_MOVIMENTO = ("ID", "EMP", "DATA", "HORA", "NOME ARQUIVO", "PASTA DESTINO")
HISTORICO_HEADERS_ANTIGO = ("nome_arquivo", "data_hora_conversao", "pasta_destino")
BASE_EMP_ATIVAS = Path.cwd() / "data" / "BaseEmpAtivas.xlsm"
BASE_EMP_ATIVAS_SHEET = "EmpAtivas"
MODELO_LANCAMENTOS = Path.cwd() / "data" / "Lancamentos_Contabeis.xlsm"


@dataclass(frozen=True)
class NomeExtrato:
    nome_original: str
    nome_limpo: str
    mes: str
    ano: str
    competencia: str
    codigo_documento: str
    banco: str
    senha: str | None
    sem_movimentacao: bool
    empresa: str
    agencia: str
    conta: str

    @property
    def nome_lancamento(self) -> str:
        partes = Path(self.nome_limpo).stem.split("_")
        partes[1] = "LANCBAN"
        return "_".join(partes) + ".xlsm"


def agora_historico() -> datetime:
    return datetime.now(TIMEZONE_HISTORICO)


def formatar_mensagem_google_chat(
    nomes_extratos: list[str],
    pdfs_sem_movimentacao: list[str] | None = None,
    pdfs_nao_legiveis: list[str] | None = None,
    atualizados_sge: int | None = None,
    momento: datetime | None = None,
    nomes_invalidos: list[str] | None = None,
    erros_senha: list[str] | None = None,
    layouts_nao_reconhecidos: list[str] | None = None,
    erros_processamento: list[str] | None = None,
    empresas_nao_cadastradas_sge: list[str] | None = None,
    erros_sge: int = 0,
    erros_sge_detalhe: list[str] | None = None,
) -> str:
    momento = momento or agora_historico()
    pdfs_sem_movimentacao = pdfs_sem_movimentacao or []
    pdfs_nao_legiveis = pdfs_nao_legiveis or []
    nomes_invalidos = nomes_invalidos or []
    erros_senha = erros_senha or []
    layouts_nao_reconhecidos = layouts_nao_reconhecidos or []
    erros_processamento = erros_processamento or []
    empresas_nao_cadastradas_sge = empresas_nao_cadastradas_sge or []
    erros_sge_detalhe = erros_sge_detalhe or []
    blocos = []

    if nomes_extratos:
        data = momento.strftime("%d/%m/%y")
        hora = momento.strftime("%H:%M")
        lista_extratos = "\n".join(nomes_extratos)
        blocos.append(
            f"Extratos salvos {data} as {hora} e atualizado na Base de Dados:\n\n"
            f"{lista_extratos}"
        )

    if pdfs_sem_movimentacao:
        lista_sem_movimentacao = "\n".join(pdfs_sem_movimentacao)
        blocos.append(
            "PDFs sem movimentacao salvos na pasta da empresa:\n\n"
            f"{lista_sem_movimentacao}"
        )

    if pdfs_nao_legiveis:
        lista_nao_legiveis = "\n".join(pdfs_nao_legiveis)
        blocos.append(
            "PDFs nao legiveis movidos para 00_INVALIDOS:\n\n"
            f"{lista_nao_legiveis}"
        )

    for titulo, itens in (
        ("PDFs com nome invalido movidos para 00_INVALIDOS:", nomes_invalidos),
        ("PDFs com erro de senha movidos para 00_INVALIDOS:", erros_senha),
        ("PDFs com layout nao reconhecido movidos para 00_INVALIDOS:", layouts_nao_reconhecidos),
        ("PDFs com erro de processamento:", erros_processamento),
    ):
        if itens:
            blocos.append(f"{titulo}\n\n" + "\n".join(itens))

    if atualizados_sge is not None and atualizados_sge > 0:
        blocos.append(
            f"Baixa dada no portal SGE: {atualizados_sge} documento(s) atualizado(s)"
        )

    if empresas_nao_cadastradas_sge:
        lista = "\n".join(empresas_nao_cadastradas_sge)
        blocos.append(
            "Convertidos sem baixa no SGE (empresa nao cadastrada):\n\n"
            f"{lista}"
        )

    if erros_sge > 0:
        blocos.append(f"Erros ao dar baixa no SGE: {erros_sge}")

    if erros_sge_detalhe:
        lista = "\n".join(erros_sge_detalhe)
        blocos.append(f"Detalhes dos erros SGE:\n\n{lista}")

    return "\n\n".join(blocos)


def enviar_notificacao_google_chat(
    nomes_extratos: list[str],
    pdfs_sem_movimentacao: list[str] | None = None,
    pdfs_nao_legiveis: list[str] | None = None,
    atualizados_sge: int | None = None,
    momento: datetime | None = None,
    nomes_invalidos: list[str] | None = None,
    erros_senha: list[str] | None = None,
    layouts_nao_reconhecidos: list[str] | None = None,
    erros_processamento: list[str] | None = None,
    empresas_nao_cadastradas_sge: list[str] | None = None,
    google_drive=None,
    pasta_raiz_id: str | None = None,
    execucao_id: str | None = None,
    tipo: str = "CONCLUSAO",
    status_execucao: str | None = None,
    total_processados: int = 0,
    total_convertidos: int = 0,
    total_sem_movimentacao: int = 0,
    total_invalidos: int = 0,
    total_desbloqueados: int = 0,
    erros_sge: int = 0,
    erros_sge_detalhe: list[str] | None = None,
) -> bool:
    pdfs_sem_movimentacao = pdfs_sem_movimentacao or []
    pdfs_nao_legiveis = pdfs_nao_legiveis or []
    nomes_invalidos = nomes_invalidos or []
    erros_senha = erros_senha or []
    layouts_nao_reconhecidos = layouts_nao_reconhecidos or []
    erros_processamento = erros_processamento or []
    erros_sge_detalhe = erros_sge_detalhe or []
    empresas_nao_cadastradas_sge = empresas_nao_cadastradas_sge or []

    if execucao_id:
        momento = momento or agora_historico()
        detalhes = formatar_mensagem_google_chat(
            nomes_extratos,
            pdfs_sem_movimentacao=pdfs_sem_movimentacao,
            pdfs_nao_legiveis=pdfs_nao_legiveis,
            momento=momento,
            nomes_invalidos=nomes_invalidos,
            erros_senha=erros_senha,
            layouts_nao_reconhecidos=layouts_nao_reconhecidos,
            erros_processamento=erros_processamento,
            empresas_nao_cadastradas_sge=empresas_nao_cadastradas_sge,
            erros_sge=erros_sge,
            erros_sge_detalhe=erros_sge_detalhe,
        )
        resumo = (
            f"Recebimento: EXTRATOS\n"
            f"Data/hora: {momento.strftime('%d/%m/%Y %H:%M:%S')}\n\n"
            "Resumo:\n"
            f"Processados: {total_processados}\n"
            f"Convertidos: {total_convertidos}\n"
            f"Sem movimentacao: {total_sem_movimentacao}\n"
            f"Invalidos (ILEGÍVEL): {total_invalidos}\n"
            f"Desbloqueados: {total_desbloqueados}\n"
            f"Atualizados no PORTAL SGE: {atualizados_sge or 0}\n"
            f"Erros no PORTAL SGE: {erros_sge}"
        )
        if detalhes:
            resumo = f"{resumo}\n\n{detalhes}"
        elif tipo == "SEM_ARQUIVOS":
            resumo = f"{resumo}\n\nNenhum PDF encontrado na pasta EXT."

        return registrar_e_enviar_notificacao(
            google_drive=google_drive,
            pasta_raiz_id=pasta_raiz_id or GOOGLE_DRIVE_FOLDER_ID,
            execucao_id=execucao_id,
            tipo=tipo,
            mensagem=resumo,
        )

    if not any(
        (
            nomes_extratos,
            pdfs_sem_movimentacao,
            pdfs_nao_legiveis,
            nomes_invalidos,
            erros_senha,
            layouts_nao_reconhecidos,
            erros_processamento,
            empresas_nao_cadastradas_sge,
        )
    ):
        logger.info("Notificacao Google Chat nao enviada: nenhum resultado para informar")
        return False

    mensagem = formatar_mensagem_google_chat(
        nomes_extratos,
        pdfs_sem_movimentacao=pdfs_sem_movimentacao,
        pdfs_nao_legiveis=pdfs_nao_legiveis,
        atualizados_sge=atualizados_sge,
        momento=momento,
        nomes_invalidos=nomes_invalidos,
        erros_senha=erros_senha,
        layouts_nao_reconhecidos=layouts_nao_reconhecidos,
        erros_processamento=erros_processamento,
        empresas_nao_cadastradas_sge=empresas_nao_cadastradas_sge,
        erros_sge=erros_sge,
        erros_sge_detalhe=erros_sge_detalhe,
    )
    payload = {
        "space_name": GOOGLE_CHAT_SPACE_NAME,
        "message": mensagem,
    }

    try:
        logger.info(
            "Enviando notificacao ao Google Chat com %s extrato(s), %s PDF(s) sem movimentacao e %s PDF(s) nao legivel(is)",
            len(nomes_extratos),
            len(pdfs_sem_movimentacao),
            len(pdfs_nao_legiveis),
        )
        response = requests.post(
            GOOGLE_CHAT_SEND_URL,
            json=payload,
            timeout=30,
        )
        resposta_texto = (response.text or "")[:500]
        logger.info(
            "Resposta Google Chat: status=%s body=%s",
            response.status_code,
            resposta_texto,
        )

        if response.status_code < 200 or response.status_code >= 300:
            logger.error(
                "Falha ao enviar notificacao para o Google Chat: status=%s body=%s",
                response.status_code,
                resposta_texto,
            )
            return False

        logger.info(
            "Notificacao enviada ao Google Chat com %s extrato(s), %s PDF(s) sem movimentacao e %s PDF(s) nao legivel(is)",
            len(nomes_extratos),
            len(pdfs_sem_movimentacao),
            len(pdfs_nao_legiveis),
        )
        return True
    except Exception:
        logger.exception("Erro ao enviar notificacao para o Google Chat")
        return False


def pdf_possui_senha(pdf_path: Path) -> bool:
    reader = PdfReader(str(pdf_path))
    return bool(reader.is_encrypted)


def sanitizar_nome_senha(nome_arquivo: str) -> str:
    """Remove o segmento SENHA e seu valor para não propagar credenciais."""
    caminho = Path(nome_arquivo)
    partes = caminho.stem.split("_")
    partes_limpas = [
        parte.strip()
        for parte in partes
        if not re.match(r"^SENHA(?:\s+.*)?$", parte.strip(), flags=re.IGNORECASE)
    ]
    return "_".join(partes_limpas) + caminho.suffix


def interpretar_nome_extrato(nome_arquivo: str) -> NomeExtrato:
    caminho = Path(nome_arquivo)
    if caminho.suffix.lower() != ".pdf":
        raise ValueError("Arquivo nao possui extensao PDF")

    partes = [parte.strip() for parte in caminho.stem.split("_")]
    if len(partes) < 6 or any(not parte for parte in partes):
        raise ValueError("Nome fora do padrao de segmentos esperado")

    periodo, codigo_documento, banco = partes[:3]
    periodo_match = re.fullmatch(r"(?P<mes>\d{2})(?P<ano>\d{2})", periodo)
    if not periodo_match:
        raise ValueError("Competencia deve seguir o formato MMAA")
    if codigo_documento.upper() != "EXTBAN":
        raise ValueError("Codigo do documento deve ser EXTBAN")
    if not banco:
        raise ValueError("Banco ausente no nome")

    indice = 3
    senha = None
    senha_match = re.fullmatch(
        r"SENHA(?:\s+(?P<senha>[^\s_]+))?",
        partes[indice],
        flags=re.IGNORECASE,
    )
    if senha_match:
        senha = senha_match.group("senha")
        if not senha:
            raise ValueError("Segmento SENHA sem valor")
        indice += 1

    sem_movimentacao = indice < len(partes) and partes[indice].upper() == "SM"
    if sem_movimentacao:
        indice += 1

    if len(partes) != indice + 3:
        raise ValueError("Nome deve terminar com EMPRESA_AG AGENCIA_CC CONTA")

    empresa = partes[indice]
    agencia_match = re.fullmatch(r"AG\s+(.+)", partes[indice + 1], flags=re.IGNORECASE)
    conta_match = re.fullmatch(r"CC\s+(.+)", partes[indice + 2], flags=re.IGNORECASE)
    if not empresa or not agencia_match or not conta_match:
        raise ValueError("Empresa, agencia ou conta ausente no nome")

    partes_limpas = partes[:3]
    if sem_movimentacao:
        partes_limpas.append("SM")
    partes_limpas.extend(partes[indice:])
    nome_limpo = "_".join(partes_limpas) + caminho.suffix
    mes = periodo_match.group("mes")
    ano = periodo_match.group("ano")

    return NomeExtrato(
        nome_original=nome_arquivo,
        nome_limpo=nome_limpo,
        mes=mes,
        ano=ano,
        competencia=f"{mes}-20{ano}",
        codigo_documento=codigo_documento.upper(),
        banco=banco.upper(),
        senha=senha,
        sem_movimentacao=sem_movimentacao,
        empresa=empresa,
        agencia=agencia_match.group(1).strip(),
        conta=conta_match.group(1).strip(),
    )


def extrair_senha_nome_pdf(nome_arquivo: str) -> tuple[str, str] | None:
    try:
        dados = interpretar_nome_extrato(nome_arquivo)
    except ValueError:
        caminho = Path(nome_arquivo)
        match = re.match(
            r"^(?P<prefixo>.+?)[\s_]+SENHA\s+(?P<senha>[^\s_]+)(?P<sufixo>.*)$",
            caminho.stem,
            flags=re.IGNORECASE,
        )
        if not match:
            return None
        senha = match.group("senha").strip()
        sufixo = match.group("sufixo").strip()
        nome_limpo_stem = match.group("prefixo").rstrip()
        if sufixo.startswith("_"):
            nome_limpo_stem = f"{nome_limpo_stem}{sufixo}"
        elif sufixo:
            nome_limpo_stem = f"{nome_limpo_stem} {sufixo}"
        return senha, f"{nome_limpo_stem}{caminho.suffix}"

    if not dados.senha:
        return None
    return dados.senha, dados.nome_limpo


def nome_indica_sem_movimentacao(nome_arquivo: str) -> bool:
    try:
        return interpretar_nome_extrato(nome_arquivo).sem_movimentacao
    except ValueError:
        partes = [parte.strip().upper() for parte in Path(nome_arquivo).stem.split("_")]
        return "SM" in partes


def nome_com_prefixo_invalido(nome_arquivo: str, prefixo: str) -> str:
    if nome_arquivo.startswith(PREFIXOS_INVALIDOS):
        return nome_arquivo
    return f"{prefixo}{nome_arquivo}"


def normalizar_nome_empresa(valor) -> str:
    return " ".join(str(valor or "").strip().upper().split())


def normalizar_cabecalho_base(valor) -> str:
    texto = normalizar_nome_empresa(valor)
    texto_sem_acentos = "".join(
        caractere
        for caractere in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(caractere)
    )

    if "RAZ" in texto_sem_acentos and "SOCIAL" in texto_sem_acentos:
        return "RAZ" + chr(195) + "O SOCIAL"

    return texto_sem_acentos


def extrair_cliente_nome_arquivo(nome_arquivo: str) -> str:
    try:
        return interpretar_nome_extrato(nome_arquivo).empresa
    except ValueError:
        pass

    stem = Path(nome_arquivo).stem
    partes = [parte.strip() for parte in stem.split("_") if parte.strip()]

    for i, parte in enumerate(partes):
        if parte.upper().startswith("AG ") or parte.upper() == "AG":
            return partes[i - 1] if i > 0 else stem.strip()

    return partes[-1] if partes else stem.strip()


def extrair_banco_nome_arquivo(nome_arquivo: str) -> str:
    partes_canonicas = [
        parte.strip()
        for parte in Path(nome_arquivo).stem.split("_")
        if parte.strip()
    ]
    if (
        len(partes_canonicas) >= 3
        and re.fullmatch(r"\d{4}", partes_canonicas[0])
        and partes_canonicas[1].upper() in {"EXTBAN", "LANCBAN"}
    ):
        return partes_canonicas[2].upper()

    try:
        return interpretar_nome_extrato(nome_arquivo).banco
    except ValueError:
        pass

    stem = Path(nome_arquivo).stem
    partes = [parte.strip() for parte in stem.split("_") if parte.strip()]

    if len(partes) < 2:
        return ""

    descricao = re.sub(
        r"^(?:EXT|LANC)BAN\s+",
        "",
        partes[1],
        flags=re.IGNORECASE,
    ).strip()

    if not descricao:
        return ""

    tokens = descricao.split()

    while tokens and re.fullmatch(r"[\d./-]+", tokens[-1]):
        tokens.pop()

    return " ".join(tokens).strip().upper()


def extrair_periodo_nome_arquivo(nome_arquivo: str) -> tuple[str, str]:
    try:
        dados = interpretar_nome_extrato(nome_arquivo)
        return dados.mes, dados.ano
    except ValueError:
        pass

    prefixo_periodo = Path(nome_arquivo).stem.split("_", maxsplit=1)[0]
    match = re.match(r"^(?P<mes>\d{2})(?P<ano>\d{2})$", prefixo_periodo)

    if not match:
        raise ValueError(f"Nome do arquivo sem periodo MMAA valido: {nome_arquivo}")

    return match.group("mes"), match.group("ano")


def normalizar_id_empresa(empresa_id) -> str:
    if empresa_id is None:
        return ""

    if isinstance(empresa_id, float) and empresa_id.is_integer():
        return str(int(empresa_id))

    return str(empresa_id).strip()


def extrair_competencia_nome_arquivo(nome_arquivo: str) -> str:
    try:
        return interpretar_nome_extrato(nome_arquivo).competencia
    except ValueError:
        pass

    stem = Path(nome_arquivo).stem
    prefixo = stem.split("_", maxsplit=1)[0]
    match = re.match(r"^(?P<mes>\d{2})(?P<ano>\d{2})$", prefixo)

    if not match:
        raise ValueError(f"Arquivo sem periodo MMAA valido: {nome_arquivo}")

    return f"{match.group('mes')}-20{match.group('ano')}"


def extrair_codigo_documento(nome_arquivo: str) -> str:
    try:
        return interpretar_nome_extrato(nome_arquivo).codigo_documento
    except ValueError:
        pass

    partes = Path(nome_arquivo).stem.split("_")

    if len(partes) < 2:
        return "EXTBAN"

    cod_doc = partes[1].strip().upper().split()[0]
    return cod_doc


def nome_pasta_empresa(empresa_id, empresa_nome: str) -> str:
    empresa_id_normalizado = normalizar_id_empresa(empresa_id)
    empresa_nome_normalizado = normalizar_nome_empresa(empresa_nome)

    if not empresa_id_normalizado or not empresa_nome_normalizado:
        raise ValueError("ID/EMP ausente para resolver pasta da empresa")

    return f"{empresa_id_normalizado}_{empresa_nome_normalizado}"


def resolver_pasta_emp_base(
    google_drive: GoogleDriveAuth,
    pasta_base_id: str,
) -> str:
    logger.info("Localizando pasta EMP dentro da pasta base SRVARQ")
    try:
        pasta_emp = google_drive.get_or_create_folder(
            folder_id_pai=pasta_base_id,
            name_folder="EMP",
        )
    except HttpError as erro:
        if erro.resp.status != 404:
            raise

        logger.warning(
            "Pasta base SRVARQ nao acessivel pelo ID configurado. Tentando localizar por nome."
        )
        pasta_base = (
            google_drive.find_folder_by_name("SRVARQ", shared_with_me=True)
            or google_drive.find_folder_by_name("SRVARQ")
        )

        if not pasta_base:
            raise

        logger.info("Pasta SRVARQ localizada por nome: %s", pasta_base["id"])
        pasta_emp = google_drive.get_or_create_folder(
            folder_id_pai=pasta_base["id"],
            name_folder="EMP",
        )

    return pasta_emp["id"]


def carregar_empresas_ativas(
    base_path: Path = BASE_EMP_ATIVAS,
    sheet_name: str = BASE_EMP_ATIVAS_SHEET,
) -> dict[str, tuple[object, str]]:
    if not base_path.exists():
        logger.warning("Base de empresas ativas nao encontrada: %s", base_path)
        return {}

    workbook = load_workbook(base_path, read_only=True, data_only=True)

    try:
        if sheet_name not in workbook.sheetnames:
            logger.warning("Aba %s nao encontrada na base: %s", sheet_name, base_path)
            return {}

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)

        if not headers:
            logger.warning("Base de empresas ativas sem cabecalho: %s", base_path)
            return {}

        normalized_headers = [
            normalizar_cabecalho_base(header)
            for header in headers
        ]

        try:
            id_index = normalized_headers.index("ID")
            emp_index = normalized_headers.index("RAZÃO SOCIAL")
        except ValueError:
            logger.warning("Base de empresas ativas sem colunas ID/Razao social: %s", base_path)
            return {}

        empresas = {}

        for row in rows:
            if not row:
                continue

            empresa = row[emp_index] if emp_index < len(row) else None
            empresa_normalizada = normalizar_nome_empresa(empresa)

            if not empresa_normalizada:
                continue

            empresa_id = row[id_index] if id_index < len(row) else ""
            empresas[empresa_normalizada] = (empresa_id or "", str(empresa).strip())

        return empresas
    finally:
        workbook.close()


def validar_preparacao_fluxo(
    base_path: Path = BASE_EMP_ATIVAS,
    modelo_path: Path = MODELO_LANCAMENTOS,
) -> None:
    if not base_path.exists():
        raise FileNotFoundError(f"Base de empresas ativas nao encontrada: {base_path}")
    if not modelo_path.exists():
        raise FileNotFoundError(f"Modelo de lancamentos nao encontrado: {modelo_path}")

    base_workbook = load_workbook(base_path, read_only=True, data_only=True)
    try:
        if BASE_EMP_ATIVAS_SHEET not in base_workbook.sheetnames:
            raise ValueError(f"Aba {BASE_EMP_ATIVAS_SHEET} nao encontrada em {base_path}")
        headers = next(
            base_workbook[BASE_EMP_ATIVAS_SHEET].iter_rows(values_only=True),
            None,
        )
        normalized_headers = [normalizar_cabecalho_base(header) for header in (headers or [])]
        if "ID" not in normalized_headers or "RAZÃO SOCIAL" not in normalized_headers:
            raise ValueError("Base de empresas ativas sem colunas ID/Razao social")
    finally:
        base_workbook.close()

    modelo_workbook = load_workbook(modelo_path, read_only=True, data_only=False)
    try:
        if "Plan1" not in modelo_workbook.sheetnames:
            raise ValueError(f"Aba Plan1 nao encontrada em {modelo_path}")
    finally:
        modelo_workbook.close()


def validar_dataframe_extrato(df) -> None:
    colunas_obrigatorias = {"DATA", "VALOR", "TIPO", "DESCRIÇÃO"}
    ausentes = colunas_obrigatorias.difference(df.columns)
    if ausentes:
        raise ValueError(
            "DataFrame do extrato sem colunas obrigatorias: "
            + ", ".join(sorted(ausentes))
        )


def enviar_ou_atualizar_arquivo(
    google_drive: GoogleDriveAuth,
    caminho_local: Path,
    folder_id_destino: str,
    type_file: str,
    name_drive: str,
):
    existente = google_drive.find_file_by_name(
        folder_id=folder_id_destino,
        name=name_drive,
        mime_type=type_file,
    )
    if existente:
        return google_drive.update_file(
            file_id=existente["id"],
            caminho_local=caminho_local,
            type_file=type_file,
            name_drive=name_drive,
        )
    return google_drive.upload(
        caminho_local=caminho_local,
        folder_id_destino=folder_id_destino,
        type_file=type_file,
        name_drive=name_drive,
    )


def buscar_empresa_por_cliente(
    cliente: str,
    empresas: dict[str, tuple[object, str]] | None = None,
) -> tuple[object, str]:
    empresas = empresas if empresas is not None else carregar_empresas_ativas()
    cliente_normalizado = normalizar_nome_empresa(cliente)
    empresa = empresas.get(cliente_normalizado)

    if not empresa:
        logger.warning("Cliente nao encontrado na BaseEmpAtivas: %s", cliente)
        return "", ""

    return empresa


def resolver_pasta_destino_emp(
    google_drive: GoogleDriveAuth,
    pasta_emp_id: str,
    arquivo_nome: str,
    empresa_id,
    empresa_nome: str,
) -> tuple[str, str]:
    mes, ano = extrair_periodo_nome_arquivo(arquivo_nome)
    nome_cliente = nome_pasta_empresa(empresa_id, empresa_nome)
    caminho_partes = ["EMP", nome_cliente, "MOV", "CONT", ano, mes, "EXT"]
    pasta_atual_id = pasta_emp_id

    for i, nome_pasta in enumerate(caminho_partes[1:]):
        if i == 0:
            pasta_atual = google_drive.list_folder_by_name(
                folder_id=pasta_atual_id,
                name_folder=nome_pasta,
            )
            if not pasta_atual:
                raise ValueError(
                    f"Pasta de cliente '{nome_cliente}' nao encontrada em EMP. "
                    "A pasta deve existir previamente no Google Drive."
                )
        else:
            pasta_atual = google_drive.get_or_create_folder(
                folder_id_pai=pasta_atual_id,
                name_folder=nome_pasta,
            )
        pasta_atual_id = pasta_atual["id"]

    return pasta_atual_id, "/".join(caminho_partes)


def separar_data_hora_historico(valor) -> tuple[str, str]:
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d"), valor.strftime("%H:%M:%S")

    texto = str(valor or "").strip()
    if not texto:
        return "", ""

    partes = texto.split(maxsplit=1)
    data = partes[0]
    hora = partes[1] if len(partes) > 1 else ""
    return data, hora


def preencher_banco_historico(worksheet) -> None:
    headers = tuple(
        cell.value
        for cell in worksheet[1]
    ) if worksheet.max_row else ()

    if "NOME ARQUIVO" not in headers or "BANCO" not in headers:
        return

    nome_arquivo_coluna = headers.index("NOME ARQUIVO") + 1
    banco_coluna = headers.index("BANCO") + 1

    for row_index in range(2, worksheet.max_row + 1):
        banco_cell = worksheet.cell(row=row_index, column=banco_coluna)

        if banco_cell.value:
            continue

        nome_arquivo = worksheet.cell(row=row_index, column=nome_arquivo_coluna).value
        banco = extrair_banco_nome_arquivo(nome_arquivo or "")

        if banco:
            banco_cell.value = banco


def migrar_historico_antigo(worksheet) -> None:
    headers = tuple(
        cell.value
        for cell in worksheet[1]
    ) if worksheet.max_row else ()

    if headers == HISTORICO_HEADERS:
        preencher_banco_historico(worksheet)
        return

    linhas_antigas = list(worksheet.iter_rows(min_row=2, values_only=True))
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(HISTORICO_HEADERS)

    if headers == HISTORICO_HEADERS_SEM_BANCO:
        for empresa_id, empresa_nome, data, hora, nome_arquivo, pasta_destino, data_hora_movimento in linhas_antigas:
            worksheet.append(
                [
                    empresa_id or "",
                    empresa_nome or "",
                    data or "",
                    hora or "",
                    nome_arquivo or "",
                    pasta_destino or "",
                    data_hora_movimento or "",
                    extrair_banco_nome_arquivo(nome_arquivo or ""),
                ]
            )
        return

    if headers == HISTORICO_HEADERS_COM_HORA_MOVIMENTO:
        for empresa_id, empresa_nome, data, hora, hora_movimento, nome_arquivo, pasta_destino in linhas_antigas:
            data_hora_movimento = ""
            if data and hora_movimento:
                data_hora_movimento = f"{data} {hora_movimento}"

            worksheet.append(
                [
                    empresa_id or "",
                    empresa_nome or "",
                    data or "",
                    hora or "",
                    nome_arquivo or "",
                    pasta_destino or "",
                    data_hora_movimento,
                    extrair_banco_nome_arquivo(nome_arquivo or ""),
                ]
            )
        return

    if headers == HISTORICO_HEADERS_SEM_MOVIMENTO:
        for empresa_id, empresa_nome, data, hora, nome_arquivo, pasta_destino in linhas_antigas:
            worksheet.append(
                [
                    empresa_id or "",
                    empresa_nome or "",
                    data or "",
                    hora or "",
                    nome_arquivo or "",
                    pasta_destino or "",
                    "",
                    extrair_banco_nome_arquivo(nome_arquivo or ""),
                ]
            )
        return

    if headers != HISTORICO_HEADERS_ANTIGO:
        return

    for nome_arquivo, data_hora_conversao, pasta_destino in linhas_antigas:
        data, hora = separar_data_hora_historico(data_hora_conversao)
        worksheet.append(
            [
                "",
                "",
                data,
                hora,
                nome_arquivo or "",
                pasta_destino or "",
                "",
                extrair_banco_nome_arquivo(nome_arquivo or ""),
            ]
        )


def renomear_e_mover_para_invalidos(
    google_drive: GoogleDriveAuth,
    arquivo_id: str,
    arquivo_nome: str,
    invalidos_id: str,
    prefixo: str,
    motivo: str,
) -> str:
    nome_final = nome_com_prefixo_invalido(arquivo_nome, prefixo)

    if nome_final != arquivo_nome:
        google_drive.rename_file(arquivo_id, nome_final)

    google_drive.move_file(
        file_id=arquivo_id,
        folder_id_destino=invalidos_id,
    )
    logger.warning("%s. Movendo para 00_INVALIDOS: %s", motivo, nome_final)
    return nome_final


def registrar_historico_conversao(
    google_drive: GoogleDriveAuth,
    pasta_raiz_id: str,
    temp_dir: Path,
    nomes_arquivos: list[str],
    pasta_destino: str,
    data_hora: datetime | None = None,
    data_hora_movimento: datetime | None = None,
    empresa_id=None,
    empresa_nome: str | None = None,
    empresas: dict[str, tuple[object, str]] | None = None,
) -> None:
    temp_dir.mkdir(parents=True, exist_ok=True)
    historico_local = temp_dir / HISTORICO_CONVERSOES
    arquivo_historico = google_drive.find_file_by_name(
        folder_id=pasta_raiz_id,
        name=HISTORICO_CONVERSOES,
        mime_type=XLSX_MIME_TYPE,
    )

    if arquivo_historico:
        logger.info("Baixando historico de conversoes existente")
        google_drive.download(
            file_id=arquivo_historico["id"],
            destino_local=historico_local,
        )
        workbook = load_workbook(historico_local)
        worksheet = workbook.active
        migrar_historico_antigo(worksheet)
    else:
        logger.info("Criando historico de conversoes")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "historico"
        worksheet.append(HISTORICO_HEADERS)

    momento = data_hora or agora_historico()
    momento_movimento = data_hora_movimento or momento
    data_formatada = momento.strftime("%Y-%m-%d")
    hora_formatada = momento.strftime("%H:%M:%S")
    data_hora_movimento_formatada = momento_movimento.strftime("%Y-%m-%d %H:%M:%S")
    if empresa_id is None or empresa_nome is None:
        cliente = extrair_cliente_nome_arquivo(nomes_arquivos[0]) if nomes_arquivos else ""
        empresa_id, empresa_nome = buscar_empresa_por_cliente(cliente, empresas=empresas)

    for nome_arquivo in nomes_arquivos:
        worksheet.append(
            [
                empresa_id,
                empresa_nome,
                data_formatada,
                hora_formatada,
                nome_arquivo,
                pasta_destino,
                data_hora_movimento_formatada,
                extrair_banco_nome_arquivo(nome_arquivo),
            ]
        )

    workbook.save(historico_local)

    if arquivo_historico:
        google_drive.update_file(
            file_id=arquivo_historico["id"],
            caminho_local=historico_local,
            type_file=XLSX_MIME_TYPE,
            name_drive=HISTORICO_CONVERSOES,
        )
        logger.info("Historico de conversoes atualizado: %s", HISTORICO_CONVERSOES)
    else:
        google_drive.upload(
            caminho_local=historico_local,
            folder_id_destino=pasta_raiz_id,
            type_file=XLSX_MIME_TYPE,
            name_drive=HISTORICO_CONVERSOES,
        )
        logger.info("Historico de conversoes criado: %s", HISTORICO_CONVERSOES)

    historico_local.unlink(missing_ok=True)


def arquivar_pdf_sem_movimentacao(
    google_drive: GoogleDriveAuth,
    pasta_raiz_id: str,
    pasta_emp_id: str,
    temp_dir: Path,
    arquivo_id: str,
    nome_arquivo: str,
    empresas: dict[str, tuple[object, str]] | None = None,
) -> dict | None:
    cliente = extrair_cliente_nome_arquivo(nome_arquivo)
    empresa_id, empresa_nome = buscar_empresa_por_cliente(cliente, empresas=empresas)

    try:
        pasta_destino_id, pasta_destino_historico = resolver_pasta_destino_emp(
            google_drive=google_drive,
            pasta_emp_id=pasta_emp_id,
            arquivo_nome=nome_arquivo,
            empresa_id=empresa_id,
            empresa_nome=empresa_nome,
        )
    except ValueError:
        logger.exception(
            "Nao foi possivel resolver pasta EMP para PDF sem movimentacao. Arquivo mantido na pasta EXT: %s",
            nome_arquivo,
        )
        return None

    momento = agora_historico()
    logger.info("Movendo PDF sem movimentacao para pasta destino EMP: %s", nome_arquivo)
    google_drive.move_file(
        file_id=arquivo_id,
        folder_id_destino=pasta_destino_id,
    )
    momento_movimento = agora_historico()

    registrar_historico_conversao(
        google_drive=google_drive,
        pasta_raiz_id=pasta_raiz_id,
        temp_dir=temp_dir,
        nomes_arquivos=[nome_arquivo],
        pasta_destino=pasta_destino_historico,
        data_hora=momento,
        data_hora_movimento=momento_movimento,
        empresa_id=empresa_id,
        empresa_nome=empresa_nome,
        empresas=empresas,
    )
    logger.info(
        "PDF sem movimentacao salvo na pasta da empresa: %s -> %s",
        nome_arquivo,
        pasta_destino_historico,
    )
    return {
        "notificacao": f"{nome_arquivo} - {pasta_destino_historico}",
        "empresa_id": empresa_id,
        "pasta_destino": pasta_destino_historico,
    }


def _executar_conversao(execucao_id: str):
    setup_logging()
    logger.info("Iniciando fluxo de conversao")
    logger.info("Validando arquivos locais obrigatorios")
    validar_preparacao_fluxo()
    empresas = carregar_empresas_ativas()

    logger.info("Autenticando Google Drive")
    google_drive = GoogleDriveAuth(
        credentials_path=GOOGLE_OAUTH_CREDENTIALS,
        token_path=GOOGLE_OAUTH_TOKEN,
        token_secret_path=GOOGLE_OAUTH_TOKEN_SECRET,
    )

    extratos_id = GOOGLE_DRIVE_FOLDER_ID
    emp_id = GOOGLE_DRIVE_EMP_FOLDER_ID
    lancamento = MODELO_LANCAMENTOS
    temp_dir = Path.cwd() / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    logger.info("Diretorio temporario preparado: %s", temp_dir)
    emp_raiz_id = resolver_pasta_emp_base(
        google_drive=google_drive,
        pasta_base_id=emp_id,
    )
    logger.info("Pasta raiz operacional do Google Drive: %s", extratos_id)

    logger.info("Criando ou recuperando pasta de entrada EXT no Google Drive")
    pasta_entrada_ext = google_drive.get_or_create_folder(
        folder_id_pai=extratos_id,
        name_folder="EXT",
    )
    entrada_ext_id = pasta_entrada_ext["id"]
    logger.info("Pasta de entrada EXT do Google Drive: %s", entrada_ext_id)

    logger.info("Criando ou recuperando pasta de invalidos no Google Drive")
    pasta_invalidos = google_drive.get_or_create_folder(
        folder_id_pai=extratos_id,
        name_folder="00_INVALIDOS",
    )
    invalidos_id = pasta_invalidos["id"]

    logger.info("Listando PDFs da pasta EXT do Google Drive")
    arquivos = google_drive.pdfs(
        folder_id=entrada_ext_id,
        pdf_type=PDF_MIME_TYPE,
    )
    logger.info("PDFs encontrados na pasta EXT do Google Drive: %s", len(arquivos))

    max_pdfs = int(os.getenv("MAX_PDFS_POR_EXECUCAO", "0"))
    if max_pdfs > 0:
        logger.info("Limitando processamento a %s PDFs (MAX_PDFS_POR_EXECUCAO)", max_pdfs)
        arquivos = arquivos[:max_pdfs]

    if not arquivos:
        logger.info("Pasta EXT vazia")

    processados = 0
    convertidos = 0
    invalidos = 0
    desbloqueados = 0
    sem_movimentacao = 0
    extratos_notificacao = []
    pdfs_sem_movimentacao_notificacao = []
    pdfs_nao_legiveis_notificacao = []
    nomes_invalidos_notificacao = []
    erros_senha_notificacao = []
    layouts_invalidos_notificacao = []
    erros_processamento_notificacao = []
    documentos_convertidos = []
    documentos_para_baixa = []

    for arquivo_drive in arquivos:
        arquivo_id = arquivo_drive["id"]
        arquivo_nome_original = arquivo_drive["name"]
        arquivo_nome_seguro = sanitizar_nome_senha(arquivo_nome_original)
        arquivo_nome = arquivo_nome_seguro
        processados += 1
        inicio_pdf = time.perf_counter()
        pdf_local = temp_dir / f"{arquivo_id}.pdf"
        pdf_sem_senha = temp_dir / f"{arquivo_id}_sem_senha.pdf"
        pdf_processamento = pdf_local
        pdf = None
        df = None
        dest_lancamento = None
        dest_excel = None

        logger.info("Processando PDF: %s", arquivo_nome_seguro)
        log_memoria(f"antes_processar_pdf_{arquivo_nome_seguro}")

        try:
            logger.info("Baixando PDF do Google Drive: %s", arquivo_nome_seguro)
            inicio_download = time.perf_counter()
            google_drive.download(
                file_id=arquivo_id,
                destino_local=pdf_local,
            )
            logger.info(
                "Download concluido: %s | tempo=%.2fs",
                arquivo_nome_seguro,
                time.perf_counter() - inicio_download,
            )

            try:
                inicio_senha = time.perf_counter()
                protegido = pdf_possui_senha(pdf_local)
                logger.info(
                    "Verificacao de senha concluida: %s | protegido=%s | tempo=%.2fs",
                    arquivo_nome_seguro,
                    protegido,
                    time.perf_counter() - inicio_senha,
                )
            except Exception:
                logger.exception("PDF invalido ou corrompido: %s", arquivo_nome_seguro)
                nome_final = renomear_e_mover_para_invalidos(
                    google_drive=google_drive,
                    arquivo_id=arquivo_id,
                    arquivo_nome=arquivo_nome_seguro,
                    invalidos_id=invalidos_id,
                    prefixo=NAO_LEGIVEL_PREFIX,
                    motivo="PDF invalido, corrompido ou nao legivel",
                )
                pdfs_nao_legiveis_notificacao.append(nome_final)
                invalidos += 1
                continue

            dados_senha = extrair_senha_nome_pdf(arquivo_nome_original)

            if protegido:
                if not dados_senha:
                    nome_final = renomear_e_mover_para_invalidos(
                        google_drive=google_drive,
                        arquivo_id=arquivo_id,
                        arquivo_nome=arquivo_nome_seguro,
                        invalidos_id=invalidos_id,
                        prefixo=ERRO_SENHA_PREFIX,
                        motivo="PDF protegido sem senha valida no nome",
                    )
                    erros_senha_notificacao.append(nome_final)
                    invalidos += 1
                    continue

                senha, arquivo_nome = dados_senha
                try:
                    remover_senha_pdf(
                        caminho_pdf=str(pdf_local),
                        senha=senha,
                        caminho_saida=str(pdf_sem_senha),
                    )
                except ValueError:
                    logger.exception(
                        "Senha invalida ou falha ao desbloquear PDF: %s",
                        arquivo_nome,
                    )
                    nome_final = renomear_e_mover_para_invalidos(
                        google_drive=google_drive,
                        arquivo_id=arquivo_id,
                        arquivo_nome=arquivo_nome,
                        invalidos_id=invalidos_id,
                        prefixo=ERRO_SENHA_PREFIX,
                        motivo="Senha invalida ou falha ao desbloquear PDF",
                    )
                    erros_senha_notificacao.append(nome_final)
                    invalidos += 1
                    continue

                google_drive.update_file(
                    file_id=arquivo_id,
                    caminho_local=pdf_sem_senha,
                    type_file=PDF_MIME_TYPE,
                    name_drive=arquivo_nome,
                )
                pdf_processamento = pdf_sem_senha
                desbloqueados += 1
                logger.info("PDF desbloqueado e atualizado no Drive: %s", arquivo_nome)
            elif dados_senha:
                _, arquivo_nome = dados_senha
                google_drive.rename_file(arquivo_id, arquivo_nome)
                logger.info("Nome do PDF normalizado no Drive: %s", arquivo_nome)
            else:
                arquivo_nome = arquivo_nome_original

            try:
                dados_nome = interpretar_nome_extrato(arquivo_nome)
            except ValueError as erro_nome:
                nomes_invalidos_notificacao.append(
                    f"{arquivo_nome} - nome fora do padrao: {erro_nome}"
                )
                continue

            arquivo_nome = dados_nome.nome_limpo
            arquivo_stem = Path(arquivo_nome).stem

            if dados_nome.sem_movimentacao:
                resultado_sm = arquivar_pdf_sem_movimentacao(
                    google_drive=google_drive,
                    pasta_raiz_id=extratos_id,
                    pasta_emp_id=emp_raiz_id,
                    temp_dir=temp_dir,
                    arquivo_id=arquivo_id,
                    nome_arquivo=arquivo_nome,
                    empresas=empresas,
                )
                if resultado_sm:
                    pdfs_sem_movimentacao_notificacao.append(resultado_sm["notificacao"])
                    sem_movimentacao += 1
                    documentos_para_baixa.append({
                        "empresa_id": resultado_sm["empresa_id"],
                        "mes": dados_nome.mes,
                        "ano": dados_nome.ano,
                        "competencia": dados_nome.competencia,
                        "codigo_documento": dados_nome.codigo_documento,
                        "banco": dados_nome.banco,
                        "agencia": dados_nome.agencia,
                        "conta": dados_nome.conta,
                        "arquivo_nome": arquivo_nome,
                        "pasta_destino": resultado_sm["pasta_destino"],
                        "quantidade_arquivos": 1,
                        "status": "Não Aplicável",
                    })
                else:
                    erros_processamento_notificacao.append(
                        f"{arquivo_nome} - empresa ou pasta nao encontrada; mantido na EXT"
                    )
                continue

            logger.info("Extraindo conteudo do PDF: %s", arquivo_nome)
            inicio_extracao = time.perf_counter()
            pdf, total_paginas = PDFExtractor(pdf_processamento).extract()
            logger.info(
                "Extracao finalizada: %s | linhas=%s | paginas=%s | tempo=%.2fs",
                arquivo_nome,
                len(pdf or []),
                total_paginas,
                time.perf_counter() - inicio_extracao,
            )

            if not pdf:
                nome_nao_legivel = renomear_e_mover_para_invalidos(
                    google_drive=google_drive,
                    arquivo_id=arquivo_id,
                    arquivo_nome=arquivo_nome,
                    invalidos_id=invalidos_id,
                    prefixo=NAO_LEGIVEL_PREFIX,
                    motivo="PDF sem texto extraivel, nao legivel ou possivelmente imagem",
                )
                pdfs_nao_legiveis_notificacao.append(nome_nao_legivel)
                invalidos += 1
                continue

            logger.info("Identificando layout e convertendo PDF em DataFrame: %s", arquivo_nome)
            inicio_parser = time.perf_counter()
            log_memoria(f"antes_parser_{arquivo_nome}")
            try:
                df = dispatch(pdf)
            except LayoutNotRecognized:
                logger.exception("Layout nao reconhecido. Movendo para invalidos: %s", arquivo_nome)
                nome_final = renomear_e_mover_para_invalidos(
                    google_drive=google_drive,
                    arquivo_id=arquivo_id,
                    arquivo_nome=arquivo_nome,
                    invalidos_id=invalidos_id,
                    prefixo=LAYOUT_INVALIDO_PREFIX,
                    motivo="Layout bancario nao reconhecido",
                )
                layouts_invalidos_notificacao.append(nome_final)
                invalidos += 1
                continue
            log_memoria(f"depois_parser_{arquivo_nome}")
            logger.info(
                "Parser concluido: %s | tempo=%.2fs",
                arquivo_nome,
                time.perf_counter() - inicio_parser,
            )

            if df is None or df.empty:
                resultado_sm = arquivar_pdf_sem_movimentacao(
                    google_drive=google_drive,
                    pasta_raiz_id=extratos_id,
                    pasta_emp_id=emp_raiz_id,
                    temp_dir=temp_dir,
                    arquivo_id=arquivo_id,
                    nome_arquivo=arquivo_nome,
                    empresas=empresas,
                )
                if resultado_sm:
                    pdfs_sem_movimentacao_notificacao.append(resultado_sm["notificacao"])
                    sem_movimentacao += 1
                    documentos_para_baixa.append({
                        "empresa_id": resultado_sm["empresa_id"],
                        "mes": dados_nome.mes,
                        "ano": dados_nome.ano,
                        "competencia": dados_nome.competencia,
                        "codigo_documento": dados_nome.codigo_documento,
                        "banco": dados_nome.banco,
                        "agencia": dados_nome.agencia,
                        "conta": dados_nome.conta,
                        "arquivo_nome": arquivo_nome,
                        "pasta_destino": resultado_sm["pasta_destino"],
                        "quantidade_arquivos": 1,
                        "status": "Não Aplicável",
                    })
                else:
                    erros_processamento_notificacao.append(
                        f"{arquivo_nome} - empresa ou pasta nao encontrada; mantido na EXT"
                    )
                continue

            validar_dataframe_extrato(df)
            empresa_id, empresa_nome = buscar_empresa_por_cliente(
                dados_nome.empresa,
                empresas=empresas,
            )

            try:
                pasta_destino_id, pasta_destino_historico = resolver_pasta_destino_emp(
                    google_drive=google_drive,
                    pasta_emp_id=emp_raiz_id,
                    arquivo_nome=arquivo_nome,
                    empresa_id=empresa_id,
                    empresa_nome=empresa_nome,
                )
            except ValueError:
                logger.exception(
                    "Nao foi possivel resolver pasta EMP para o PDF. Arquivo mantido na EXT: %s",
                    arquivo_nome,
                )
                erros_processamento_notificacao.append(
                    f"{arquivo_nome} - empresa ou pasta nao encontrada; mantido na EXT"
                )
                continue

            dest_lancamento = temp_dir / dados_nome.nome_lancamento
            dest_excel = temp_dir / f"{arquivo_stem}.xlsx"

            inicio_geracao = time.perf_counter()
            logger.info("Gerando arquivos finais: %s", arquivo_nome)
            log_memoria(f"antes_gerar_excel_{arquivo_nome}")

            logger.info("Copiando modelo de lancamento para: %s", dest_lancamento)
            shutil.copy2(lancamento, dest_lancamento)

            logger.info("Preenchendo planilha de lancamento: %s", dest_lancamento)
            planilha_lancamento(df, dest_lancamento)

            logger.info("Gerando planilha totalizada: %s", dest_excel)
            df_totalizado = totalizador(df)
            df_totalizado.to_excel(dest_excel, index=False)
            del df_totalizado
            momento_conversao = agora_historico()
            log_memoria(f"depois_gerar_excel_{arquivo_nome}")

            logger.info("Enviando XLSM para o Google Drive: %s", dest_lancamento.name)
            enviar_ou_atualizar_arquivo(
                google_drive=google_drive,
                caminho_local=dest_lancamento,
                folder_id_destino=pasta_destino_id,
                type_file=XLSM_MIME_TYPE,
                name_drive=dest_lancamento.name,
            )

            logger.info("Enviando XLSX para o Google Drive: %s", dest_excel.name)
            enviar_ou_atualizar_arquivo(
                google_drive=google_drive,
                caminho_local=dest_excel,
                folder_id_destino=pasta_destino_id,
                type_file=XLSX_MIME_TYPE,
                name_drive=dest_excel.name,
            )
            logger.info(
                "Arquivos finais gerados e enviados: %s | tempo=%.2fs",
                arquivo_nome,
                time.perf_counter() - inicio_geracao,
            )

            logger.info("Movendo PDF original para pasta destino EMP: %s", arquivo_nome)
            google_drive.move_file(
                file_id=arquivo_id,
                folder_id_destino=pasta_destino_id,
            )
            momento_movimento = agora_historico()

            registrar_historico_conversao(
                google_drive=google_drive,
                pasta_raiz_id=extratos_id,
                temp_dir=temp_dir,
                nomes_arquivos=[arquivo_nome, dest_excel.name, dest_lancamento.name],
                pasta_destino=pasta_destino_historico,
                data_hora=momento_conversao,
                data_hora_movimento=momento_movimento,
                empresa_id=empresa_id,
                empresa_nome=empresa_nome,
            )

            documentos_convertidos.append({
                "empresa_id": empresa_id,
                "mes": dados_nome.mes,
                "ano": dados_nome.ano,
                "competencia": dados_nome.competencia,
                "codigo_documento": dados_nome.codigo_documento,
                "banco": dados_nome.banco,
                "agencia": dados_nome.agencia,
                "conta": dados_nome.conta,
                "arquivo_nome": arquivo_nome,
                "pasta_destino": pasta_destino_historico,
                "momento": momento_conversao,
            })

            documentos_para_baixa.append({
                "empresa_id": empresa_id,
                "mes": dados_nome.mes,
                "ano": dados_nome.ano,
                "competencia": dados_nome.competencia,
                "codigo_documento": dados_nome.codigo_documento,
                "banco": dados_nome.banco,
                "agencia": dados_nome.agencia,
                "conta": dados_nome.conta,
                "arquivo_nome": arquivo_nome,
                "pasta_destino": pasta_destino_historico,
                "quantidade_arquivos": 3,
                "status": "Enviado",
            })

            convertidos += 1
            extratos_notificacao.append(arquivo_stem)
            logger.info(
                "PDF processado com sucesso: %s | tempo_total=%.2fs",
                arquivo_nome,
                time.perf_counter() - inicio_pdf,
            )
            logger.info("PDF convertido com sucesso: %s", arquivo_nome)
            logger.info("Arquivos enviados para a pasta: %s", pasta_destino_historico)
        except Exception:
            logger.exception("Erro ao processar PDF: %s", arquivo_nome_seguro)
            erros_processamento_notificacao.append(
                f"{arquivo_nome_seguro} - erro inesperado; mantido na EXT"
            )
        finally:
            pdf_local.unlink(missing_ok=True)
            pdf_sem_senha.unlink(missing_ok=True)
            if dest_lancamento is not None:
                dest_lancamento.unlink(missing_ok=True)
            if dest_excel is not None:
                dest_excel.unlink(missing_ok=True)
            logger.info("Arquivos temporarios limpos para: %s", arquivo_nome_seguro)

            pdf = None
            df = None
            dest_lancamento = None
            dest_excel = None
            gc.collect()
            log_memoria(f"apos_gc_collect_{arquivo_nome_seguro}")

    logger.info("Atualizando controle no portal SGE para %s documento(s)", len(documentos_para_baixa))
    atualizados_sge = 0
    erros_sge = 0
    erros_sge_notificacao: list[str] = []
    sem_baixa_nao_cadastrada: list[str] = []

    for doc in documentos_para_baixa:
        try:
            id_existente = buscar_id_empresa_supabase(
                empresa_codigo=str(doc["empresa_id"]),
                competencia=doc["competencia"],
                codigo_documento=doc["codigo_documento"],
            )
            if not id_existente:
                logger.warning(
                    "Controle nao cadastrado no SGE: empresa=%s competencia=%s cod_doc=%s — ignorado",
                    doc["empresa_id"],
                    doc["competencia"],
                    doc["codigo_documento"],
                )
                sem_baixa_nao_cadastrada.append(
                    f"{doc['arquivo_nome']} — competencia {doc['competencia']}"
                )
                continue
            local = f"Google Drive / {doc['pasta_destino']}"
            sucesso, detalhe_erro = baixar_controle_supabase(
                empresa_codigo=str(doc["empresa_id"]),
                codigo_documento=doc["codigo_documento"],
                competencia=doc["competencia"],
                banco=doc["banco"],
                agencia=doc["agencia"],
                conta=doc["conta"],
                nome_arquivo=doc["arquivo_nome"],
                local_arquivo=local,
                quantidade_arquivos=doc["quantidade_arquivos"],
                status=doc["status"],
            )
            if sucesso:
                atualizados_sge += 1
            else:
                erros_sge += 1
                if detalhe_erro:
                    erros_sge_notificacao.append(detalhe_erro)
        except Exception:
            erros_sge += 1
            logger.exception("Erro ao atualizar controle no SGE: %s", doc["arquivo_nome"])

    notificacao_kwargs = {
        "pdfs_sem_movimentacao": pdfs_sem_movimentacao_notificacao,
        "pdfs_nao_legiveis": pdfs_nao_legiveis_notificacao,
        "atualizados_sge": atualizados_sge,
    }
    if nomes_invalidos_notificacao:
        notificacao_kwargs["nomes_invalidos"] = nomes_invalidos_notificacao
    if erros_senha_notificacao:
        notificacao_kwargs["erros_senha"] = erros_senha_notificacao
    if layouts_invalidos_notificacao:
        notificacao_kwargs["layouts_nao_reconhecidos"] = layouts_invalidos_notificacao
    if erros_processamento_notificacao:
        notificacao_kwargs["erros_processamento"] = erros_processamento_notificacao
    if erros_sge_notificacao:
        notificacao_kwargs["erros_sge_detalhe"] = erros_sge_notificacao
    if sem_baixa_nao_cadastrada:
        notificacao_kwargs["empresas_nao_cadastradas_sge"] = sem_baixa_nao_cadastrada

    possui_alertas = any(
        (
            invalidos,
            erros_sge,
            sem_baixa_nao_cadastrada,
            nomes_invalidos_notificacao,
            erros_senha_notificacao,
            layouts_invalidos_notificacao,
            erros_processamento_notificacao,
        )
    )
    notificacao_kwargs.update(
        {
            "google_drive": google_drive,
            "pasta_raiz_id": extratos_id,
            "execucao_id": execucao_id,
            "tipo": "SEM_ARQUIVOS" if not arquivos else "CONCLUSAO",
            "status_execucao": "SUCESSO COM ALERTAS" if possui_alertas else "SUCESSO",
            "total_processados": processados,
            "total_convertidos": convertidos,
            "total_sem_movimentacao": sem_movimentacao,
            "total_invalidos": invalidos,
            "total_desbloqueados": desbloqueados,
            "erros_sge": erros_sge,
        }
    )

    enviar_notificacao_google_chat(
        extratos_notificacao,
        **notificacao_kwargs,
    )

    logger.info(
        (
            "Fluxo de conversao concluido. "
            "Processados=%s Convertidos=%s SemMovimentacao=%s Invalidos=%s "
            "Desbloqueados=%s ErrosProcessamento=%s "
            "AtualizadosSGE=%s ErrosSGE=%s"
        ),
        processados,
        convertidos,
        sem_movimentacao,
        invalidos,
        desbloqueados,
        len(erros_processamento_notificacao),
        atualizados_sge,
        erros_sge,
    )


def executar_conversao():
    execucao_id = str(uuid4())
    try:
        return _executar_conversao(execucao_id)
    except Exception as error:
        logger.exception("Falha critica na execucao da conversao: %s", execucao_id)
        erro_seguro = re.sub(
            r"SENHA\s+[^\s_]+",
            "SENHA ***",
            str(error),
            flags=re.IGNORECASE,
        )[:1_000]
        google_drive = None
        try:
            google_drive = GoogleDriveAuth(
                credentials_path=GOOGLE_OAUTH_CREDENTIALS,
                token_path=GOOGLE_OAUTH_TOKEN,
                token_secret_path=GOOGLE_OAUTH_TOKEN_SECRET,
            )
        except Exception:
            logger.exception(
                "Google Drive indisponivel para persistir notificacao de falha: %s",
                execucao_id,
            )

        enviar_notificacao_google_chat(
            [],
            erros_processamento=[
                f"Falha critica: {type(error).__name__}: {erro_seguro}"
            ],
            google_drive=google_drive,
            pasta_raiz_id=GOOGLE_DRIVE_FOLDER_ID,
            execucao_id=execucao_id,
            tipo="FALHA",
            status_execucao="FALHA",
        )
        raise


def main():
    executar_conversao()


if __name__ == "__main__":
    main()
