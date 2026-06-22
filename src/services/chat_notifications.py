import logging
import os
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from threading import RLock
from typing import Callable
from uuid import uuid4

import requests
from openpyxl import Workbook, load_workbook

from src.app.gdrive.settings import (
    GOOGLE_DRIVE_FOLDER_ID,
    GOOGLE_OAUTH_CREDENTIALS,
    GOOGLE_OAUTH_TOKEN,
    GOOGLE_OAUTH_TOKEN_SECRET,
    XLSX_MIME_TYPE,
)


logger = logging.getLogger(__name__)

GOOGLE_CHAT_SEND_URL = os.getenv(
    "GOOGLE_CHAT_SEND_URL",
    "https://google-chat-api.onrender.com/google-chat/send",
)
GOOGLE_CHAT_SPACE_NAME = os.getenv(
    "GOOGLE_CHAT_SPACE_NAME",
    "spaces/AAQAQEHQc-k",
)
CHAT_OUTBOX_NAME = "CONTROLE_NOTIFICACOES_CHAT.xlsx"
LOCAL_FALLBACK_PATH = Path.cwd() / "logs" / "CONTROLE_NOTIFICACOES_CHAT_LOCAL.xlsx"
OUTBOX_HEADERS = (
    "ID",
    "EXECUCAO_ID",
    "PARTE",
    "TOTAL_PARTES",
    "CRIADA_EM",
    "STATUS",
    "TENTATIVAS",
    "ULTIMA_TENTATIVA",
    "ENVIADA_EM",
    "ULTIMO_ERRO",
    "TIPO",
    "MENSAGEM",
)
STATUS_PENDENTE = "PENDENTE"
STATUS_ENVIADA = "ENVIADA"
MAX_MESSAGE_CHARS = 3_500
IMMEDIATE_RETRY_DELAYS = (2, 5, 10)
OUTBOX_WORKER_INTERVAL_SECONDS = 300
OUTBOX_WORKER_BATCH_SIZE = 20

_outbox_lock = RLock()


@dataclass
class NotificationRecord:
    id: str
    execucao_id: str
    parte: int
    total_partes: int
    criada_em: str
    status: str
    tentativas: int
    ultima_tentativa: str
    enviada_em: str
    ultimo_erro: str
    tipo: str
    mensagem: str

    @classmethod
    def from_row(cls, row) -> "NotificationRecord":
        valores = list(row) + [None] * (len(OUTBOX_HEADERS) - len(row))
        return cls(
            id=str(valores[0] or ""),
            execucao_id=str(valores[1] or ""),
            parte=int(valores[2] or 0),
            total_partes=int(valores[3] or 0),
            criada_em=str(valores[4] or ""),
            status=str(valores[5] or STATUS_PENDENTE),
            tentativas=int(valores[6] or 0),
            ultima_tentativa=str(valores[7] or ""),
            enviada_em=str(valores[8] or ""),
            ultimo_erro=str(valores[9] or ""),
            tipo=str(valores[10] or "CONCLUSAO"),
            mensagem=str(valores[11] or ""),
        )

    def to_row(self) -> list[object]:
        return [
            self.id,
            self.execucao_id,
            self.parte,
            self.total_partes,
            self.criada_em,
            self.status,
            self.tentativas,
            self.ultima_tentativa,
            self.enviada_em,
            self.ultimo_erro,
            self.tipo,
            self.mensagem,
        ]


@dataclass(frozen=True)
class SendResult:
    success: bool
    attempts: int
    error: str = ""


def dividir_mensagem(mensagem: str, limite: int = MAX_MESSAGE_CHARS) -> list[str]:
    mensagem = mensagem.strip()
    if not mensagem:
        return ["Execucao concluida sem detalhes adicionais."]

    partes = []
    restante = mensagem
    while len(restante) > limite:
        corte = restante.rfind("\n", 0, limite + 1)
        if corte < limite // 2:
            corte = restante.rfind(" ", 0, limite + 1)
        if corte <= 0:
            corte = limite
        partes.append(restante[:corte].rstrip())
        restante = restante[corte:].lstrip()
    if restante:
        partes.append(restante)
    return partes


def criar_registros_notificacao(
    execucao_id: str,
    tipo: str,
    mensagem: str,
    momento: datetime | None = None,
) -> list[NotificationRecord]:
    momento_texto = (momento or datetime.now().astimezone()).isoformat(timespec="seconds")
    partes = dividir_mensagem(mensagem)
    total = len(partes)
    return [
        NotificationRecord(
            id=str(uuid4()),
            execucao_id=execucao_id,
            parte=indice,
            total_partes=total,
            criada_em=momento_texto,
            status=STATUS_PENDENTE,
            tentativas=0,
            ultima_tentativa="",
            enviada_em="",
            ultimo_erro="",
            tipo=tipo,
            mensagem=parte,
        )
        for indice, parte in enumerate(partes, start=1)
    ]


def _prepare_worksheet(workbook):
    worksheet = workbook.active
    worksheet.title = "notificacoes"
    primeira_linha = tuple(
        cell.value
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1), [])
    )
    if not any(primeira_linha):
        for column, value in enumerate(OUTBOX_HEADERS, start=1):
            worksheet.cell(row=1, column=column, value=value)
    elif primeira_linha != OUTBOX_HEADERS:
        raise ValueError("Cabecalho invalido no controle de notificacoes do Chat")
    return worksheet


def _records_from_worksheet(worksheet) -> list[NotificationRecord]:
    return [
        NotificationRecord.from_row(row)
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    ]


def _upsert_records(worksheet, records: list[NotificationRecord]) -> None:
    rows_by_id = {
        str(worksheet.cell(row=row, column=1).value): row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=1).value
    }
    for record in records:
        row_number = rows_by_id.get(record.id)
        if row_number is None:
            worksheet.append(record.to_row())
            rows_by_id[record.id] = worksheet.max_row
            continue
        for column, value in enumerate(record.to_row(), start=1):
            worksheet.cell(row=row_number, column=column, value=value)


def _with_drive_workbook(
    google_drive,
    pasta_raiz_id: str,
    callback: Callable,
    *,
    persist: bool,
):
    with _outbox_lock, tempfile.TemporaryDirectory() as temp_dir:
        local_path = Path(temp_dir) / CHAT_OUTBOX_NAME
        arquivo_drive = google_drive.find_file_by_name(
            folder_id=pasta_raiz_id,
            name=CHAT_OUTBOX_NAME,
            mime_type=XLSX_MIME_TYPE,
        )
        if arquivo_drive:
            google_drive.download(arquivo_drive["id"], local_path)
            workbook = load_workbook(local_path)
        else:
            workbook = Workbook()

        try:
            worksheet = _prepare_worksheet(workbook)
            result = callback(worksheet)
            if persist:
                workbook.save(local_path)
        finally:
            workbook.close()

        if persist:
            if arquivo_drive:
                google_drive.update_file(
                    file_id=arquivo_drive["id"],
                    caminho_local=local_path,
                    type_file=XLSX_MIME_TYPE,
                    name_drive=CHAT_OUTBOX_NAME,
                )
            else:
                google_drive.upload(
                    caminho_local=local_path,
                    folder_id_destino=pasta_raiz_id,
                    type_file=XLSX_MIME_TYPE,
                    name_drive=CHAT_OUTBOX_NAME,
                )
        return result


def persistir_registros_drive(google_drive, pasta_raiz_id: str, records) -> None:
    _with_drive_workbook(
        google_drive,
        pasta_raiz_id,
        lambda worksheet: _upsert_records(worksheet, records),
        persist=True,
    )


def listar_registros_drive(google_drive, pasta_raiz_id: str) -> list[NotificationRecord]:
    return _with_drive_workbook(
        google_drive,
        pasta_raiz_id,
        _records_from_worksheet,
        persist=False,
    )


def _load_local_workbook():
    LOCAL_FALLBACK_PATH.parent.mkdir(parents=True, exist_ok=True)
    if LOCAL_FALLBACK_PATH.exists():
        return load_workbook(LOCAL_FALLBACK_PATH)
    return Workbook()


def persistir_registros_local(records: list[NotificationRecord]) -> None:
    with _outbox_lock:
        workbook = _load_local_workbook()
        try:
            worksheet = _prepare_worksheet(workbook)
            _upsert_records(worksheet, records)
            temp_path = LOCAL_FALLBACK_PATH.with_suffix(".tmp.xlsx")
            workbook.save(temp_path)
        finally:
            workbook.close()
        os.replace(temp_path, LOCAL_FALLBACK_PATH)


def listar_registros_local() -> list[NotificationRecord]:
    with _outbox_lock:
        if not LOCAL_FALLBACK_PATH.exists():
            return []
        workbook = load_workbook(LOCAL_FALLBACK_PATH)
        try:
            return _records_from_worksheet(_prepare_worksheet(workbook))
        finally:
            workbook.close()


def sincronizar_fallback_local(google_drive, pasta_raiz_id: str) -> int:
    records = listar_registros_local()
    if not records:
        return 0
    persistir_registros_drive(google_drive, pasta_raiz_id, records)
    with _outbox_lock:
        LOCAL_FALLBACK_PATH.unlink(missing_ok=True)
    logger.info("Fallback local de notificacoes sincronizado: %s registro(s)", len(records))
    return len(records)


def tentar_enviar_mensagem(
    mensagem: str,
    *,
    max_attempts: int = 3,
    retry_delays: tuple[int, ...] = IMMEDIATE_RETRY_DELAYS,
) -> SendResult:
    ultimo_erro = ""
    for attempt in range(1, max_attempts + 1):
        try:
            response = requests.post(
                GOOGLE_CHAT_SEND_URL,
                json={"space_name": GOOGLE_CHAT_SPACE_NAME, "message": mensagem},
                timeout=30,
            )
            if 200 <= response.status_code < 300:
                return SendResult(True, attempt)
            ultimo_erro = f"HTTP {response.status_code}: {(response.text or '')[:500]}"
        except Exception as error:
            ultimo_erro = f"{type(error).__name__}: {str(error)[:500]}"

        logger.warning(
            "Tentativa %s/%s de notificacao ao Google Chat falhou: %s",
            attempt,
            max_attempts,
            ultimo_erro,
        )
        if attempt < max_attempts:
            delay = retry_delays[min(attempt - 1, len(retry_delays) - 1)]
            time.sleep(delay)

    return SendResult(False, max_attempts, ultimo_erro)


def _apply_send_result(record: NotificationRecord, result: SendResult) -> None:
    agora = datetime.now().astimezone().isoformat(timespec="seconds")
    record.tentativas += result.attempts
    record.ultima_tentativa = agora
    record.ultimo_erro = result.error
    if result.success:
        record.status = STATUS_ENVIADA
        record.enviada_em = agora
        record.ultimo_erro = ""


def registrar_e_enviar_notificacao(
    *,
    google_drive,
    pasta_raiz_id: str,
    execucao_id: str,
    tipo: str,
    mensagem: str,
) -> bool:
    records = criar_registros_notificacao(execucao_id, tipo, mensagem)
    persisted_on_drive = False
    try:
        if google_drive is not None:
            sincronizar_fallback_local(google_drive, pasta_raiz_id)
            persistir_registros_drive(google_drive, pasta_raiz_id, records)
            persisted_on_drive = True
        else:
            persistir_registros_local(records)
    except Exception:
        logger.exception("Falha ao persistir notificacao no Drive; usando fallback local")
        persistir_registros_local(records)

    for record in records:
        result = tentar_enviar_mensagem(record.mensagem)
        _apply_send_result(record, result)

    try:
        if persisted_on_drive:
            persistir_registros_drive(google_drive, pasta_raiz_id, records)
        else:
            persistir_registros_local(records)
    except Exception:
        logger.exception("Falha ao atualizar status da notificacao; preservando fallback local")
        persistir_registros_local(records)

    return all(record.status == STATUS_ENVIADA for record in records)


def _create_google_drive():
    from src.app.gdrive.google_drive_auth import GoogleDriveAuth

    return GoogleDriveAuth(
        credentials_path=GOOGLE_OAUTH_CREDENTIALS,
        token_path=GOOGLE_OAUTH_TOKEN,
        token_secret_path=GOOGLE_OAUTH_TOKEN_SECRET,
    )


def reprocessar_notificacoes_pendentes(
    google_drive=None,
    pasta_raiz_id: str = GOOGLE_DRIVE_FOLDER_ID,
    limite: int = OUTBOX_WORKER_BATCH_SIZE,
) -> dict[str, int]:
    resultados = {"pendentes": 0, "enviadas": 0, "falhas": 0}
    try:
        google_drive = google_drive or _create_google_drive()
        sincronizar_fallback_local(google_drive, pasta_raiz_id)
        records = listar_registros_drive(google_drive, pasta_raiz_id)
        pendentes = sorted(
            (record for record in records if record.status == STATUS_PENDENTE),
            key=lambda record: (record.criada_em, record.parte),
        )[:limite]
        resultados["pendentes"] = len(pendentes)

        for record in pendentes:
            result = tentar_enviar_mensagem(record.mensagem, max_attempts=1)
            _apply_send_result(record, result)
            if result.success:
                resultados["enviadas"] += 1
            else:
                resultados["falhas"] += 1

        if pendentes:
            persistir_registros_drive(google_drive, pasta_raiz_id, pendentes)
    except Exception:
        logger.exception("Erro ao reprocessar fila de notificacoes do Google Chat")
        try:
            records = [
                record
                for record in listar_registros_local()
                if record.status == STATUS_PENDENTE
            ][:limite]
            resultados["pendentes"] = len(records)
            for record in records:
                result = tentar_enviar_mensagem(record.mensagem, max_attempts=1)
                _apply_send_result(record, result)
                resultados["enviadas" if result.success else "falhas"] += 1
            if records:
                persistir_registros_local(records)
        except Exception:
            logger.exception("Erro ao reprocessar fallback local de notificacoes")
    return resultados
