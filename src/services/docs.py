import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.app.gdrive.google_drive_auth import GoogleDriveAuth
from src.app.gdrive.settings import (
    GOOGLE_DRIVE_EMP_FOLDER_ID,
    GOOGLE_DRIVE_FOLDER_ID,
    GOOGLE_OAUTH_CREDENTIALS,
    GOOGLE_OAUTH_TOKEN,
    GOOGLE_OAUTH_TOKEN_SECRET,
    XLSX_MIME_TYPE,
)
from src.app.supabase.supabase_api import (
    baixar_controle_supabase,
    carregar_caminhos_documentos_api,
)
from src.services.chat_notifications import registrar_e_enviar_notificacao
from src.services.conversao import (
    agora_historico,
    buscar_empresa_por_cliente,
    carregar_empresas_ativas,
    nome_pasta_empresa,
    normalizar_id_empresa,
    normalizar_nome_empresa,
    resolver_pasta_emp_base,
)
from src.utils.logging_config import setup_logging


logger = logging.getLogger(__name__)

DOCS_FOLDER_NAME = "DOCS"
HISTORICO_DOCS = "HISTORICO_DOCS.xlsx"

# Alias de codigos: grafia que aparece no NOME DO ARQUIVO -> grafia canonica
# usada na planilha DOCS E CAMINHO.xlsx e no codigo_documento da baixa do SGE.
ALIAS_CODIGO = {
    "EXTBANAPL": "EXTAPL",
    "FATCAR": "FATCART",
    "EXTMAQCART": "EXTMAQCAR",
}

# Codigos cujo nome traz um segmento de BANCO (e, opcionalmente, AG/CC).
# Para os demais, o segmento apos o codigo ja e o cliente.
CODIGOS_COM_BANCO = {
    "EXTBAN",
    "EXTAPL",
    "EXTCOTCAP",
    "EXTMAQCAR",
    "ENCCTBANC",
    "FATCART",
}
HISTORICO_DOCS_HEADERS = (
    "ID",
    "EMP",
    "DATA",
    "HORA",
    "CODIGO",
    "NOME ARQUIVO",
    "PASTA DESTINO",
    "DATA HORA MOVIMENTO",
    "STATUS SGE",
)


def normalizar_codigo(valor) -> str:
    if valor is None:
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor)).strip().upper()

    return str(valor).strip().upper()


def parse_nome_documento(nome_arquivo: str) -> dict[str, str]:
    caminho = Path(nome_arquivo)
    partes = [parte.strip() for parte in caminho.stem.split("_") if parte.strip()]

    if len(partes) < 3:
        raise ValueError(f"Nome do arquivo DOCS sem padrao MMAA_CODIGO_CLIENTE: {nome_arquivo}")

    periodo = partes[0]
    if not re.fullmatch(r"\d{4}", periodo):
        raise ValueError(f"Nome do arquivo DOCS sem periodo MMAA valido: {nome_arquivo}")

    # O segmento do codigo pode trazer um subtipo colado, separado por espaco
    # (ex.: "EXTAPL FIC GIRO", "EXTAPL CDB", "CONTEMP PRONAMP"). O codigo e a
    # primeira palavra; o resto e subtipo (informativo). Aplica alias para a
    # grafia canonica (ex.: EXTBANAPL -> EXTAPL, FATCAR -> FATCART).
    codigo_segmento = normalizar_codigo(partes[1])
    codigo_base = codigo_segmento.split(" ", 1)[0]
    codigo = ALIAS_CODIGO.get(codigo_base, codigo_base)

    banco = ""
    agencia = ""
    conta = ""
    cliente = partes[-1]

    if codigo in CODIGOS_COM_BANCO and len(partes) > 3:
        banco = partes[2]
        partes_resto = []
        for p in partes[3:]:
            if p.upper().startswith("AG "):
                agencia = p[3:].strip()
            elif p.upper().startswith("CC "):
                conta = p[3:].strip()
            else:
                partes_resto.append(p)
        # Documento bancario sem segmento de cliente (so banco + AG/CC): nao
        # usar o "CC ..." como cliente; deixar vazio para virar pendencia.
        cliente = partes_resto[-1] if partes_resto else ""

    return {
        "mes": periodo[:2],
        "ano": periodo[-2:],
        "codigo": codigo,
        "cliente": cliente,
        "banco": banco,
        "agencia": agencia,
        "conta": conta,
    }


# Pastas raiz de destino validas dentro de SRVARQ no Google Drive.
RAIZES_DESTINO = {"EMP", "PUBLICO"}
# Alias de raiz: grafia legada da planilha -> nome real da pasta no Drive.
ALIAS_RAIZ = {"PUB": "PUBLICO"}


def montar_destino_docs(
    destino_template: str,
    empresa_chave: str,
    ano: str,
    mes: str,
) -> tuple[str, list[str]]:
    """Resolve o template de destino em (raiz, partes relativas a raiz).

    A raiz e a pasta de topo dentro de SRVARQ ("EMP" ou "PUBLICO"); sem
    prefixo reconhecido, assume "EMP" (ex.: template iniciando em {EMPRESA}).
    """
    destino = (
        destino_template
        .replace("{EMPRESA}", empresa_chave)
        .replace("{ANO}", ano)
        .replace("{MES}", mes)
        .replace("{MÊS}", mes)
    )
    destino = destino.replace("\\", "/").strip("/")
    partes = [parte for parte in destino.split("/") if parte]

    if partes and partes[0].upper() == "SRVARQ":
        partes = partes[1:]

    raiz = "EMP"
    if partes:
        cabeca = ALIAS_RAIZ.get(partes[0].upper(), partes[0].upper())
        if cabeca in RAIZES_DESTINO:
            raiz = cabeca
            partes = partes[1:]

    return raiz, partes


def resolver_raiz_destino(
    google_drive: GoogleDriveAuth,
    raizes_ids: dict[str, str],
    raiz_nome: str,
) -> str:
    """Retorna o folder_id da raiz de destino, resolvendo sob demanda.

    ``raizes_ids`` funciona como cache mutavel {nome_raiz: folder_id}; raizes
    ainda nao resolvidas (ex.: PUBLICO) sao localizadas dentro de SRVARQ.
    """
    if raiz_nome not in raizes_ids:
        raizes_ids[raiz_nome] = resolver_pasta_emp_base(
            google_drive=google_drive,
            pasta_base_id=GOOGLE_DRIVE_EMP_FOLDER_ID,
            nome_pasta=raiz_nome,
        )
    return raizes_ids[raiz_nome]


def resolver_pasta_destino_docs(
    google_drive: GoogleDriveAuth,
    raiz_folder_id: str,
    destino_partes: list[str],
    raiz_nome: str = "EMP",
) -> tuple[str, str]:
    pasta_atual_id = raiz_folder_id

    for i, nome_pasta in enumerate(destino_partes):
        if i == 0:
            pasta = google_drive.list_folder_by_name(
                folder_id=pasta_atual_id,
                name_folder=nome_pasta,
            )
            if not pasta:
                raise ValueError(
                    f"Pasta de cliente '{nome_pasta}' nao encontrada em {raiz_nome}. "
                    "A pasta deve existir previamente no Google Drive."
                )
        else:
            pasta = google_drive.get_or_create_folder(
                folder_id_pai=pasta_atual_id,
                name_folder=nome_pasta,
            )
        pasta_atual_id = pasta["id"]

    return pasta_atual_id, "/".join([raiz_nome, *destino_partes])


def registrar_historico_docs(
    google_drive: GoogleDriveAuth,
    pasta_raiz_id: str,
    temp_dir: Path,
    registros: list[dict[str, object]],
) -> None:
    if not registros:
        logger.info("Historico DOCS nao atualizado: nenhum documento movido")
        return

    temp_dir.mkdir(parents=True, exist_ok=True)
    historico_local = temp_dir / HISTORICO_DOCS
    arquivo_historico = google_drive.find_file_by_name(
        folder_id=pasta_raiz_id,
        name=HISTORICO_DOCS,
        mime_type=XLSX_MIME_TYPE,
    )

    if arquivo_historico:
        logger.info("Baixando historico DOCS existente")
        google_drive.download(
            file_id=arquivo_historico["id"],
            destino_local=historico_local,
        )
        workbook = load_workbook(historico_local)
        worksheet = workbook.active
        headers = tuple(cell.value for cell in worksheet[1]) if worksheet.max_row else ()

        if headers != HISTORICO_DOCS_HEADERS:
            linhas_antigas = list(worksheet.iter_rows(min_row=2, values_only=True))
            worksheet.delete_rows(1, worksheet.max_row)
            worksheet.append(HISTORICO_DOCS_HEADERS)
            for linha in linhas_antigas:
                valores = list(linha or [])
                worksheet.append((valores + [""] * len(HISTORICO_DOCS_HEADERS))[:len(HISTORICO_DOCS_HEADERS)])
    else:
        logger.info("Criando historico DOCS")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "historico"
        worksheet.append(HISTORICO_DOCS_HEADERS)

    for registro in registros:
        worksheet.append(
            [
                registro["empresa_id"],
                registro["empresa_nome"],
                registro["data"],
                registro["hora"],
                registro["codigo"],
                registro["nome_arquivo"],
                registro["pasta_destino"],
                registro["data_hora_movimento"],
                registro.get("status_sge", ""),
            ]
        )

    workbook.save(historico_local)

    if arquivo_historico:
        google_drive.update_file(
            file_id=arquivo_historico["id"],
            caminho_local=historico_local,
            type_file=XLSX_MIME_TYPE,
            name_drive=HISTORICO_DOCS,
        )
        logger.info("Historico DOCS atualizado: %s", HISTORICO_DOCS)
    else:
        google_drive.upload(
            caminho_local=historico_local,
            folder_id_destino=pasta_raiz_id,
            type_file=XLSX_MIME_TYPE,
            name_drive=HISTORICO_DOCS,
        )
        logger.info("Historico DOCS criado: %s", HISTORICO_DOCS)

    historico_local.unlink(missing_ok=True)


def formatar_mensagem_docs_google_chat(
    nomes_documentos: list[str],
    momento: datetime | None = None,
    atualizados_sge: int | None = None,
    empresas_nao_cadastradas_sge: list[str] | None = None,
    erros_sge: int = 0,
    erros_sge_detalhe: list[str] | None = None,
    pendencias: list[str] | None = None,
) -> str:
    momento = momento or agora_historico()
    data = momento.strftime("%d/%m/%y")
    hora = momento.strftime("%H:%M")
    pendencias = pendencias or []
    blocos = []

    if nomes_documentos:
        lista_documentos = "\n".join(nomes_documentos)
        blocos.append(
            f"Documentos salvos {data} as {hora} e atualizado na Base de Dados:\n\n"
            f"{lista_documentos}"
        )

    if pendencias:
        lista_pendencias = "\n".join(pendencias)
        blocos.append(
            "Documentos pendentes em DOCS (nao movidos) - revisar:\n\n"
            f"{lista_pendencias}"
        )

    if atualizados_sge is not None and atualizados_sge > 0:
        blocos.append(
            f"Baixa dada no portal SGE: {atualizados_sge} documento(s) atualizado(s)"
        )

    if erros_sge > 0:
        blocos.append(f"Erros ao dar baixa no SGE: {erros_sge}")

    if erros_sge_detalhe:
        lista_detalhes = "\n".join(erros_sge_detalhe)
        blocos.append(f"Detalhes dos erros SGE:\n\n{lista_detalhes}")

    if empresas_nao_cadastradas_sge:
        lista = "\n".join(empresas_nao_cadastradas_sge)
        blocos.append(
            "Documentos sem baixa no SGE (empresa nao cadastrada):\n\n"
            f"{lista}"
        )

    return "\n\n".join(blocos)


def enviar_notificacao_docs_google_chat(
    nomes_documentos: list[str],
    momento: datetime | None = None,
    atualizados_sge: int | None = None,
    empresas_nao_cadastradas_sge: list[str] | None = None,
    erros_sge: int = 0,
    erros_sge_detalhe: list[str] | None = None,
    pendencias: list[str] | None = None,
    google_drive=None,
    pasta_raiz_id: str = "",
    execucao_id: str = "",
) -> bool:
    pendencias = pendencias or []
    if not nomes_documentos and not empresas_nao_cadastradas_sge and not pendencias:
        logger.info("Notificacao Google Chat DOCS nao enviada: nenhum documento movido nem pendencia")
        return False

    mensagem = formatar_mensagem_docs_google_chat(
        nomes_documentos,
        momento=momento,
        atualizados_sge=atualizados_sge,
        empresas_nao_cadastradas_sge=empresas_nao_cadastradas_sge,
        erros_sge=erros_sge,
        erros_sge_detalhe=erros_sge_detalhe,
        pendencias=pendencias,
    )

    return registrar_e_enviar_notificacao(
        google_drive=google_drive,
        pasta_raiz_id=pasta_raiz_id,
        execucao_id=execucao_id,
        tipo="DOCS",
        mensagem=mensagem,
    )


def _mime_type_historico(arquivo: dict) -> bool:
    return arquivo.get("name") == HISTORICO_DOCS


def _contar_arquivos_pasta(google_drive: GoogleDriveAuth, pasta_id: str) -> int:
    """Conta arquivos (nao-pastas) diretamente dentro de uma pasta do Drive."""
    filhos = google_drive.list_children(pasta_id)
    return sum(
        1
        for f in filhos
        if f.get("mimeType") != GoogleDriveAuth.FOLDER_MIME_TYPE
        and f.get("name", "").lower() != "desktop.ini"
    )


def _processar_item_docs(
    google_drive: GoogleDriveAuth,
    item: dict,
    is_pasta: bool,
    caminhos: dict[str, str],
    empresas,
    raizes_ids: dict[str, str],
    registros_historico: list[dict[str, object]],
    documentos_para_baixa: list[dict[str, object]],
    documentos_notificacao: list[str],
    pendencias_notificacao: list[str],
) -> str:
    """Processa um arquivo OU pasta DOCS.

    Identifica o codigo pelo nome (MMAA_CODIGO_CLIENTE), resolve o destino na
    planilha de caminhos, move para a pasta da empresa e prepara a baixa no SGE.
    Para pastas, a quantidade de arquivos e a contagem de itens dentro dela.

    Quando o item nao pode ser tratado (nome fora do padrao, codigo nao
    cadastrado, empresa nao encontrada ou erro), registra o motivo em
    ``pendencias_notificacao`` para aparecer na notificacao do Chat.

    Retorna 'movido', 'ignorado' ou 'erro'.
    """
    nome = item["name"]
    logger.info("Processando %s DOCS: %s", "pasta" if is_pasta else "documento", nome)

    try:
        dados_nome = parse_nome_documento(nome)
    except ValueError as erro:
        logger.warning("%s", erro)
        pendencias_notificacao.append(f"{nome} - nome fora do padrao MMAA_CODIGO_CLIENTE")
        return "ignorado"

    codigo = dados_nome["codigo"]
    destino_template = caminhos.get(codigo)
    if not destino_template:
        logger.warning("Codigo DOCS nao cadastrado no SGE (Documentos): %s", codigo)
        pendencias_notificacao.append(
            f"{nome} - codigo '{codigo}' nao cadastrado em Documentos no portal SGE"
        )
        return "ignorado"

    cliente = dados_nome["cliente"]
    if not cliente:
        logger.warning("Documento DOCS sem cliente identificavel no nome: %s", nome)
        pendencias_notificacao.append(f"{nome} - documento sem cliente identificavel no nome")
        return "ignorado"

    empresa_id, empresa_nome = buscar_empresa_por_cliente(cliente, empresas=empresas)
    if not empresa_id or not empresa_nome:
        logger.warning(
            "Empresa DOCS nao encontrada para cliente %s. Mantido em DOCS: %s",
            cliente,
            nome,
        )
        pendencias_notificacao.append(
            f"{nome} - empresa/cliente '{cliente}' nao encontrada"
        )
        return "ignorado"

    try:
        quantidade = _contar_arquivos_pasta(google_drive, item["id"]) if is_pasta else 1

        empresa_chave = nome_pasta_empresa(empresa_id, empresa_nome)
        raiz_nome, destino_partes = montar_destino_docs(
            destino_template=destino_template,
            empresa_chave=empresa_chave,
            ano=dados_nome["ano"],
            mes=dados_nome["mes"],
        )
        raiz_folder_id = resolver_raiz_destino(google_drive, raizes_ids, raiz_nome)
        pasta_destino_id, pasta_destino_historico = resolver_pasta_destino_docs(
            google_drive=google_drive,
            raiz_folder_id=raiz_folder_id,
            destino_partes=destino_partes,
            raiz_nome=raiz_nome,
        )

        logger.info(
            "Movendo %s DOCS para: %s",
            "pasta" if is_pasta else "documento",
            pasta_destino_historico,
        )
        google_drive.move_file(file_id=item["id"], folder_id_destino=pasta_destino_id)
        momento_movimento = agora_historico()
        registros_historico.append(
            {
                "empresa_id": normalizar_id_empresa(empresa_id),
                "empresa_nome": normalizar_nome_empresa(empresa_nome),
                "data": momento_movimento.strftime("%Y-%m-%d"),
                "hora": momento_movimento.strftime("%H:%M:%S"),
                "codigo": codigo,
                "nome_arquivo": nome,
                "pasta_destino": pasta_destino_historico,
                "data_hora_movimento": momento_movimento.strftime("%Y-%m-%d %H:%M:%S"),
                "status_sge": "",
            }
        )
        documentos_para_baixa.append(
            {
                "empresa_id": empresa_id,
                "competencia": f'{dados_nome["mes"]}-20{dados_nome["ano"]}',
                "codigo_documento": dados_nome["codigo"],
                "banco": dados_nome["banco"],
                "agencia": dados_nome["agencia"],
                "conta": dados_nome["conta"],
                "arquivo_nome": nome,
                "pasta_destino": pasta_destino_historico,
                "quantidade_arquivos": quantidade,
                "indice_historico": len(registros_historico) - 1,
                "data_recebimento": momento_movimento.strftime("%Y-%m-%d"),
            }
        )
        documentos_notificacao.append(f"{nome} - {pasta_destino_historico}")
        return "movido"
    except Exception:
        logger.exception("Erro ao mover documento DOCS: %s", nome)
        pendencias_notificacao.append(f"{nome} - erro ao mover/processar")
        return "erro"


def executar_docs() -> dict[str, int]:
    setup_logging()
    execucao_id = f"DOCS-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    logger.info("Iniciando fluxo DOCS [execucao=%s]", execucao_id)
    logger.info("Autenticando Google Drive")

    google_drive = GoogleDriveAuth(
        credentials_path=GOOGLE_OAUTH_CREDENTIALS,
        token_path=GOOGLE_OAUTH_TOKEN,
        token_secret_path=GOOGLE_OAUTH_TOKEN_SECRET,
    )

    raiz_id = GOOGLE_DRIVE_FOLDER_ID
    temp_dir = Path.cwd() / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    logger.info("Diretorio temporario preparado: %s", temp_dir)
    logger.info("Pasta raiz operacional DOCS: %s", raiz_id)

    docs_folder = google_drive.get_or_create_folder(
        folder_id_pai=raiz_id,
        name_folder=DOCS_FOLDER_NAME,
    )
    docs_folder_id = docs_folder["id"]
    logger.info("Pasta DOCS localizada: %s", docs_folder_id)

    logger.info("Localizando pasta EMP dentro da pasta base SRVARQ para DOCS")
    # Cache de raizes de destino (EMP resolvida ja; PUBLICO sob demanda).
    raizes_ids: dict[str, str] = {
        "EMP": resolver_pasta_emp_base(
            google_drive=google_drive,
            pasta_base_id=GOOGLE_DRIVE_EMP_FOLDER_ID,
        )
    }

    try:
        caminhos = carregar_caminhos_documentos_api()
    except Exception:
        logger.exception("Falha ao carregar destino final dos documentos no SGE")
        return {"processados": 0, "movidos": 0, "ignorados": 0, "erros": 1}

    if not caminhos:
        logger.warning("Nenhum documento ativo com destino final cadastrado no SGE")
        return {"processados": 0, "movidos": 0, "ignorados": 0, "erros": 1}

    empresas = carregar_empresas_ativas()
    filhos = google_drive.list_children(docs_folder_id)
    arquivos = [
        f
        for f in filhos
        if f.get("mimeType") != GoogleDriveAuth.FOLDER_MIME_TYPE
        and f.get("name", "").lower() != "desktop.ini"
        and not _mime_type_historico(f)
    ]
    pastas = [
        f for f in filhos if f.get("mimeType") == GoogleDriveAuth.FOLDER_MIME_TYPE
    ]
    logger.info(
        "Itens na pasta DOCS: %s arquivo(s) | %s subpasta(s)", len(arquivos), len(pastas)
    )

    processados = 0
    movidos = 0
    ignorados = 0
    erros = 0
    registros_historico: list[dict[str, object]] = []
    documentos_notificacao: list[str] = []
    documentos_para_baixa: list[dict[str, object]] = []
    pendencias_notificacao: list[str] = []

    itens = [(a, False) for a in arquivos] + [(p, True) for p in pastas]
    for item, is_pasta in itens:
        processados += 1
        resultado = _processar_item_docs(
            google_drive=google_drive,
            item=item,
            is_pasta=is_pasta,
            caminhos=caminhos,
            empresas=empresas,
            raizes_ids=raizes_ids,
            registros_historico=registros_historico,
            documentos_para_baixa=documentos_para_baixa,
            documentos_notificacao=documentos_notificacao,
            pendencias_notificacao=pendencias_notificacao,
        )
        if resultado == "movido":
            movidos += 1
        elif resultado == "ignorado":
            ignorados += 1
        else:
            erros += 1

    logger.info("Atualizando controle no portal SGE para %s documento(s)", len(documentos_para_baixa))
    atualizados_sge = 0
    erros_sge = 0
    erros_sge_notificacao: list[str] = []
    sem_baixa_nao_cadastrada: list[str] = []

    for doc in documentos_para_baixa:
        try:
            local = f'Google Drive / {doc["pasta_destino"]}'
            sucesso, detalhe_erro = baixar_controle_supabase(
                empresa_codigo=str(doc["empresa_id"]),
                codigo_documento=doc["codigo_documento"],
                competencia=doc["competencia"],
                banco=doc["banco"],
                agencia=doc["agencia"],
                conta=doc["conta"],
                nome_arquivo=doc["arquivo_nome"],
                local_arquivo=local,
                quantidade_arquivos=doc.get("quantidade_arquivos", 1),
                status="Enviado",
                data_recebimento=doc["data_recebimento"],
            )
            if sucesso:
                atualizados_sge += 1
                registros_historico[doc["indice_historico"]]["status_sge"] = "Baixado"
            else:
                erros_sge += 1
                registros_historico[doc["indice_historico"]]["status_sge"] = "Erro"
                if detalhe_erro:
                    erros_sge_notificacao.append(detalhe_erro)
        except Exception:
            erros_sge += 1
            registros_historico[doc["indice_historico"]]["status_sge"] = "Erro"
            logger.exception("Erro ao atualizar controle no SGE: %s", doc["arquivo_nome"])

    registrar_historico_docs(
        google_drive=google_drive,
        pasta_raiz_id=raiz_id,
        temp_dir=temp_dir,
        registros=registros_historico,
    )
    enviar_notificacao_docs_google_chat(
        documentos_notificacao,
        google_drive=google_drive,
        pasta_raiz_id=raiz_id,
        execucao_id=execucao_id,
        atualizados_sge=atualizados_sge,
        empresas_nao_cadastradas_sge=sem_baixa_nao_cadastrada,
        erros_sge=erros_sge,
        erros_sge_detalhe=erros_sge_notificacao,
        pendencias=pendencias_notificacao,
    )

    logger.info(
        "Fluxo DOCS concluido. Processados=%s Movidos=%s Ignorados=%s Erros=%s "
        "Pendencias=%s AtualizadosSGE=%s ErrosSGE=%s",
        processados,
        movidos,
        ignorados,
        erros,
        len(pendencias_notificacao),
        atualizados_sge,
        erros_sge,
    )

    return {
        "processados": processados,
        "movidos": movidos,
        "ignorados": ignorados,
        "erros": erros,
        "pendencias": len(pendencias_notificacao),
        "atualizados_sge": atualizados_sge,
        "erros_sge": erros_sge,
    }


def main():
    executar_docs()


if __name__ == "__main__":
    main()
