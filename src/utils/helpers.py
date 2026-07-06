import os
import logging
import shutil
import pandas as pd
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date
from pypdf import PdfReader, PdfWriter

logger = logging.getLogger(__name__)


def planilha_lancamento(df, destino):
    destino = Path(destino)
    destino_tmp = destino.with_name(f"{destino.stem}_tmp{destino.suffix}")

    shutil.copy2(destino, destino_tmp)

    wb = None

    try:
        wb = load_workbook(destino_tmp, keep_vba=True)
        ws = wb["Plan1"]

        linha_inicial = 2

        for posicao, (_, row) in enumerate(df.iterrows()):
            linha_excel = linha_inicial + posicao

            # DATA -> coluna A
            celula_data = ws[f"A{linha_excel}"]
            celula_data.number_format = "@"
            celula_data.value = str(row["DATA"])

            # VALOR -> coluna D
            celula_valor = ws[f"D{linha_excel}"]
            celula_valor.value = abs(row["VALOR"])

            fonte = copy(celula_valor.font)
            tipo = str(row["TIPO"]).upper().strip()

            if tipo == "D":
                fonte.color = "FFFF0000"  # vermelho
            else:
                fonte.color = "FF000000"  # preto

            celula_valor.font = fonte

            # DESCRIÇÃO -> coluna F
            ws[f"F{linha_excel}"] = row["DESCRIÇÃO"]

        wb.save(destino_tmp)

    finally:
        if wb is not None:
            if hasattr(wb, "vba_archive") and wb.vba_archive is not None:
                try:
                    wb.vba_archive.close()
                except Exception:
                    pass

                try:
                    wb.vba_archive = None
                except Exception:
                    pass

            try:
                wb.close()
            except Exception:
                pass

            del wb

    os.replace(destino_tmp, destino)

    print(f"Arquivo preenchido com sucesso: {destino}")

def totalizador(df):
    df = df.reset_index(drop=True)
    saida = df[df["TIPO"] == "D"]["VALOR"].sum()
    entrada = df[df["TIPO"] == "C"]["VALOR"].sum()
    df.loc[len(df)] = {
        "DATA": "",
        "DESCRIÇÃO": "TOTAL ENTRADAS",
        "VALOR": entrada,
        "TIPO": "C"
    }

    df.loc[len(df)] = {
        "DATA": "",
        "DESCRIÇÃO": "TOTAL SAÍDAS",
        "VALOR": saida,
        "TIPO": "D"
    }
    return df

def remover_senha_pdf(caminho_pdf: str, senha: str, caminho_saida: str | None = None):
    entrada = Path(caminho_pdf)

    if caminho_saida is None:
        caminho_saida = entrada.with_name(f"{entrada.stem}_sem_senha.pdf")

    reader = PdfReader(str(entrada))

    if reader.is_encrypted:
        resultado = reader.decrypt(senha)

        if resultado == 0:
            raise ValueError("Senha inválida ou não foi possível descriptografar o PDF.")

    writer = PdfWriter()

    for page in reader.pages:
        writer.add_page(page)

    with open(caminho_saida, "wb") as arquivo_saida:
        writer.write(arquivo_saida)

    return caminho_saida

def processar_extrato_banco(xls, aba):
    """Extrai lançamentos de uma aba de extrato bancário no layout FECFIN (*PART).

    O cabeçalho e as colunas de observação disponíveis variam entre bancos
    e até entre contas do mesmo banco (ex: "SAÍDA" vs "SAIDA", "OBS" vs
    "OBS INTERNA" vs "OBS INT" vs "OBS CONTABILIDADE"). Em vez de assumir
    uma posição fixa de cabeçalho e nomes fixos de coluna, esta função
    localiza o cabeçalho pelo conteúdo e usa apenas as colunas de
    observação que realmente existem naquela aba.

    Retorna um DataFrame vazio (colunas DATA/DESCRIÇÃO/VALOR/TIPO) quando
    a aba não tem um layout reconhecível ou não tem lançamentos reais
    (ex: abas-modelo só com SALDO ANTERIOR/FINAL).
    """
    vazio = pd.DataFrame(columns=["DATA", "DESCRIÇÃO", "VALOR", "TIPO"])
    bruto = pd.read_excel(xls, sheet_name=aba, header=None)

    linha_cabecalho = None

    for i, row in bruto.iterrows():
        valores = [str(v).strip().upper() for v in row if pd.notna(v)]
        if "DATA" in valores and any("HIST" in v for v in valores):
            linha_cabecalho = i
            break

    if linha_cabecalho is None:
        return vazio

    df = bruto[linha_cabecalho + 1:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in bruto.iloc[linha_cabecalho]]

    colunas_desc = ["HISTÓRICO", "TIPO", "DOC", "Nº DOC", "OBS", "OBS INTERNA",
                    "OBS INT", "OBS P/ INTERNAS", "OBS CONTABILIDADE", "DADOS BANCÁRIOS"]
    presentes = [c for c in colunas_desc if c in df.columns]

    if "DATA" not in df.columns or not presentes:
        return vazio

    col_entrada = next((c for c in df.columns if str(c).strip().upper() == "ENTRADA"), None)
    col_saida = next((c for c in df.columns if str(c).strip().upper() in ("SAÍDA", "SAIDA")), None)

    if col_entrada is None and col_saida is None:
        return vazio

    entrada = pd.to_numeric(df[col_entrada], errors="coerce").fillna(0) if col_entrada else 0
    saida = pd.to_numeric(df[col_saida], errors="coerce").fillna(0) if col_saida else 0

    df["DESCRIÇÃO"] = df[presentes].fillna("").astype(str).agg(" ".join, axis=1)
    df["DESCRIÇÃO"] = df["DESCRIÇÃO"].str.replace("nan", "", regex=False).str.strip().str.upper()

    df["VALOR"] = entrada.where(entrada != 0, saida * -1) if col_saida is not None else entrada
    df = df.loc[~df["DESCRIÇÃO"].str.contains("SALDO", na=False)]
    df = df.loc[df["DESCRIÇÃO"] != ""]  # descarta linhas em branco (sobra de formatação/linhas de modelo)

    if df.empty:
        return vazio

    df["TIPO"] = df["VALOR"].apply(lambda v: "C" if v > 0 else "D")
    df["VALOR"] = df["VALOR"].abs()
    df["DATA"] = df["DATA"].apply(lambda x: pd.to_datetime(x, errors="coerce").strftime("%d/%m/%Y") if pd.notnull(x) else "")

    return df[["DATA", "DESCRIÇÃO", "VALOR", "TIPO"]]

def normalizar_parc(valor):
    if pd.isna(valor):
        return ""

    # Quando o Excel transformou 01/02 em data
    if isinstance(valor, (pd.Timestamp, datetime, date)):
        return f"{valor.day:02}/{valor.month:02}"

    # Quando vem número 1.0, 2.0 etc.
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip()

def corrigir_xls_html_para_xlsx(
    caminho_xls: str | Path,
    caminho_xlsx: str | Path | None = None,
    deletar_original: bool = True
) -> Path:
    """
    Converte um arquivo .xls/.html que na verdade é HTML para .xlsx válido.

    Se deletar_original=True, remove o arquivo original somente depois
    que o .xlsx for criado com sucesso.

    Retorna:
        Path do arquivo .xlsx gerado.
    """

    caminho_xls = Path(caminho_xls)

    if not caminho_xls.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_xls}")

    if caminho_xlsx is None:
        caminho_xlsx = caminho_xls.with_suffix(".xlsx")
    else:
        caminho_xlsx = Path(caminho_xlsx)

        # Se passou apenas o nome do arquivo, salva na mesma pasta do .xls
        if not caminho_xlsx.is_absolute():
            caminho_xlsx = caminho_xls.parent / caminho_xlsx

    # Lê tabelas HTML dentro do arquivo .xls
    tabelas = pd.read_html(caminho_xls)

    if not tabelas:
        raise ValueError(f"Nenhuma tabela HTML encontrada em: {caminho_xls}")

    df = tabelas[0]

    # Remove linhas e colunas totalmente vazias
    df = df.dropna(how="all").dropna(axis=1, how="all")

    # Garante que a pasta de destino exista
    caminho_xlsx.parent.mkdir(parents=True, exist_ok=True)

    # Salva como .xlsx real
    with pd.ExcelWriter(caminho_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Planilha1")

    # Só apaga o original se o .xlsx realmente foi criado
    if deletar_original and caminho_xlsx.exists() and caminho_xlsx.stat().st_size > 0:
        caminho_xls.unlink()

    return caminho_xlsx

